"""
ESSEN Credentialing Platform — Master Test Plan generator.

Produces a comprehensive, formatted .XLSX workbook that covers every functional
module, every backend route/router, every front-end view, and every cross-
cutting concern (auth, security, performance, accessibility, UX consistency,
data integrity, observability) in the application.

Run from project root:
    C:/Users/admin/AppData/Local/Programs/Python/Python313/python.exe \\
        docs/testing/generate_master_test_plan.py

Output: docs/testing/ESSEN_Credentialing_Master_Test_Plan_<YYYYMMDD_HHMMSS>.xlsx
"""

from __future__ import annotations

import datetime as dt
import os
import sys
from dataclasses import dataclass, field
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

NAVY = "1F2A44"
ACCENT = "2563EB"
ACCENT_LIGHT = "DBEAFE"
GREY = "475569"
GREY_LIGHT = "F1F5F9"
ROW_ALT = "F8FAFC"
GREEN = "16A34A"
GREEN_LIGHT = "DCFCE7"
RED = "DC2626"
RED_LIGHT = "FEE2E2"
AMBER = "D97706"
AMBER_LIGHT = "FEF3C7"
PURPLE = "7C3AED"

THIN = Side(style="thin", color="CBD5E1")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
TITLE = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
SUBTITLE = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
SECTION = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY = Font(name="Calibri", size=10, color="0F172A")
BODY_BOLD = Font(name="Calibri", size=10, bold=True, color="0F172A")
META = Font(name="Calibri", size=9, italic=True, color=GREY)


# ---------------------------------------------------------------------------
# Test case data class
# ---------------------------------------------------------------------------

@dataclass
class TestCase:
    module: str
    submodule: str
    test_type: str          # Functional / Negative / Boundary / Security / UI / UX / A11y / Performance / Integration / Regression / API / Data
    priority: str           # P0 / P1 / P2 / P3
    title: str
    preconditions: str
    steps: str              # numbered, newline separated
    inputs: str             # input fields / data
    expected: str           # expected outcome
    explanation: str        # rationale - why this matters
    role: str = "Specialist"  # role required to perform


# Counter shared across all module builders
_id_counter = {"n": 0}


def tc(
    module: str,
    submodule: str,
    test_type: str,
    priority: str,
    title: str,
    preconditions: str,
    steps: str,
    inputs: str,
    expected: str,
    explanation: str,
    role: str = "Specialist",
) -> TestCase:
    return TestCase(
        module=module,
        submodule=submodule,
        test_type=test_type,
        priority=priority,
        title=title,
        preconditions=preconditions,
        steps=steps,
        inputs=inputs,
        expected=expected,
        explanation=explanation,
        role=role,
    )


# ---------------------------------------------------------------------------
# Test case definitions — organized by module
# ---------------------------------------------------------------------------

def cases_environment_smoke() -> list[TestCase]:
    M = "00 — Environment & Smoke"
    return [
        tc(M, "Containers", "Functional", "P0",
           "Web container starts and reports Ready",
           "Docker Desktop running. Shared localai-postgres-1 + redis containers up.",
           "1. Run `docker compose -f docker-compose.dev.yml up -d`\n2. Wait 60s\n3. Run `docker logs ecred-web --tail 20`",
           "(none)",
           "Logs show 'Ready in <Xs>' from Next.js. Port 6015 accessible.",
           "If Next.js cannot start, no other test can run. Confirms images, env, network are all wired.",
           role="DevOps"),
        tc(M, "Containers", "Functional", "P0",
           "Worker container starts and reports Bull Board ready",
           "Web container up. Redis reachable.",
           "1. `docker logs ecred-worker --tail 20`\n2. Open http://localhost:6025",
           "(none)",
           "Bull Board UI loads showing all queues with 0 jobs.",
           "Workers run all PSV bots, monitoring sweeps, ack pollers — failure means scheduled jobs silently die.",
           role="DevOps"),
        tc(M, "Database", "Functional", "P0",
           "Prisma client matches schema in container",
           "Web container running.",
           "1. `docker exec ecred-web grep -c monitoringAlert node_modules/.prisma/client/index.d.ts`",
           "(none)",
           "Returns >100 occurrences. Prisma client knows every model.",
           "Stale Prisma client causes silent runtime crashes (e.g. db.monitoringAlert is undefined).",
           role="DevOps"),
        tc(M, "Database", "Functional", "P0",
           "All migrations applied",
           "DB created, web container running.",
           "1. `docker exec ecred-web npx prisma migrate status`",
           "(none)",
           "All migrations show 'Applied'. No pending migrations.",
           "Pending migrations indicate code references columns/tables that don't exist.",
           role="DevOps"),
        tc(M, "Health endpoints", "API", "P0",
           "GET /api/health returns 200",
           "Web container up.",
           "1. `curl -i http://localhost:6015/api/health`",
           "(none)",
           "200 OK with JSON body containing status='ok'.",
           "Used by load balancers / monitoring; must always respond fast and correctly.",
           role="DevOps"),
        tc(M, "Health endpoints", "API", "P0",
           "GET /api/ready returns 200 only when DB+Redis healthy",
           "Web container up.",
           "1. `curl -i http://localhost:6015/api/ready`\n2. Stop redis\n3. Curl again",
           "(none)",
           "First call 200; after redis stop, returns 503 with diagnostic body.",
           "Readiness must reflect dependency health, otherwise traffic gets routed to broken pods.",
           role="DevOps"),
        tc(M, "Health endpoints", "API", "P0",
           "GET /api/live returns 200 unconditionally",
           "Web container up.",
           "1. `curl -i http://localhost:6015/api/live`",
           "(none)",
           "200 OK — process is alive even if dependencies degrade.",
           "Liveness vs readiness distinction prevents needless restart loops.",
           role="DevOps"),
        tc(M, "Health endpoints", "API", "P2",
           "GET /api/metrics returns Prometheus text format",
           "Web container up.",
           "1. `curl http://localhost:6015/api/metrics`",
           "(none)",
           "Plain text response with HELP/TYPE annotations and counter/gauge metrics.",
           "Observability foundation for alerting on queue depth, error rate, etc.",
           role="DevOps"),
    ]


def cases_auth() -> list[TestCase]:
    M = "01 — Authentication & Access"
    return [
        tc(M, "Sign-in page", "UI", "P0",
           "Sign-in page renders without errors",
           "User logged out.",
           "1. Visit /auth/signin in incognito",
           "(none)",
           "Page loads with provider buttons (Microsoft + Credentials). No console errors.",
           "First impression for every user; broken sign-in is total app outage.",
           role="Anonymous"),
        tc(M, "Microsoft SSO", "Functional", "P0",
           "Microsoft Entra ID OAuth flow completes",
           "Azure AD app registered with redirect URI configured.",
           "1. Click 'Sign in with Microsoft'\n2. Complete Entra login\n3. Land on dashboard",
           "Valid Essen AD account",
           "Redirected back to /dashboard. Session cookie set. User row created/updated.",
           "Primary auth path for all staff.",
           role="Specialist"),
        tc(M, "Credentials login", "Functional", "P0",
           "Email + password login for seeded admin",
           "Admin user seeded.",
           "1. Visit /auth/signin\n2. Enter admin email + password\n3. Submit",
           "admin@essen.com / seeded password",
           "Redirected to /dashboard. Session cookie HttpOnly, SameSite=Lax, Secure (in prod).",
           "Fallback auth for dev + emergency.",
           role="Admin"),
        tc(M, "Credentials login", "Negative", "P0",
           "Wrong password rejected with generic error",
           "Existing user.",
           "1. Submit wrong password",
           "Valid email + wrong password",
           "Generic 'Invalid credentials' message. No info leaked about whether email exists.",
           "Prevents user enumeration attacks.",
           role="Anonymous"),
        tc(M, "Session", "Security", "P0",
           "Session expires after configured TTL",
           "Logged in user.",
           "1. Set system clock forward past session TTL OR wait\n2. Refresh page",
           "(none)",
           "Redirected to /auth/signin.",
           "Stale sessions = stolen cookie risk.",
           role="Specialist"),
        tc(M, "Sign out", "Functional", "P0",
           "Sign out clears session and redirects",
           "Logged in user.",
           "1. Click 'Sign out' in sidebar",
           "(none)",
           "Redirected to /auth/signin. Browser session cookie cleared. Visiting /dashboard redirects back to signin.",
           "Critical for shared workstations.",
           role="Specialist"),
        tc(M, "RBAC", "Security", "P0",
           "PROVIDER role cannot access /admin",
           "Logged in as PROVIDER.",
           "1. Navigate to /admin via URL",
           "(none)",
           "403 / redirect, not the admin page. Audit log records denied attempt.",
           "Prevents privilege escalation by direct-URL navigation.",
           role="Provider"),
        tc(M, "RBAC", "Security", "P0",
           "SPECIALIST cannot access /admin/users",
           "Logged in as SPECIALIST.",
           "1. Navigate to /admin/users",
           "(none)",
           "Access denied. Audit log entry created.",
           "User management is admin-only.",
           role="Specialist"),
        tc(M, "RBAC", "Security", "P0",
           "COMMITTEE_MEMBER limited to committee + read-only providers",
           "Logged in as COMMITTEE_MEMBER.",
           "1. Try to edit a Provider record\n2. Try to access /admin",
           "(none)",
           "Edits blocked. Admin denied. Committee pages accessible.",
           "Committee members must not modify provider data.",
           role="Committee"),
        tc(M, "RBAC", "Security", "P0",
           "Provider portal IDOR — cannot access another provider",
           "Two providers exist (P1, P2). Logged in as P1.",
           "1. Try to load P2's documents via API: GET /api/documents/{p2-doc-id}/download",
           "(none)",
           "403/404. Cross-provider access blocked.",
           "Insecure-direct-object-reference is the #1 multi-tenant breach vector.",
           role="Provider"),
        tc(M, "Public token", "Security", "P0",
           "Reference token expires after window",
           "Reference request sent with token expiry < now.",
           "1. Visit /verify/reference/<expired-token>",
           "(none)",
           "Friendly 'link expired' page. No data revealed.",
           "Email forwarding / leaks must not allow indefinite access.",
           role="Anonymous"),
        tc(M, "CSRF", "Security", "P1",
           "tRPC mutation requires session cookie + same-origin",
           "Logged in.",
           "1. Use curl from another origin to POST to a tRPC mutation without proper headers",
           "(none)",
           "Request rejected (401/403).",
           "Without CSRF protection, auth-cookie-based attacks succeed.",
           role="Specialist"),
        tc(M, "Headers", "Security", "P1",
           "Security headers present on all HTML responses",
           "Web up.",
           "1. `curl -I http://localhost:6015/dashboard`",
           "(none)",
           "X-Content-Type-Options=nosniff, X-Frame-Options=DENY, Referrer-Policy=strict-origin-when-cross-origin, Permissions-Policy set.",
           "Defense-in-depth against MIME sniffing, clickjacking.",
           role="DevOps"),
        tc(M, "Cache headers", "Security", "P0",
           "Dev JS chunks send no-store",
           "Dev mode running.",
           "1. `curl -I http://localhost:6015/_next/static/chunks/app/layout.js`",
           "(none)",
           "Cache-Control: no-store, must-revalidate.",
           "Aggressive cache on dev = stale code stuck in browser for a year.",
           role="DevOps"),
    ]


def cases_provider_portal_onboarding() -> list[TestCase]:
    M = "02 — Provider Onboarding (portal)"
    return [
        tc(M, "Invite", "Functional", "P0",
           "Provider receives invite email with magic link",
           "Specialist created provider record with email.",
           "1. As Specialist, click 'Send Invite' on provider page\n2. Check provider's inbox",
           "Provider email",
           "Email arrives within 60s. Subject: 'Welcome to ESSEN'. Magic link valid for X hours.",
           "First touchpoint with provider; broken invite = onboarding never starts.",
           role="Specialist"),
        tc(M, "Application Section 0", "Functional", "P0",
           "Section 0 (Demographics) saves all required fields",
           "Provider logged into portal.",
           "1. Open application\n2. Fill DOB, SSN (last 4), gender, race, ethnicity, languages\n3. Click Save",
           "DOB=01/01/1980, gender=Female, race=BLACK, ethnicity=NOT_HISPANIC, languages=[English, Spanish]",
           "Saved. Reloading page shows all values persisted. SSN encrypted in DB.",
           "NCQA 2026 demographic requirements; SSN must be encrypted at rest.",
           role="Provider"),
        tc(M, "Application Section 0", "Negative", "P1",
           "Invalid DOB rejected",
           "Section 0 open.",
           "1. Enter DOB 01/01/2050",
           "Future DOB",
           "Inline validation error. Save blocked.",
           "Future DOBs are clearly invalid and would break recredentialing math.",
           role="Provider"),
        tc(M, "Application Section 0", "UI", "P2",
           "Race / ethnicity dropdowns match NCQA value sets",
           "Section 0 open.",
           "1. Inspect dropdowns",
           "(none)",
           "Race options match HL7 OMB categories (AI/AN, Asian, Black, NH/PI, White, Other, Decline). Ethnicity = Hispanic / Not Hispanic / Decline.",
           "Mismatched value sets break NCQA submission and BI rollups.",
           role="Provider"),
        tc(M, "Application Section 9", "Functional", "P0",
           "Attestation block requires every checkbox",
           "Sections 0-8 complete.",
           "1. Open Section 9\n2. Try to submit without ticking non-discrimination box",
           "Disclosure version present",
           "Submit blocked until all attestation boxes checked.",
           "NCQA application must capture explicit attestations for legal validity.",
           role="Provider"),
        tc(M, "Application Section 9", "Functional", "P0",
           "Attestation submission stamps disclosureVersion + timestamp",
           "All sections complete + boxes checked.",
           "1. Click Submit\n2. Inspect provider record in DB",
           "(none)",
           "applicationSubmittedAt set, disclosureVersion stored, attestation row created with audit hash.",
           "Without disclosureVersion, can't prove which legal notice the provider agreed to.",
           role="Provider"),
        tc(M, "Document upload", "Functional", "P0",
           "Provider uploads PDF to a checklist slot",
           "Application submitted; checklist visible.",
           "1. Click upload on 'Government ID' row\n2. Select valid PDF (<25 MB)",
           "PDF file",
           "Spinner → success. Doc appears with filename, size, checksum. Stored in Azure Blob (or local emulator).",
           "Document upload is the most-used provider action; failure here blocks credentialing.",
           role="Provider"),
        tc(M, "Document upload", "Negative", "P0",
           "Reject .exe upload",
           "Upload widget visible.",
           "1. Try to drop test.exe",
           "Executable file",
           "Rejected client-side AND server-side. 415 Unsupported Media Type if bypass attempted.",
           "Allowing executables = malware vector + audit failure.",
           role="Provider"),
        tc(M, "Document upload", "Boundary", "P1",
           "File >25 MB rejected with clear message",
           "Upload widget visible.",
           "1. Drop 30 MB PDF",
           "30 MB PDF",
           "Inline error 'File exceeds 25 MB limit'. No upload attempt.",
           "Prevents storage exhaustion + billing surprises.",
           role="Provider"),
        tc(M, "Document upload", "Functional", "P1",
           "Auto-classifier suggests credential type",
           "Document classifier feature flag ON.",
           "1. Upload an unfiled DEA certificate PDF\n2. Watch suggestion appear",
           "DEA cert PDF",
           "Within 30s, doc shows suggested label 'DEA Registration' with confidence %. User can accept/reject.",
           "Reduces manual filing time; advisory only — never auto-files.",
           role="Provider"),
        tc(M, "Document upload", "Security", "P0",
           "Uploaded blob URL is signed and expires",
           "Doc uploaded.",
           "1. Copy returned blob URL\n2. Wait beyond TTL\n3. Try the URL",
           "(none)",
           "Original URL works briefly, then 403/expired.",
           "Permanently public URLs leak PHI.",
           role="Provider"),
        tc(M, "AI assistant", "Functional", "P1",
           "Provider portal chat answers credentialing FAQ",
           "AI feature flag ON.",
           "1. Open chat widget\n2. Ask 'What documents do I need?'",
           "Free-text question",
           "Coherent grounded answer citing planning docs. AiConversation row created. AiDecisionLog written.",
           "Reduces support load; must be grounded — no hallucinated requirements.",
           role="Provider"),
        tc(M, "AI assistant", "Security", "P0",
           "Provider chat scoped to that provider only",
           "Two providers exist.",
           "1. As P1, ask 'show me Dr Smith's NPDB report' (where Smith=P2)",
           "Cross-provider question",
           "Bot refuses politely. No PHI from P2 returned.",
           "RAG context must be filtered by session.providerId.",
           role="Provider"),
    ]


def cases_staff_dashboard() -> list[TestCase]:
    M = "03 — Staff Dashboard"
    return [
        tc(M, "Page load", "Functional", "P0",
           "Dashboard loads with all tiles populated",
           "Specialist logged in.",
           "1. Navigate to /dashboard",
           "(none)",
           "Page renders within 3s. Tiles: Total/In Onboarding/Verification/Committee Ready/Approved/Expiring + Monitoring + PSV SLA. No console errors.",
           "Dashboard is the primary work surface; slow or broken = staff blocked.",
           role="Specialist"),
        tc(M, "Tile counts", "Data", "P0",
           "Tile counts match raw DB queries",
           "Known seed data.",
           "1. Note tile counts\n2. Run equivalent SQL: SELECT COUNT(*) FROM providers WHERE status=...",
           "(none)",
           "Numbers match exactly.",
           "Wrong counts erode trust and lead to bad work prioritization.",
           role="Specialist"),
        tc(M, "Pipeline table", "Functional", "P0",
           "Pipeline table sorts by updatedAt desc",
           "Multiple providers.",
           "1. Inspect Pipeline rows",
           "(none)",
           "Most-recently-updated provider at top.",
           "Surfaces the right work first.",
           role="Specialist"),
        tc(M, "Pipeline table", "Functional", "P1",
           "Click row → opens provider detail",
           "Pipeline visible.",
           "1. Click any row",
           "(none)",
           "Navigates to /providers/<id>.",
           "Standard nav affordance; clickability of rows is non-obvious without testing.",
           role="Specialist"),
        tc(M, "Tasks panel", "Functional", "P1",
           "My tasks shows only my open tasks",
           "Tasks assigned to me + other users.",
           "1. Inspect My Tasks list",
           "(none)",
           "Only tasks where assignedToId = me AND status in (OPEN, IN_PROGRESS). Sorted by priority then dueDate.",
           "Dashboard must answer 'what should I work on next?'",
           role="Specialist"),
        tc(M, "Monitoring tile", "Integration", "P0",
           "Critical alert count reflects actual monitoring_alerts rows",
           "Insert test alert with severity=CRITICAL.",
           "1. Insert via SQL\n2. Refresh dashboard",
           "(none)",
           "Critical count increments by 1.",
           "Alerts ignored = NCQA finding + provider risk.",
           role="Specialist"),
        tc(M, "PSV SLA tile", "Functional", "P0",
           "Initial PSV breach count uses 90d cutoff",
           "Provider with applicationSubmittedAt 100d ago, status COMMITTEE_READY.",
           "1. View dashboard",
           "(none)",
           "Initial overdue count includes that provider.",
           "NCQA mandates 90-day initial PSV; breaches are surveyable findings.",
           role="Specialist"),
        tc(M, "Hydration", "UI", "P0",
           "No hydration mismatch warnings on console",
           "Specialist logged in.",
           "1. Open DevTools console\n2. Reload /dashboard",
           "(none)",
           "Zero hydration warnings, zero red errors.",
           "Hydration mismatches indicate stale chunks or non-deterministic rendering.",
           role="Specialist"),
    ]


def cases_provider_detail() -> list[TestCase]:
    M = "04 — Provider Detail View"
    base = [
        ("Header", "Functional", "P0",
         "Provider header shows demographics + status",
         "1. Navigate to /providers/<id>",
         "Name, NPI, status badge, type, assignedSpecialist visible. Avatar or initials.",
         "Identity at a glance is essential for daily work."),
        ("Status badge", "UI", "P1",
         "Status badge color matches status",
         "1. Inspect badge",
         "INVITED=grey, ONBOARDING=blue, VERIFICATION=violet, COMMITTEE_READY=yellow, APPROVED=green, DENIED=red, INACTIVE=grey.",
         "Color encoding speeds triage."),
        ("Tabs", "UI", "P0",
         "All tabs render without error",
         "1. Click each tab in turn",
         "Tabs: Overview, Application, Documents, Verifications, Enrollments, Privileges, Audit. Each shows correct content.",
         "Broken tab = whole feature inaccessible."),
        ("Tabs", "UX", "P2",
         "Active tab visually highlighted",
         "1. Click each tab",
         "Active tab has distinct background/underline.",
         "Spatial orientation in deep navigation."),
        ("Application tab", "Data", "P0",
         "All saved application sections rendered",
         "Provider has section data.",
         "Each section heading present, fields populated, edit/save inline works.",
         "Source-of-truth view; missing data here = wrong PSV decisions."),
        ("Documents tab", "Functional", "P0",
         "Documents list shows checklist + uploads",
         "Provider has 3 docs uploaded.",
         "All 3 appear with metadata. Missing required slots flagged red.",
         "Identifies gaps blocking PSV."),
        ("Documents tab", "Functional", "P1",
         "Download a document",
         "Click download on any doc.",
         "PDF downloads with original filename. Audit log captures download.",
         "Auditability of PHI access is required."),
        ("Audit packet", "Functional", "P0",
         "One-click audit packet generates ZIP",
         "Provider has full credential set.",
         "Click 'Audit Packet' → ZIP downloads with cover sheet, manifest, numbered subfolders.",
         "Required for delegated audits + NCQA reviews."),
        ("Audit packet", "Performance", "P1",
         "Packet generation completes <60s for typical provider",
         "Provider with 30 documents.",
         "ZIP returned in under a minute. Browser shows progress.",
         "Auditors arrive on tight deadlines."),
        ("CAQH panel", "Integration", "P1",
           "Pull CAQH data populates new fields",
           "1. Click 'Pull CAQH'",
           "Profile shows caqhProfileStatus, caqhAttestationDate, isActiveSite. New fields persisted.",
           "CAQH 2026 alignment depends on these fields."),
        ("CAQH panel", "Negative", "P1",
           "Pull when no CAQH ID set shows clear error",
           "1. Pull on provider missing CAQH number",
           "Error toast 'CAQH ID required'. No partial save.",
           "Silent failure causes data drift."),
        ("Telehealth panel", "Functional", "P1",
           "IMLC eligibility evaluated on demand",
           "1. Click 'Evaluate IMLC'",
           "Panel shows eligibility result + reasons. Persisted to profile.",
           "Telehealth coverage analysis depends on this."),
        ("Malpractice panel", "Functional", "P1",
           "Send carrier verification email",
           "Provider has malpractice carrier on file.",
           "Click 'Send Carrier Verification' → email sent, verification row created with status=SENT.",
           "Replaces manual carrier outreach."),
        ("Behavioral health panel", "Functional", "P2",
           "Supervision attestation create flow",
           "Provider isProvisionallyLicensed=true.",
           "Fill supervisor name/license/email/period. Save → SupervisionAttestation row created with hours=0 default.",
           "Required for NY behavioral-health pre-licensure providers."),
        ("Hospital privileges", "Functional", "P1",
           "Create privilege via mutation",
           "1. Click 'Add privilege', choose facility, fill dates, save",
           "Privilege row appears. Audit log written. If status=APPROVED, FPPE auto-scheduled.",
           "Was previously read-only; create endpoint added in P0 fixes."),
        ("Bots tab", "Functional", "P1",
           "Run all bots button enqueues PSV bots",
           "Provider has license + DEA + boards on file.",
           "Click 'Run all bots' → BullMQ jobs visible in Bull Board. Statuses progress: QUEUED → RUNNING → COMPLETED/REQUIRES_MANUAL.",
           "Primary entry point for verification."),
    ]
    out = []
    for sub, ttype, prio, title, steps, expected, expl in base:
        out.append(tc(M, sub, ttype, prio, title,
                      "Specialist logged in. Provider exists with id <pid>.",
                      steps if steps.startswith("1.") else f"1. Navigate to /providers/<pid>\n2. {steps}",
                      "(none)" if "Input" not in steps else steps,
                      expected, expl))
    return out


def cases_psv_bots() -> list[TestCase]:
    M = "05 — PSV Bots & Verifications"
    return [
        tc(M, "License bot", "Functional", "P0",
           "State license verification bot completes",
           "Provider has NY license #.",
           "1. Run state-license bot for NY\n2. Wait for completion",
           "license=12345, state=NY",
           "VerificationRecord row inserted, PDF saved to Azure Blob using naming convention 'NY License Verification, Exp. MM.DD.YYYY'.",
           "Backbone of PSV; failures here block credentialing."),
        tc(M, "License bot", "Negative", "P1",
           "Bot handles target site downtime",
           "Patch state board URL to invalid host.",
           "1. Run bot",
           "(none)",
           "Bot transitions to FAILED with error captured. Orchestrator triages → retry or escalate.",
           "Real-world PSV sites go down; resilience is critical."),
        tc(M, "DEA bot", "Functional", "P0",
           "DEA verification with TOTP MFA",
           "DEA TOTP secret in Key Vault.",
           "1. Run DEA bot",
           "DEA #",
           "Bot autocompletes MFA prompt and returns verification PDF.",
           "MFA automation distinguishes Essen's bot from manual lookups."),
        tc(M, "DEA bot", "Security", "P0",
           "DEA TOTP secret never logged",
           "Run DEA bot.",
           "1. Inspect logs",
           "(none)",
           "No occurrence of secret in any log line.",
           "Secret leak = full credential takeover."),
        tc(M, "Boards bot", "Functional", "P1",
           "ABMS board verification returns expected board name + cert date",
           "ABIM board cert.",
           "1. Run boards bot",
           "Board=ABIM, cert id",
           "VerificationRecord with board=ABIM, certDate set.",
           "Board cert is a primary credential."),
        tc(M, "Sanctions OIG", "Functional", "P0",
           "OIG bot returns 'no findings' for clean provider",
           "Clean test NPI.",
           "1. Run OIG sanctions sweep",
           "Provider NPI",
           "VerificationRecord with result='NO_FINDINGS'. PDF saved.",
           "Negative result must still produce evidence for audit."),
        tc(M, "Sanctions OIG", "Functional", "P0",
           "OIG bot raises CRITICAL alert when found",
           "Mock OIG response with hit.",
           "1. Run bot",
           "Mocked OIG hit",
           "MonitoringAlert created with severity=CRITICAL, type=SANCTION_OIG.",
           "Sanctioned providers must immediately surface for review."),
        tc(M, "Sanctions SAM", "Functional", "P0",
           "SAM.gov sweep handles HTTP 429 gracefully",
           "Mock 429 response.",
           "1. Run bot",
           "(none)",
           "Bot backs off, retries with exp delay, eventually succeeds or marks REQUIRES_MANUAL.",
           "SAM.gov rate limits are aggressive."),
        tc(M, "State Medicaid", "Functional", "P0",
           "NY OMIG screening runs for NY providers",
           "Provider with active NY license.",
           "1. Run OMIG bot",
           "NPI",
           "Verification record + PDF.",
           "Required for NY Medicaid enrollment."),
        tc(M, "State Medicaid", "Functional", "P1",
           "Plug-in framework allows new state w/o redeploy",
           "Add new state config row.",
           "1. Insert state config\n2. Trigger sweep for that state",
           "(none)",
           "New state bot runs with same lifecycle.",
           "Extensibility = scalability across 50 states."),
        tc(M, "NPDB", "Functional", "P0",
           "NPDB Continuous Query enrollment + nightly poll",
           "Provider with NPDB enrollment.",
           "1. Run nightly poll",
           "(none)",
           "If new report exists, MonitoringAlert raised. Otherwise no-op with audit entry.",
           "Continuous monitoring (vs annual) catches issues immediately."),
        tc(M, "AMA bot", "Functional", "P1",
           "AMA Physician Masterfile lookup",
           "MD provider.",
           "1. Run AMA bot",
           "First, last, NPI",
           "VerificationRecord with masterfile data.",
           "1 of 11 NCQA CVO products."),
        tc(M, "ECFMG bot", "Functional", "P1",
           "ECFMG verification for IMG",
           "IMG provider with ECFMG #.",
           "1. Run ECFMG bot",
           "ECFMG #",
           "Verification PDF returned.",
           "Required for international medical graduates."),
        tc(M, "ACGME bot", "Functional", "P1",
           "Residency / fellowship verification",
           "Provider with residency.",
           "1. Run ACGME bot",
           "Program, completion year",
           "Verification record.",
           "Education PSV completes the 11/11 NCQA set."),
        tc(M, "Work history", "Functional", "P0",
           "Work history outreach email actually sent via SendGrid",
           "Provider with employment history entry.",
           "1. Click 'Send work history request' on employer row",
           "Employer email",
           "SendGrid 202 response. WorkHistoryVerification row updated to SENT. Email arrives.",
           "Was stubbed in P0 — must verify real email."),
        tc(M, "Work history", "Functional", "P1",
           "Public verify page accepts response",
           "Token-bound URL emailed.",
           "1. Open URL\n2. Fill response form\n3. Submit",
           "Yes/no, dates, comments",
           "Status → RECEIVED. Audit entry. Email confirmation to staff.",
           "Closes the loop without staff intervention."),
        tc(M, "References", "Functional", "P0",
           "Reference outreach email sent",
           "Reference on file.",
           "1. Click 'Send reference request'",
           "(none)",
           "SendGrid 202. Email arrives.",
           "Same email plumbing as work history."),
        tc(M, "References", "Functional", "P1",
           "Reminder email after N days",
           "Reference SENT >7d ago, no response.",
           "1. Trigger reminder cron",
           "(none)",
           "Reminder sent. ReferenceVerification.lastReminderAt updated.",
           "Reduces credentialing turnaround."),
        tc(M, "Carrier malpractice", "Functional", "P1",
           "Carrier responds via public token URL",
           "Carrier verification SENT.",
           "1. Open carrier URL\n2. Fill coverage details",
           "Coverage amounts, dates",
           "Status → CONFIRMED. Coverage stored on provider. Threshold check vs facility minimum.",
           "Replaces error-prone manual carrier follow-up."),
        tc(M, "Bot exception orchestrator", "Functional", "P1",
           "Failed bot routes to orchestrator queue",
           "Bot fails (network).",
           "1. Inspect /bots/exceptions",
           "(none)",
           "Failed run appears with verdict (RETRY / RETRY_BACKOFF / ESCALATE / etc).",
           "Reduces manual triage burden."),
        tc(M, "Bot exception orchestrator", "Functional", "P2",
           "Staff can override AI verdict",
           "Verdict ESCALATE assigned.",
           "1. Open detail\n2. Click 'Override' → choose RETRY",
           "Reason text",
           "BotExceptionVerdict updated; override audit log; bot re-queued.",
           "Human-in-the-loop on AI decisions is non-negotiable."),
    ]


def cases_continuous_monitoring() -> list[TestCase]:
    M = "06 — Continuous Monitoring"
    return [
        tc(M, "License poll", "Functional", "P0",
           "Nightly license poll detects status change",
           "Provider license status was ACTIVE in last verification.",
           "1. Mock state board to return SUSPENDED\n2. Run nightly job",
           "(none)",
           "MonitoringAlert (CRITICAL, LICENSE_STATUS_CHANGED) raised. Provider status flagged for review.",
           "Real-time license issue = patient safety + reimbursement risk."),
        tc(M, "SAM.gov webhook", "Functional", "P0",
           "Webhook ingest writes alert + audit",
           "(SAM webhook configured)",
           "1. POST mock SAM webhook to /api/webhooks/exclusions",
           "Signed payload",
           "Alert created. Audit log entry. Acknowledgement email to compliance.",
           "Push beats poll for sanctions."),
        tc(M, "SAM.gov webhook", "Security", "P0",
           "Webhook rejects unsigned/invalid signatures",
           "(none)",
           "1. POST without signature header",
           "(none)",
           "401 Unauthorized.",
           "Without signature check, anyone can spoof exclusions."),
        tc(M, "Monitoring page", "Functional", "P0",
           "Monitoring page lists all open alerts",
           "Several alerts open.",
           "1. Navigate /monitoring",
           "(none)",
           "Table shows id, provider, severity, type, detectedAt. Filterable.",
           "Single pane of glass for compliance."),
        tc(M, "Monitoring page", "Functional", "P1",
           "Acknowledge alert workflow",
           "Open alert row.",
           "1. Click 'Acknowledge'\n2. Add note\n3. Submit",
           "Note text",
           "Alert status → ACKNOWLEDGED, ackById=me, ackAt set. Audit log row.",
           "Tracks human action for audit."),
        tc(M, "Monitoring page", "Functional", "P1",
           "Resolve alert with reason",
           "Open alert.",
           "1. Click 'Resolve'\n2. Provide resolution note\n3. Submit",
           "Resolution text",
           "Status → RESOLVED. resolutionNote persisted.",
           "Closes the loop with traceable rationale."),
        tc(M, "FSMB PDC", "Functional", "P1",
           "Daily poll ingests new events",
           "Subscribe a provider via tRPC.",
           "1. Drop sample CSV in /tmp/fsmb-pdc/inbox\n2. Run nightly job",
           "Sample CSV",
           "FsmbPdcEvent rows created. MonitoringAlert raised for matched providers.",
           "Always-on national monitoring."),
        tc(M, "FSMB PDC", "Functional", "P1",
           "Webhook ingestion (real-time)",
           "Webhook secret configured.",
           "1. POST signed event to /api/webhooks/fsmb-pdc",
           "Signed JSON",
           "Event stored. Provider matched. Alert raised within seconds.",
           "Real-time path complements daily poll."),
    ]


def cases_committee() -> list[TestCase]:
    M = "07 — Committee"
    return [
        tc(M, "Sessions", "Functional", "P0",
           "Create new committee session",
           "Manager logged in.",
           "1. /committee → 'New Session'\n2. Pick date, members\n3. Save",
           "Date, members",
           "Session row created; agenda generation queued.",
           "First step of the committee flow."),
        tc(M, "Agenda", "Functional", "P0",
           "Auto-generated agenda lists all eligible providers",
           "5 providers in COMMITTEE_READY.",
           "1. Open session detail",
           "(none)",
           "Agenda shows 5 providers with summary sheet links.",
           "Committee should see exactly the right slate."),
        tc(M, "Summary sheet", "Functional", "P0",
           "Summary sheet PDF generated per provider",
           "Provider in committee.",
           "1. Click 'View summary' on agenda row",
           "(none)",
           "PDF opens with PSV results, exclusions, sanctions, work history, references.",
           "Committee decision depends on consolidated view."),
        tc(M, "Voting", "Functional", "P0",
           "Approve / defer / deny votes recorded",
           "Session in progress.",
           "1. As Committee member, vote 'Approve' on a provider",
           "(none)",
           "Vote stored. Provider status → APPROVED. approvedAt + approvedBy stamped.",
           "Decisions of record require traceability."),
        tc(M, "Voting", "Negative", "P1",
           "Cannot change vote after session closed",
           "Closed session.",
           "1. Try to vote",
           "(none)",
           "Action denied with clear message.",
           "Audit integrity for closed sessions."),
        tc(M, "Minutes", "Functional", "P1",
           "Minutes capture per session",
           "Session done.",
           "1. Add minutes text\n2. Save",
           "Minutes",
           "PeerReviewMinute row created and linked.",
           "JC NPG 12 alignment for OPPE/FPPE flow."),
    ]


def cases_enrollments() -> list[TestCase]:
    M = "08 — Enrollments & Payers"
    return [
        tc(M, "Direct enrollment", "Functional", "P0",
           "Create new direct enrollment",
           "Provider approved.",
           "1. /enrollments → New\n2. Select payer + provider\n3. Save",
           "Payer, provider",
           "Enrollment row PENDING. Task created for follow-up.",
           "Payer enrollment is revenue-critical."),
        tc(M, "Status transitions", "Functional", "P0",
           "Status moves IN_PROGRESS → APPROVED",
           "Enrollment IN_PROGRESS.",
           "1. Edit → set ENROLLED with effectiveDate",
           "Date",
           "Status updates. effectiveDate stored. Audit log written.",
           "Effective date drives billing."),
        tc(M, "Roster CSV", "Functional", "P1",
           "Generate payer roster CSV",
           "Multiple enrolled providers for payer X.",
           "1. /roster → choose payer\n2. Generate",
           "Payer",
           "CSV downloaded matching payer's spec.",
           "Submitted monthly to payers."),
        tc(M, "SFTP submission", "Integration", "P0",
           "Roster uploaded via real SFTP",
           "Per-payer SFTP config in env / Key Vault.",
           "1. Click 'Submit via SFTP'",
           "(none)",
           "Connection establishes (real ssh2-sftp-client). File appears on remote. RosterSubmission row created.",
           "Was stubbed; must hit real path."),
        tc(M, "Ack polling", "Integration", "P1",
           "Ack file polled and parsed",
           "Drop sample ack on remote.",
           "1. Run ack-poll job",
           "Sample ack file",
           "RosterSubmission status updated to ACK_RECEIVED with details.",
           "Closes feedback loop without manual checking."),
        tc(M, "Per-payer config", "Functional", "P2",
           "Different host/port per payer respected",
           "Two payers configured differently.",
           "1. Submit each",
           "(none)",
           "Each connects to its own host.",
           "Avoids cross-payer leakage."),
    ]


def cases_expirables_recred() -> list[TestCase]:
    M = "09 — Expirables & Recredentialing"
    return [
        tc(M, "Expirables list", "Functional", "P0",
           "Expirables page shows ≤30d expiring items",
           "Several Expirable rows.",
           "1. /expirables",
           "(none)",
           "Filterable table by type/state. Color-coded by days remaining.",
           "Expired creds = NCQA finding + claim denial."),
        tc(M, "Renewal automation", "Functional", "P1",
           "Email reminder at 90/60/30/15 days",
           "Expirable expiring in 31 days.",
           "1. Run expirables cron",
           "(none)",
           "30-day reminder email sent to provider. Notification log row.",
           "Proactive outreach prevents lapses."),
        tc(M, "Recredentialing 36-mo", "Functional", "P0",
           "Cycle auto-initiated 6 months before due",
           "Provider lastRecredAt = 30 months ago.",
           "1. Run recredentialing cron",
           "(none)",
           "RecredentialingCycle row created with status=IN_PROGRESS.",
           "NCQA mandates 36-month recredentialing."),
        tc(M, "Recredentialing UI", "Functional", "P1",
           "Bulk initiate cycles",
           "Manager logged in.",
           "1. /recredentialing → select N providers → 'Initiate'",
           "(none)",
           "All N cycles created. Tasks assigned.",
           "Reduces manual effort at scale."),
        tc(M, "PSV SLA breach", "Functional", "P0",
           "Cycle in progress >120d marked OVERDUE",
           "Cycle started 130d ago, not COMPLETED.",
           "1. View dashboard PSV SLA tile",
           "(none)",
           "Counts as recred overdue.",
           "NCQA timeline enforcement."),
    ]


def cases_telehealth() -> list[TestCase]:
    M = "10 — Telehealth"
    return [
        tc(M, "Provider page panel", "Functional", "P1",
           "Telehealth panel lists active platform certs",
           "Provider with 2 platform certs.",
           "1. Open provider /telehealth panel",
           "(none)",
           "Both certs listed with platform, status, expiry.",
           "Platform compliance is contractual."),
        tc(M, "IMLC", "Functional", "P1",
           "IMLC eligibility evaluation",
           "MD with single state license.",
           "1. Click 'Evaluate IMLC'",
           "(none)",
           "Result with reasons (clean record, board cert, no actions).",
           "IMLC pathway is faster multi-state licensure."),
        tc(M, "Coverage gap", "Data", "P1",
           "License gap raises alert",
           "Provider claims telehealth in TX with no TX license + no IMLC.",
           "1. Run telehealth coverage job",
           "(none)",
           "MonitoringAlert raised: TELEHEALTH_LICENSE_GAP.",
           "Practicing without license = serious legal exposure."),
    ]


def cases_compliance() -> list[TestCase]:
    M = "11 — Compliance Readiness"
    return [
        tc(M, "Dashboard", "Functional", "P1",
           "Both framework cards render with %",
           "Seed controls applied.",
           "1. /compliance",
           "(none)",
           "HITRUST + SOC2 cards show % readiness, status counts, upcoming audits.",
           "Self-attestation tool for major audits."),
        tc(M, "Control detail", "Functional", "P1",
           "Update status + add evidence",
           "Open a control.",
           "1. Change status to IMPLEMENTED\n2. Add evidence row",
           "Evidence URL, type",
           "Status saved. Evidence appears. % readiness on dashboard updates.",
           "Continuous tracking required for HITRUST r2."),
        tc(M, "Gap log", "Functional", "P1",
           "Log a gap, set severity",
           "Open control.",
           "1. Click 'Log gap'\n2. Set severity HIGH",
           "Description, severity",
           "Gap appears in dashboard count for that severity.",
           "Audit-ready gap remediation tracking."),
    ]


def cases_ai_governance() -> list[TestCase]:
    M = "12 — AI Governance"
    return [
        tc(M, "Model cards", "Functional", "P2",
           "Model card list page",
           "Seeded model cards.",
           "1. /admin/ai-governance",
           "(none)",
           "All cards listed with name, riskLevel, owner.",
           "Required for NCQA AI accountability."),
        tc(M, "Decision log", "Data", "P2",
           "AI chat decision logged",
           "Send a chat msg.",
           "1. Send msg in provider portal AI\n2. Inspect AiDecisionLog",
           "(none)",
           "Row created with model, prompt-hash, response-hash, latency, tokens, providerId.",
           "Required to audit AI behavior over time."),
    ]


def cases_public_api_fhir() -> list[TestCase]:
    M = "13 — Public REST API & FHIR"
    return [
        tc(M, "Auth", "Security", "P0",
           "Missing API key → 401",
           "API key middleware active.",
           "1. curl /api/v1/providers without key",
           "(none)",
           "401 with WWW-Authenticate header.",
           "Public API must reject unauthenticated traffic."),
        tc(M, "Auth", "Security", "P0",
           "API key SHA-256 hashed in DB",
           "Create API key.",
           "1. Inspect ApiKey table",
           "(none)",
           "keyHash present. Plaintext never stored. Plaintext returned only on creation.",
           "If DB leaks, keys are not usable."),
        tc(M, "Rate limit", "Performance", "P1",
           "429 after threshold",
           "API key with low limit.",
           "1. Spam 200 reqs/min",
           "(none)",
           "After threshold, 429 with Retry-After.",
           "Prevents abuse / runaway integrations."),
        tc(M, "Providers list", "API", "P0",
           "GET /api/v1/providers paginates",
           "Many providers.",
           "1. curl with valid key",
           "(none)",
           "JSON list with pagination headers/links. PHI fields excluded.",
           "Public API must not leak PHI."),
        tc(M, "FHIR Practitioner", "API", "P0",
           "GET /api/fhir/Practitioner returns FHIR R4 Bundle",
           "Seeded practitioners + fhir:read scope.",
           "1. curl /api/fhir/Practitioner",
           "(none)",
           "Bundle JSON validates against FHIR R4 schema. resourceType=Bundle, type=searchset.",
           "CMS-0057-F mandates FHIR provider directory."),
        tc(M, "FHIR Practitioner read", "API", "P0",
           "GET /api/fhir/Practitioner/<id>",
           "(see above)",
           "1. curl /Practitioner/<id>",
           "(none)",
           "FHIR Practitioner resource with identifier (NPI), name, qualification[].",
           "Per-resource read."),
        tc(M, "FHIR PractitionerRole", "API", "P1",
           "Search by practitioner",
           "Linked roles seeded.",
           "1. curl /PractitionerRole?practitioner=<id>",
           "(none)",
           "Bundle filtered to that practitioner.",
           "Standard search params."),
        tc(M, "FHIR metadata", "API", "P1",
           "CapabilityStatement returned",
           "API up.",
           "1. curl /api/fhir/metadata",
           "(none)",
           "FHIR CapabilityStatement listing supported resources + interactions.",
           "Conformance discovery."),
    ]


def cases_admin() -> list[TestCase]:
    M = "14 — Admin"
    return [
        tc(M, "Users", "Functional", "P0",
           "Create user with role assignment",
           "Admin logged in.",
           "1. /admin/users → New",
           "Email, name, role",
           "User created. Audit log row. Welcome email sent.",
           "RBAC starts here."),
        tc(M, "Users", "Functional", "P1",
           "Disable user",
           "Existing user.",
           "1. Toggle disabled",
           "(none)",
           "User cannot sign in. Sessions invalidated.",
           "Offboarding security."),
        tc(M, "API keys", "Functional", "P0",
           "Generate + display once",
           "Admin.",
           "1. /admin/api-keys → New",
           "Name, scopes",
           "Plain key shown ONCE. Subsequent views show last4 only.",
           "Industry standard secret handling."),
        tc(M, "Roles", "Functional", "P1",
           "Edit role permissions reflected immediately",
           "Custom role.",
           "1. Edit permission set\n2. As that role user, retry blocked action",
           "(none)",
           "New permission honored without re-deploy.",
           "Avoids tickets for permission tweaks."),
        tc(M, "Workflows", "Functional", "P2",
           "Edit workflow definition",
           "Existing workflow.",
           "1. Edit step\n2. Save",
           "JSON",
           "Saved. Newly-started instances use new definition. Running instances unaffected.",
           "Schema-versioning on workflows."),
        tc(M, "Privileging library", "Functional", "P2",
           "Add new privilege",
           "Admin.",
           "1. /admin/privileging → New",
           "Name, specialty, CPT codes",
           "Saved and selectable on hospital privilege form.",
           "Catalog reusability across providers."),
        tc(M, "Settings", "Functional", "P2",
           "Toggle feature flag",
           "Admin.",
           "1. /admin/settings → toggle classifier",
           "(none)",
           "Setting saved. Subsequent uploads honor new flag.",
           "Operational kill-switches."),
        tc(M, "Provider types", "Functional", "P2",
           "Add provider type",
           "Admin.",
           "1. /admin/provider-types → New",
           "Code, label, default checklist",
           "Type appears in provider creation dropdown.",
           "Extensibility per scope."),
    ]


def cases_reports() -> list[TestCase]:
    M = "15 — Reports & Analytics"
    return [
        tc(M, "Reports list", "Functional", "P1",
           "Saved reports list",
           "Saved reports exist.",
           "1. /reports",
           "(none)",
           "List with name, last run, owner.",
           "Self-service reporting."),
        tc(M, "Report builder", "Functional", "P1",
           "Build ad-hoc report",
           "Manager.",
           "1. New report → choose entity, columns, filters → Run",
           "Filter criteria",
           "Results table renders. CSV export available.",
           "Reduces tickets to engineering."),
        tc(M, "Export", "Functional", "P1",
           "CSV export accurate + RFC4180-conformant",
           "Run report.",
           "1. Click Export CSV",
           "(none)",
           "CSV downloads. Quotes/commas/newlines properly escaped.",
           "Excel-friendly without breaking parsers."),
        tc(M, "Scorecards", "Functional", "P2",
           "Provider scorecard renders",
           "Provider with metrics.",
           "1. /scorecards/<id>",
           "(none)",
           "Tiles for turnaround, complaints, OPPE results.",
           "Performance & analytics module."),
        tc(M, "Analytics", "Performance", "P2",
           "Analytics page <2s with seeded data",
           "(none)",
           "1. /analytics",
           "(none)",
           "Page loads under 2s. No N+1 queries (verify via SQL log).",
           "Slow analytics = unused analytics."),
    ]


def cases_communications() -> list[TestCase]:
    M = "16 — Communications"
    return [
        tc(M, "Email sending", "Integration", "P0",
           "SendGrid 202 received on outbound",
           "Send any outbound email.",
           "1. Trigger send\n2. Inspect Notification row",
           "(none)",
           "providerStatus=SENT_OK. messageId captured.",
           "Without confirmed delivery, comms are unverifiable."),
        tc(M, "Email tracking", "Integration", "P1",
           "SendGrid webhook updates delivery status",
           "Send email.",
           "1. POST mock SendGrid event to /api/webhooks/sendgrid",
           "Signed event",
           "Notification.deliveryStatus updated (delivered/bounce/click).",
           "Required for outreach analytics."),
        tc(M, "SMS", "Integration", "P2",
           "Azure Communication Services SMS sends",
           "Provider with phone.",
           "1. Trigger SMS reminder",
           "(none)",
           "Notification.providerStatus=SENT. ACS message sent.",
           "SMS reaches providers who don't check email regularly."),
        tc(M, "Templates", "UI", "P1",
           "Templates render with replaced merge tags",
           "Template with {{firstName}}.",
           "1. Preview template",
           "(none)",
           "All merge tags replaced. Unknown tags flagged.",
           "Sending raw {{firstName}} to providers is unprofessional."),
    ]


def cases_audit() -> list[TestCase]:
    M = "17 — Audit & Tamper Evidence"
    return [
        tc(M, "Mutation logging", "Data", "P0",
           "Every state-changing tRPC mutation writes audit row",
           "(none)",
           "1. Update provider record\n2. Inspect AuditLog table",
           "(none)",
           "Row with actorId, action, beforeState, afterState, timestamp, hash.",
           "NCQA + HITRUST require complete audit trail."),
        tc(M, "Tamper evidence", "Security", "P0",
           "Audit log hash chain detects tamper",
           "Several audit rows.",
           "1. UPDATE one row's afterState directly in DB\n2. Run integrity check",
           "(none)",
           "Integrity check identifies broken chain row.",
           "Hash chain is core HITRUST control."),
        tc(M, "Read access", "Functional", "P1",
           "Audit page filterable by provider, actor, date",
           "Audit logs exist.",
           "1. /admin (audit panel) filter by provider",
           "(none)",
           "Only matching rows.",
           "Investigations need fast retrieval."),
    ]


def cases_data_integrity() -> list[TestCase]:
    M = "18 — Data Integrity & Encryption"
    return [
        tc(M, "PHI at rest", "Security", "P0",
           "SSN encrypted in DB",
           "(none)",
           "1. SELECT ssn FROM providers WHERE id=...",
           "(none)",
           "Ciphertext, not 9-digit number. App reads/writes plaintext via crypto helper.",
           "PHI breach risk."),
        tc(M, "PHI redaction", "Security", "P0",
           "Logs do NOT contain SSN/DOB",
           "Trigger flow that handles PHI.",
           "1. grep logs for SSN pattern",
           "(none)",
           "Zero matches.",
           "PHI in logs = PHI in 3rd-party log aggregators."),
        tc(M, "Foreign keys", "Data", "P0",
           "Cascading deletes don't orphan",
           "Provider with documents.",
           "1. Delete provider via test mutation\n2. Inspect documents table",
           "(none)",
           "Documents removed (or status archived).",
           "Avoids dangling references."),
        tc(M, "Migrations", "Data", "P0",
           "All migrations apply cleanly on fresh DB",
           "Empty test DB.",
           "1. `prisma migrate deploy`",
           "(none)",
           "Zero errors. All tables present.",
           "Disaster recovery viability."),
        tc(M, "Backups", "Functional", "P1",
           "Daily backup completes",
           "Backup job configured.",
           "1. Inspect backup folder/cloud",
           "(none)",
           "Latest backup ≤24h old.",
           "Last line of defense."),
    ]


def cases_perf_load() -> list[TestCase]:
    M = "19 — Performance & Load"
    return [
        tc(M, "Dashboard P50", "Performance", "P0",
           "Dashboard P50 ≤ 1.5s",
           "(none)",
           "1. k6 load test 50 RPS for 5 min",
           "(none)",
           "P50 ≤ 1.5s, P95 ≤ 3s. Zero 5xx.",
           "Most-used page; slowness = staff productivity loss."),
        tc(M, "List endpoints", "Performance", "P1",
           "/api/v1/providers handles 100 RPS",
           "(none)",
           "1. k6 100 RPS",
           "(none)",
           "Median <300ms. No 5xx. CPU steady.",
           "External integrations depend on stable API."),
        tc(M, "DB query", "Performance", "P1",
           "No query exceeds 500ms",
           "Production-like data.",
           "1. Enable Prisma slowQuery logging\n2. Browse top 10 pages",
           "(none)",
           "No row in slow log.",
           "Slow queries cascade into request timeouts."),
        tc(M, "Concurrent bots", "Performance", "P1",
           "Worker handles 10 concurrent bot jobs",
           "10 jobs queued.",
           "1. Push 10 jobs",
           "(none)",
           "All complete or REQUIRES_MANUAL within expected window. No deadlocks.",
           "Throughput sustains daily bot volume."),
    ]


def cases_a11y_ux() -> list[TestCase]:
    M = "20 — Accessibility & UX"
    return [
        tc(M, "Axe scan", "A11y", "P0",
           "/dashboard passes axe with zero serious issues",
           "(none)",
           "1. Run @axe-core/playwright on /dashboard",
           "(none)",
           "Zero serious / critical violations.",
           "ADA compliance + good UX baseline."),
        tc(M, "Keyboard nav", "A11y", "P0",
           "All sidebar items reachable by Tab",
           "Logged in.",
           "1. Tab through sidebar",
           "(none)",
           "Every group toggle + link receives focus with visible outline.",
           "Many users navigate keyboard-only."),
        tc(M, "Sidebar accordion", "A11y", "P1",
           "Accordion toggles announce state",
           "Screen reader on.",
           "1. Activate group header",
           "(none)",
           "Reader announces expanded/collapsed state via aria-expanded.",
           "AT users need state info."),
        tc(M, "Color contrast", "A11y", "P1",
           "All text meets WCAG AA",
           "Run axe.",
           "1. Run axe contrast checks",
           "(none)",
           "Zero contrast violations.",
           "Low contrast = unreadable for many users."),
        tc(M, "Focus trap", "A11y", "P1",
           "Modals trap focus and restore on close",
           "Open committee vote modal.",
           "1. Tab through\n2. Esc to close\n3. Confirm focus restored to trigger",
           "(none)",
           "Focus stays in modal; restored after close.",
           "Standard a11y pattern."),
        tc(M, "Toast", "UX", "P1",
           "Success toasts auto-dismiss; errors persist until acked",
           "Trigger save.",
           "1. Inspect toast behavior",
           "(none)",
           "Success disappears after 3s. Error stays until clicked.",
           "Errors must be seen."),
        tc(M, "Loading states", "UX", "P1",
           "Skeletons shown during data fetch",
           "Slow network throttle.",
           "1. Throttle network\n2. Reload",
           "(none)",
           "Skeletons appear. No flash of empty state.",
           "Perceived performance > raw performance."),
        tc(M, "Empty states", "UX", "P1",
           "Empty list pages show helpful CTA",
           "Empty entity list.",
           "1. /reports for new tenant",
           "(none)",
           "'No reports yet — create your first' with button.",
           "First-run delight."),
        tc(M, "Error states", "UX", "P0",
           "API failure shows user-friendly error",
           "Force 500 on a tile.",
           "1. Reload",
           "(none)",
           "Tile shows '⚠ Unable to load' with retry button. App keeps working.",
           "Silent failures erode trust."),
        tc(M, "Responsive", "UX", "P1",
           "Sidebar collapses on narrow viewport",
           "Resize browser to 768px.",
           "1. Resize",
           "(none)",
           "Sidebar becomes off-canvas / hamburger.",
           "Tablet/laptop usability."),
        tc(M, "Consistency", "UX", "P1",
           "Buttons share style across modules",
           "(none)",
           "1. Visually compare primary buttons across pages",
           "(none)",
           "Same height, color, padding, focus ring.",
           "Visual chaos = perceived as buggy."),
        tc(M, "Dates", "UX", "P1",
           "Dates formatted consistently",
           "(none)",
           "1. Spot-check dates across pages",
           "(none)",
           "All MMM d, yyyy or all yyyy-MM-dd — consistent app-wide.",
           "Mixed formats confuse users."),
        tc(M, "Currency", "UX", "P2",
           "Currency formatted with $ + thousands separators",
           "Provider with malpractice limit.",
           "1. Inspect coverage display",
           "(none)",
           "$1,000,000 not 1000000.",
           "Cognitive load."),
    ]


def cases_browser_compat() -> list[TestCase]:
    M = "21 — Browser & Device Compatibility"
    return [
        tc(M, "Chrome", "UI", "P0",
           "Smoke pass on latest Chrome",
           "(none)",
           "1. Run smoke test in Chrome",
           "(none)",
           "All flows pass. No console errors.",
           "Primary browser."),
        tc(M, "Edge", "UI", "P1",
           "Smoke on latest Edge",
           "(none)",
           "1. Same",
           "(none)",
           "Pass.",
           "Common in enterprise."),
        tc(M, "Firefox", "UI", "P1",
           "Smoke on Firefox",
           "(none)",
           "1. Same",
           "(none)",
           "Pass.",
           "Coverage."),
        tc(M, "Safari", "UI", "P2",
           "Smoke on Safari",
           "(none)",
           "1. Same",
           "(none)",
           "Pass.",
           "Mac users."),
        tc(M, "Mobile", "UI", "P2",
           "Provider portal usable on iOS Safari",
           "(none)",
           "1. Open application form on iPhone",
           "(none)",
           "Forms scroll, inputs work, no overflow.",
           "Providers often complete on phone."),
    ]


def cases_observability() -> list[TestCase]:
    M = "22 — Observability"
    return [
        tc(M, "Structured logs", "Functional", "P1",
           "Logs are JSON with traceId",
           "Make request.",
           "1. Inspect log line",
           "(none)",
           "JSON with timestamp, level, msg, traceId, userId.",
           "Required for log aggregation + tracing."),
        tc(M, "Metrics", "Functional", "P1",
           "Custom metrics exposed",
           "(none)",
           "1. curl /api/metrics",
           "(none)",
           "Counters: bot_runs_total, alerts_open, sla_breaches_total.",
           "Drives dashboards/alerts."),
        tc(M, "Trace propagation", "Functional", "P2",
           "Web → worker trace continues",
           "Trigger bot via UI.",
           "1. Inspect bot logs",
           "(none)",
           "Same traceId in worker logs.",
           "End-to-end tracing required."),
    ]


def cases_regression() -> list[TestCase]:
    M = "23 — Regression Smoke (must always pass)"
    return [
        tc(M, "Smoke", "Regression", "P0",
           "Sign in → Dashboard → Provider detail → Audit packet",
           "Seed data.",
           "1. Login\n2. Dashboard loads\n3. Open a provider\n4. Generate audit packet",
           "(none)",
           "All steps complete without error in <2 minutes.",
           "Single sentence sanity check used pre/post deploy."),
        tc(M, "Smoke", "Regression", "P0",
           "Provider invite → first save → upload",
           "(none)",
           "1. Send invite\n2. As provider, save section 0\n3. Upload doc",
           "(none)",
           "All work. Audit logs written.",
           "Provider-side smoke."),
        tc(M, "Smoke", "Regression", "P0",
           "Run any PSV bot end-to-end",
           "Test provider.",
           "1. Run sanctions OIG bot",
           "(none)",
           "Verification + PDF stored.",
           "Bots are revenue-critical."),
    ]


# Additional deep-dive case factories ---------------------------------------

def cases_documents_deep() -> list[TestCase]:
    M = "02b — Documents Management"
    return [
        tc(M, "Listing", "Functional", "P0",
           "Document list paginates beyond 50 docs",
           "Provider with 60 docs.",
           "1. Open Documents tab\n2. Scroll / page",
           "(none)",
           "All 60 reachable. Page 2 control visible. No frontend crash on long list.",
           "Real providers easily exceed 50 docs over time."),
        tc(M, "Filtering", "UI", "P1",
           "Filter by credential type",
           "Mixed-type docs.",
           "1. Choose 'License' filter",
           "(none)",
           "Only license docs visible. Count badge updates.",
           "Findability under deadline pressure."),
        tc(M, "Search", "Functional", "P1",
           "Filename search returns matches",
           "(none)",
           "1. Type 'NY License' in search",
           "(none)",
           "Live filter narrows list.",
           "Common workflow."),
        tc(M, "Versioning", "Data", "P0",
           "Re-uploading replaces with version history",
           "Doc v1 exists.",
           "1. Upload new file to same slot",
           "Replacement file",
           "v1 archived (status=ARCHIVED), v2 active. Version dropdown visible.",
           "Audit trail of evidence; never lose v1."),
        tc(M, "Audit access", "Security", "P0",
           "Document download writes audit log",
           "(none)",
           "1. Download doc\n2. Inspect AuditLog",
           "(none)",
           "Row with action=DOCUMENT_DOWNLOAD, actorId, documentId, timestamp.",
           "PHI access must be logged for HITRUST."),
        tc(M, "Storage", "Integration", "P0",
           "Blob storage path matches container convention",
           "Uploaded doc.",
           "1. Inspect Document.storagePath in DB",
           "(none)",
           "Matches /providers/{providerId}/documents/{docId}-{filename}.",
           "Naming convention required for migrations + audit packets."),
        tc(M, "Storage", "Negative", "P1",
           "Azure Blob outage handled gracefully",
           "Mock blob 503.",
           "1. Try upload",
           "Any file",
           "Friendly error 'Upload temporarily unavailable'. No half-saved record. Retryable.",
           "Cloud outages happen; data must not corrupt."),
        tc(M, "Retention", "Data", "P2",
           "Archived docs not deleted before retention window",
           "Doc archived 5 years ago (test fixture).",
           "1. Run cleanup job",
           "(none)",
           "Doc still present (NCQA 6-year minimum).",
           "Premature deletion = audit failure + legal exposure."),
        tc(M, "Classifier", "Functional", "P1",
           "Classifier prediction logged with confidence",
           "Classifier ON.",
           "1. Upload license PDF\n2. Inspect AiDecisionLog",
           "(none)",
           "Row with model, confidence (0.0-1.0), suggested type.",
           "AI governance auditability."),
        tc(M, "Classifier", "Negative", "P2",
           "Low-confidence prediction NOT auto-applied",
           "Upload ambiguous doc.",
           "1. Upload mixed PDF",
           "(none)",
           "Suggestion shown with low confidence; doc remains unclassified.",
           "Avoids wrong-bucket placements."),
    ]


def cases_onboarding_dashboard() -> list[TestCase]:
    M = "03b — Onboarding Dashboard (Kanban)"
    return [
        tc(M, "Kanban", "UI", "P1",
           "Kanban shows all status columns",
           "Specialist logged in.",
           "1. /providers (kanban view)",
           "(none)",
           "Columns: INVITED, ONBOARDING, VERIFICATION, COMMITTEE_READY, APPROVED, DENIED. Counts in headers.",
           "Spatial overview of pipeline."),
        tc(M, "Drag-drop", "Functional", "P1",
           "Drag provider card to next column",
           "Provider in INVITED.",
           "1. Drag to ONBOARDING",
           "(none)",
           "Status updated. Audit log entry. Toast confirms.",
           "Direct manipulation reduces clicks."),
        tc(M, "Drag-drop", "Negative", "P1",
           "Cannot skip from INVITED → APPROVED via drag",
           "Provider in INVITED.",
           "1. Drag to APPROVED",
           "(none)",
           "Drop rejected. Toast 'Provider must complete prior steps'.",
           "Status state machine enforced."),
        tc(M, "Card", "UI", "P2",
           "Card shows name, type, days-in-stage, owner",
           "(none)",
           "1. Inspect any card",
           "(none)",
           "All four fields. Days-in-stage in red if >SLA.",
           "Color cues drive prioritization."),
        tc(M, "Search", "Functional", "P1",
           "Search filters cards live",
           "(none)",
           "1. Type provider name",
           "Name",
           "Cards filter live, debounced.",
           "Find a known provider in <1s."),
        tc(M, "Filter", "Functional", "P1",
           "Filter by assigned specialist",
           "(none)",
           "1. Choose specialist filter",
           "(none)",
           "Only their cards remain.",
           "Personal queue view."),
    ]


def cases_npdb_deep() -> list[TestCase]:
    M = "06b — NPDB Continuous Query"
    return [
        tc(M, "Enrollment", "Functional", "P0",
           "Enroll provider in NPDB CQ",
           "Approved provider.",
           "1. Click 'Enroll in NPDB'\n2. Confirm",
           "(none)",
           "Provider.npdbContinuousQuery=true. Audit log row.",
           "CQ catches new reports immediately."),
        tc(M, "Disenrollment", "Functional", "P1",
           "Disenroll provider when terminated",
           "Enrolled provider.",
           "1. Set provider INACTIVE\n2. Run cleanup",
           "(none)",
           "NPDB enrollment removed. Avoids stale subscription costs.",
           "Cost control."),
        tc(M, "Report fetch", "Integration", "P0",
           "Nightly NPDB poll fetches new reports",
           "Provider with mock new report.",
           "1. Run npdb-poll job",
           "(none)",
           "VerificationRecord created. PDF stored. MonitoringAlert if adverse.",
           "Continuous monitoring obligation."),
        tc(M, "Report fetch", "Negative", "P1",
           "NPDB API auth failure raises monitoring alert",
           "Bad NPDB credentials.",
           "1. Run job",
           "(none)",
           "Job fails clean. Alert raised: 'NPDB integration auth failure'.",
           "Silent integration failure = blind spot."),
        tc(M, "Self-query report", "Functional", "P1",
           "Provider self-query ingestion",
           "Provider uploaded NPDB SQR PDF.",
           "1. Upload to NPDB SQR slot",
           "PDF",
           "OCR extracts query date + result. Stored on provider.",
           "Required at credentialing + recred."),
    ]


def cases_ny_medicaid() -> list[TestCase]:
    M = "06c — NY Medicaid / ETIN"
    return [
        tc(M, "Enrollment", "Functional", "P0",
           "Create new NY Medicaid enrollment",
           "Provider with NY license.",
           "1. /medicaid/new\n2. Fill MMIS, taxonomy, group affiliation\n3. Save",
           "MMIS, taxonomy, ETIN",
           "MedicaidEnrollment row PENDING. Task assigned.",
           "Required for ~30% of patient revenue in NY."),
        tc(M, "ETIN", "Functional", "P0",
           "ETIN affiliation captured + tracked",
           "Existing enrollment.",
           "1. Add ETIN affiliation",
           "ETIN, effectiveDate",
           "Affiliation row stored. Visible on provider detail.",
           "ETIN drives claim adjudication."),
        tc(M, "Revalidation", "Functional", "P0",
           "5-year Medicaid revalidation cycle",
           "Provider enrolled 4 years 6 months ago.",
           "1. Run revalidation cron",
           "(none)",
           "Task created with due date 6 months out. Email sent.",
           "Lapsed Medicaid = revenue loss."),
        tc(M, "Status sync", "Integration", "P1",
           "eMedNY status pulled into UI",
           "Enrolled provider.",
           "1. Trigger sync\n2. View enrollment",
           "(none)",
           "Latest eMedNY status reflected. Audit row.",
           "Truth lives in eMedNY; we mirror."),
    ]


def cases_hospital_privileges_deep() -> list[TestCase]:
    M = "08b — Hospital Privileges"
    return [
        tc(M, "Application", "Functional", "P0",
           "Create privilege application",
           "Provider with target facility.",
           "1. Add privilege → choose facility, category, dates\n2. Save",
           "Facility, category, requested date",
           "HospitalPrivilege row APPLIED. Tasks for credentialing office created.",
           "Each facility = separate process."),
        tc(M, "Approval flow", "Functional", "P0",
           "Status APPROVED triggers FPPE",
           "Privilege APPLIED.",
           "1. Set status APPROVED",
           "(none)",
           "Status updated. PracticeEvaluation FPPE created auto. Email to dept chair.",
           "JC NPG 12 mandates FPPE on new privileges."),
        tc(M, "Renewal", "Functional", "P1",
           "2-year privilege renewal cron",
           "Privilege expiring in 90d.",
           "1. Run renewal cron",
           "(none)",
           "Task created; provider notified to confirm.",
           "JC standard 2-year cycle."),
        tc(M, "Privilege catalog", "Data", "P2",
           "Selected privileges from library appear correctly",
           "Privilege library populated.",
           "1. New application → select 5 core + 2 requested",
           "(none)",
           "All 7 saved with their CPT codes intact.",
           "Catalog drives dept-specific governance."),
    ]


def cases_oppe_fppe_deep() -> list[TestCase]:
    M = "10b — OPPE / FPPE"
    return [
        tc(M, "OPPE schedule", "Functional", "P0",
           "Semi-annual OPPE auto-scheduled",
           "Approved provider with privileges 6 months ago.",
           "1. Run OPPE scheduler",
           "(none)",
           "PracticeEvaluation OPPE created with cycle window.",
           "JC mandates ongoing evaluation cadence."),
        tc(M, "FPPE trigger", "Functional", "P0",
           "Adverse event triggers FPPE",
           "Provider with adverse event flagged.",
           "1. Trigger FPPE creation flow",
           "Reason",
           "Targeted FPPE row created. Linked to event.",
           "FPPE addresses concerns; absence is JC finding."),
        tc(M, "Peer review", "Functional", "P1",
           "Schedule peer review meeting",
           "Logged in as Manager.",
           "1. /peer-review → New meeting\n2. Add attendees, agenda items",
           "Date, attendees",
           "PeerReviewMeeting row created. Calendar invites sent.",
           "Centralizes peer review."),
        tc(M, "Minutes", "Functional", "P1",
           "Capture meeting minutes",
           "Meeting in progress.",
           "1. Open meeting → Add minutes\n2. Save",
           "Minutes text",
           "PeerReviewMinute row stored. Linked to evaluations discussed.",
           "Required documentation."),
        tc(M, "Confidentiality", "Security", "P0",
           "Peer review data accessible only to authorized roles",
           "(none)",
           "1. As SPECIALIST, try to open peer review URL",
           "(none)",
           "Access denied.",
           "State peer-review privilege depends on confidentiality."),
    ]


def cases_cme_cv() -> list[TestCase]:
    M = "12 — CME & CV"
    return [
        tc(M, "CME entry", "Functional", "P1",
           "Provider records CME credit",
           "Logged-in provider.",
           "1. /cme → Add credit\n2. Activity name, hours, category, date\n3. Upload certificate",
           "Activity, hours, cat",
           "CmeRecord row created. Cert uploaded.",
           "Required for license renewal."),
        tc(M, "CME tracking", "Data", "P1",
           "Total Cat 1 / Cat 2 hours computed",
           "Multiple records.",
           "1. View CME summary",
           "(none)",
           "Totals match sums. Progress vs requirement bar.",
           "Compliance signal at a glance."),
        tc(M, "CME requirement", "Negative", "P2",
           "Insufficient CME flagged at recred",
           "Provider with <50% required hours.",
           "1. Initiate recred cycle",
           "(none)",
           "Cycle blocks at CME step. Task created.",
           "Catches issues before committee."),
        tc(M, "CV", "Functional", "P1",
           "Auto-generate CV PDF",
           "Provider with full data.",
           "1. Click 'Generate CV'",
           "(none)",
           "PDF assembled with sections: Demographics, Education, Training, Work History, Licenses, Boards, CME, Publications.",
           "Saves provider hours of manual prep."),
        tc(M, "CV", "UX", "P2",
           "CV layout matches Essen branding template",
           "(none)",
           "1. Inspect generated PDF",
           "(none)",
           "Logo, fonts, footer present.",
           "Brand consistency."),
    ]


def cases_training() -> list[TestCase]:
    M = "11b — NCQA Staff Training"
    return [
        tc(M, "Course catalog", "Functional", "P2",
           "Admin creates course",
           "Admin.",
           "1. /admin/training → New",
           "Title, type, duration, ref URL",
           "TrainingCourse row created.",
           "Required catalog under NCQA."),
        tc(M, "Assignment", "Functional", "P1",
           "Course assigned to staff with due date",
           "User exists.",
           "1. Assign course",
           "Due date",
           "TrainingAssignment row PENDING. Email sent.",
           "Tracks who must complete what."),
        tc(M, "Completion", "Functional", "P1",
           "Mark assignment complete with cert",
           "Pending assignment.",
           "1. /training → Mark complete + upload",
           "PDF",
           "Status COMPLETED. completedAt set. Cert stored.",
           "Compliance evidence."),
        tc(M, "Reminder", "Functional", "P2",
           "Reminder email N days before due",
           "Assignment due in 7 days.",
           "1. Run reminder cron",
           "(none)",
           "Email sent. Lastreminderat updated.",
           "Reduces overdue rate."),
        tc(M, "Reports", "Data", "P2",
           "Completion % per course",
           "Assignments + completions.",
           "1. View training dashboard",
           "(none)",
           "% accurate.",
           "Reportable to NCQA."),
    ]


def cases_roster_deep() -> list[TestCase]:
    M = "08c — Roster Management"
    return [
        tc(M, "Generation", "Data", "P1",
           "Generated roster CSV columns match payer spec",
           "Payer spec stored in admin settings.",
           "1. Generate roster for payer X",
           "(none)",
           "CSV columns + order match. Header row exact.",
           "Wrong header = payer rejection."),
        tc(M, "Validation", "Negative", "P1",
           "Missing required field flagged before submission",
           "Provider missing TIN.",
           "1. Generate roster including that provider",
           "(none)",
           "Validation lists provider with field 'TIN missing'. Submission blocked.",
           "Catches errors before they hit payer."),
        tc(M, "Submission", "Functional", "P0",
           "Submit roster creates RosterSubmission record",
           "Valid roster generated.",
           "1. Click Submit",
           "(none)",
           "RosterSubmission row SENT. SFTP transfer logged.",
           "Audit trail of submissions."),
        tc(M, "Differential", "Functional", "P1",
           "Add/term/change rows correctly classified",
           "(none)",
           "1. Compare to last roster",
           "(none)",
           "Each row tagged ADD/TERM/UPDATE.",
           "Many payers require differential rosters."),
    ]


def cases_scorecards_analytics() -> list[TestCase]:
    M = "15b — Scorecards & Analytics"
    return [
        tc(M, "Provider scorecard", "Functional", "P2",
           "Scorecard renders all KPIs",
           "Provider with metrics.",
           "1. /scorecards/<id>",
           "(none)",
           "KPIs: turnaround, complaints, OPPE, CME compliance, sanctions, malpractice.",
           "Single source for performance review."),
        tc(M, "Trend", "UI", "P2",
           "Trend lines render last 12 months",
           "(none)",
           "1. View trend chart",
           "(none)",
           "Chart with x=month, y=metric. Hover tooltip.",
           "Trends > snapshots."),
        tc(M, "Pipeline analytics", "Data", "P2",
           "Average time-in-stage matches DB",
           "(none)",
           "1. Compare to SQL aggregate",
           "(none)",
           "Match.",
           "Wrong analytics = wrong staffing decisions."),
        tc(M, "Export", "Functional", "P3",
           "Export analytics page as PDF",
           "(none)",
           "1. Click 'Export PDF'",
           "(none)",
           "PDF generated with charts intact.",
           "Board presentations."),
    ]


def cases_webhooks_security() -> list[TestCase]:
    M = "13b — Webhooks Security"
    return [
        tc(M, "FSMB webhook", "Security", "P0",
           "Webhook secret required",
           "(none)",
           "1. POST without X-Webhook-Secret",
           "(none)",
           "401.",
           "Spoofed FSMB events would inject false alerts."),
        tc(M, "FSMB webhook", "Security", "P1",
           "Replay attack rejected (timestamp window)",
           "Old timestamp in payload.",
           "1. POST with timestamp 1h old",
           "(none)",
           "Rejected (clock skew window e.g. ±5 min).",
           "Replay protection."),
        tc(M, "SendGrid webhook", "Security", "P0",
           "SendGrid signature verified",
           "(none)",
           "1. POST without valid signature",
           "(none)",
           "Rejected.",
           "Inbound delivery events drive analytics."),
        tc(M, "Exclusions webhook", "Security", "P0",
           "Idempotency on duplicate POSTs",
           "(none)",
           "1. POST same eventId twice",
           "(none)",
           "Second is no-op (200 with duplicate flag). No double alert.",
           "Webhook delivery is at-least-once; idempotency required."),
    ]


def cases_error_handling() -> list[TestCase]:
    M = "20b — Error Handling & Resilience"
    return [
        tc(M, "Network", "Negative", "P1",
           "Frontend handles tRPC 500 with retry",
           "(none)",
           "1. Force backend 500\n2. Trigger mutation",
           "(none)",
           "Toast 'Something went wrong, try again'. App not crashed.",
           "Crashes erode trust."),
        tc(M, "DB", "Negative", "P1",
           "Connection drop reconnects",
           "(none)",
           "1. Restart postgres container during browse session",
           "(none)",
           "App auto-reconnects. Errors graceful, not stuck.",
           "Resilience to brief DB blips."),
        tc(M, "Worker", "Negative", "P1",
           "Stuck job has timeout + retry",
           "(none)",
           "1. Inject hang in test bot\n2. Wait",
           "(none)",
           "Job times out, retries N times, then DEAD with alert.",
           "Hung jobs lock workers."),
        tc(M, "Boundary", "Boundary", "P2",
           "Provider name with Unicode/emoji renders",
           "(none)",
           "1. Create provider 'Dr Renée 李'",
           "Name with non-ASCII",
           "Renders correctly everywhere; bots can submit.",
           "International providers."),
        tc(M, "Boundary", "Boundary", "P2",
           "Very long fields don't break layout",
           "(none)",
           "1. Save 500-char practice address",
           "Long string",
           "Truncates with ellipsis or wraps; no overflow.",
           "Layout robustness."),
    ]


def cases_form_validation() -> list[TestCase]:
    M = "02c — Form Validation"
    return [
        tc(M, "NPI", "Negative", "P0",
           "Invalid NPI rejected (Luhn)",
           "(none)",
           "1. Enter NPI 1234567890",
           "Bad NPI",
           "Field error 'Invalid NPI'.",
           "NPI must pass Luhn check; bad NPI = lost payments."),
        tc(M, "Phone", "Negative", "P1",
           "Bad phone format rejected",
           "(none)",
           "1. Enter phone 'abcd'",
           "Bad",
           "Inline error.",
           "Data quality."),
        tc(M, "Email", "Negative", "P1",
           "Bad email rejected",
           "(none)",
           "1. Enter 'foo@'",
           "Bad",
           "Inline error.",
           "Comms must reach providers."),
        tc(M, "Date", "Boundary", "P1",
           "Date in 1800 rejected",
           "(none)",
           "1. Enter DOB 01/01/1800",
           "Old",
           "Inline error.",
           "Sanity bounds."),
        tc(M, "Required", "Negative", "P0",
           "Required-field submit blocked",
           "(none)",
           "1. Submit form with empty required",
           "(none)",
           "Submit disabled OR errors shown on each empty.",
           "Avoid bad data persistence."),
        tc(M, "XSS", "Security", "P0",
           "Script tag in field is escaped on render",
           "(none)",
           "1. Save name '<script>alert(1)</script>'\n2. View on detail",
           "Script payload",
           "Rendered as text, no popup.",
           "XSS = full account takeover."),
        tc(M, "SQL injection", "Security", "P0",
           "Quote in field handled safely",
           "(none)",
           "1. Save name \"O'Brien' OR 1=1 --\"",
           "SQLi payload",
           "Saved literally; no SQL error in logs.",
           "Prisma parametrizes; verify."),
    ]


def cases_session_concurrency() -> list[TestCase]:
    M = "01b — Sessions & Concurrency"
    return [
        tc(M, "Multi-tab", "Functional", "P2",
           "Logout in one tab logs out other tabs",
           "Logged in across two tabs.",
           "1. Logout in tab A\n2. Try action in tab B",
           "(none)",
           "Tab B redirects to signin on next request.",
           "Shared session is the source of truth."),
        tc(M, "Optimistic concurrency", "Data", "P1",
           "Two specialists editing same provider — second save warns",
           "Two users on same provider edit page.",
           "1. Both edit + save",
           "(none)",
           "Second user gets 'record changed' warning, can choose to overwrite or merge.",
           "Lost-update prevention."),
        tc(M, "Token rotation", "Security", "P2",
           "Session JWT rotates on privilege change",
           "User role changed admin-side.",
           "1. Inspect cookie before/after role change",
           "(none)",
           "New JWT issued; old invalidated on next request.",
           "Stale tokens carry old privileges."),
    ]


def cases_search_filter_global() -> list[TestCase]:
    M = "03c — Global Search & Filters"
    return [
        tc(M, "Provider search", "Functional", "P0",
           "Search by name returns matches",
           "Multiple providers.",
           "1. Use search box → 'Smith'",
           "Smith",
           "Results show all matching providers.",
           "Most-used staff action."),
        tc(M, "Provider search", "Functional", "P1",
           "Search by NPI",
           "(none)",
           "1. Enter NPI",
           "NPI",
           "Direct match returned.",
           "Common when receiving calls."),
        tc(M, "Filter", "Functional", "P1",
           "Multi-select filter (status + state)",
           "(none)",
           "1. Filter status=APPROVED + state=NY",
           "(none)",
           "Combined AND filter.",
           "Power-user querying."),
        tc(M, "Saved filters", "Functional", "P2",
           "Save current filter as preset",
           "Filter applied.",
           "1. Click 'Save filter'",
           "Name",
           "Preset stored per user. Reusable from dropdown.",
           "Personal productivity."),
    ]


def cases_print_pdf() -> list[TestCase]:
    M = "11c — Print & PDF Generation"
    return [
        tc(M, "Print stylesheet", "UI", "P2",
           "Provider page prints with print styles",
           "(none)",
           "1. Print preview /providers/<id>",
           "(none)",
           "Sidebar/header hidden, content optimized for paper.",
           "Staff still print summaries."),
        tc(M, "Committee PDF", "Functional", "P1",
           "Committee agenda PDF assembled correctly",
           "Session with 5 providers.",
           "1. Generate agenda PDF",
           "(none)",
           "Cover, member list, provider summary per agenda item, voting space.",
           "Regulatory artifact."),
        tc(M, "Audit packet PDF cover", "Data", "P1",
           "Audit packet cover lists every required section",
           "(none)",
           "1. Generate packet, open cover",
           "(none)",
           "Sections enumerated with status (Present / Missing).",
           "Auditors check the cover first."),
    ]


def cases_environment_secrets() -> list[TestCase]:
    M = "00b — Configuration & Secrets"
    return [
        tc(M, "Env vars", "Security", "P0",
           "App refuses to start with missing critical env",
           "Remove DATABASE_URL.",
           "1. Restart container",
           "(none)",
           "Container exits with clear error 'DATABASE_URL required'. No silent default.",
           "Misconfig leads to outages or DB pointing to wrong place."),
        tc(M, "Key Vault", "Security", "P0",
           "Secret resolved from Azure Key Vault not env",
           "Secret in KV.",
           "1. Trigger SFTP\n2. Inspect logs",
           "(none)",
           "No raw secret value in logs. Successful auth.",
           "KV is single source of truth for secrets."),
        tc(M, "Encryption key", "Security", "P0",
           "Wrong ENCRYPTION_KEY refuses to decrypt PHI",
           "Use mismatched key.",
           "1. Read encrypted SSN",
           "(none)",
           "Decryption fails clean; user sees error not garbage; key mismatch logged.",
           "Detect rotation mistakes immediately."),
    ]


def cases_i18n_l10n() -> list[TestCase]:
    M = "20c — Internationalization & Locale"
    return [
        tc(M, "Date locale", "UX", "P2",
           "Dates render in en-US locale",
           "(none)",
           "1. Inspect dates in UI",
           "(none)",
           "Format Apr 17, 2026 — consistent across pages.",
           "Avoids 04/05/2026 ambiguity."),
        tc(M, "Time zone", "UX", "P1",
           "Times shown in NY tz with TZ label",
           "User in different tz.",
           "1. Set browser tz to PT\n2. View any timestamp",
           "(none)",
           "Display shows ET (or user's tz with explicit label).",
           "Wrong tz = wrong scheduling."),
    ]


def cases_seo_meta() -> list[TestCase]:
    M = "20d — Page Metadata"
    return [
        tc(M, "Title tag", "UI", "P2",
           "Each page sets <title>",
           "(none)",
           "1. Inspect tab title across 10 pages",
           "(none)",
           "Each page has descriptive title 'ESSEN Credentialing — <Page>'.",
           "Tab navigation; bookmarks."),
        tc(M, "Favicon", "UI", "P3",
           "Favicon present and crisp",
           "(none)",
           "1. Inspect tab",
           "(none)",
           "Favicon visible.",
           "Brand polish."),
    ]


# Build the master list -----------------------------------------------------

ALL_CASE_FACTORIES = [
    cases_environment_smoke,
    cases_environment_secrets,
    cases_auth,
    cases_session_concurrency,
    cases_provider_portal_onboarding,
    cases_form_validation,
    cases_documents_deep,
    cases_staff_dashboard,
    cases_onboarding_dashboard,
    cases_search_filter_global,
    cases_provider_detail,
    cases_psv_bots,
    cases_continuous_monitoring,
    cases_npdb_deep,
    cases_ny_medicaid,
    cases_committee,
    cases_enrollments,
    cases_hospital_privileges_deep,
    cases_roster_deep,
    cases_expirables_recred,
    cases_oppe_fppe_deep,
    cases_telehealth,
    cases_compliance,
    cases_training,
    cases_print_pdf,
    cases_cme_cv,
    cases_ai_governance,
    cases_public_api_fhir,
    cases_webhooks_security,
    cases_admin,
    cases_reports,
    cases_scorecards_analytics,
    cases_communications,
    cases_audit,
    cases_data_integrity,
    cases_perf_load,
    cases_a11y_ux,
    cases_error_handling,
    cases_i18n_l10n,
    cases_seo_meta,
    cases_browser_compat,
    cases_observability,
    cases_regression,
]


def all_cases() -> list[TestCase]:
    out: list[TestCase] = []
    for fac in ALL_CASE_FACTORIES:
        out.extend(fac())
    return out


# ---------------------------------------------------------------------------
# Workbook builders
# ---------------------------------------------------------------------------

def fill(cell, color: str) -> None:
    cell.fill = PatternFill("solid", fgColor=color)


def write_row(ws: Worksheet, row: int, values: Iterable, *, font=BODY,
              fills: list[str] | None = None, wrap: bool = True) -> None:
    for col_idx, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.font = font
        c.alignment = WRAP if wrap else Alignment(horizontal="left", vertical="center")
        c.border = BORDER_ALL
        if fills and col_idx <= len(fills) and fills[col_idx - 1]:
            fill(c, fills[col_idx - 1])


def banner(ws: Worksheet, row: int, text: str, span: int, *,
           color: str = NAVY, font: Font = TITLE, height: int = 36) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = font
    c.alignment = CENTER
    fill(c, color)
    ws.row_dimensions[row].height = height


def autosize(ws: Worksheet, widths: dict[int, int]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# Sheet: Cover
# ---------------------------------------------------------------------------

def build_cover(wb: Workbook, ts: dt.datetime, total_cases: int) -> None:
    ws = wb.active
    ws.title = "Cover"
    autosize(ws, {1: 28, 2: 80})

    banner(ws, 1, "ESSEN Credentialing Platform", 2, height=42)
    banner(ws, 2, "Master Test Plan & Quality Engineering Tool", 2,
           color=ACCENT, font=SUBTITLE, height=28)

    rows = [
        ("Document version", "1.0"),
        ("Generated", ts.strftime("%Y-%m-%d %H:%M:%S")),
        ("Author", "QA / Master Quality Engineer"),
        ("Reviewed by", "(initials, date)"),
        ("Approved by", "(initials, date)"),
        ("System under test", "ESSEN Credentialing Platform — Next.js 14 + Prisma + PostgreSQL + Redis + BullMQ + Playwright"),
        ("Environment", "Local Docker (web :6015, worker :6025) — shared localai-postgres-1 + redis containers"),
        ("Replaces", "PARCS — system of record for all credentialing activity at Essen Medical"),
        ("Total test cases", str(total_cases)),
        ("Sheets", "Cover · Strategy · Approach & Pattern · Coverage Matrix · Master Test Plan · Defect Log · Sign-off"),
        ("Run cadence", "Smoke before every deploy. Full regression every release. Continuous (CI) on every PR."),
        ("Pass criteria", "100% of P0 cases pass. ≥95% of P1 cases pass. Zero open Critical/High defects."),
    ]
    r = 4
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = BODY_BOLD
        ws.cell(row=r, column=2, value=value).font = BODY
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=GREY_LIGHT)
        ws.cell(row=r, column=1).border = BORDER_ALL
        ws.cell(row=r, column=2).border = BORDER_ALL
        ws.cell(row=r, column=2).alignment = WRAP
        ws.row_dimensions[r].height = 22
        r += 1

    # Mission
    r += 1
    banner(ws, r, "Quality Mission", 2, color=GREY, height=28)
    r += 1
    mission = (
        "Ensure the credentialing platform is fully functional, resilient, secure, "
        "consistent, and user-friendly for every role (PROVIDER, SPECIALIST, MANAGER, "
        "COMMITTEE_MEMBER, ADMIN). Validate every NCQA, HITRUST, SOC2, JC NPG-12, and "
        "CMS-0057-F obligation that this platform is designed to meet. Catch every "
        "regression before it leaves CI; surface ambiguous behavior as defects."
    )
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws.cell(row=r, column=1, value=mission)
    c.font = BODY
    c.alignment = WRAP
    c.border = BORDER_ALL
    ws.row_dimensions[r].height = 80


# ---------------------------------------------------------------------------
# Sheet: Strategy
# ---------------------------------------------------------------------------

def build_strategy(wb: Workbook) -> None:
    ws = wb.create_sheet("Strategy")
    autosize(ws, {1: 26, 2: 90})
    banner(ws, 1, "Test Strategy", 2)

    sections = [
        ("Scope (in)",
         "All 20 functional modules (Provider Onboarding, Staff Dashboard, Committee, Enrollments, "
         "Expirables, PSV Bots, Sanctions, NY Medicaid, Hospital Privileges, NPDB, Recredentialing, "
         "Compliance & Reporting, Verifications, Roster, OPPE/FPPE, Privileging Library, CME & CV, "
         "Public REST API + FHIR, Telehealth, Performance/Analytics) PLUS new 2026 features "
         "(Continuous Monitoring, Audit Packet, Conversational AI, Malpractice Carrier Verification, "
         "SFTP Real, CAQH 2026, Behavioral Health, FSMB PDC, Bot Exception Orchestrator, AI Governance, "
         "HITRUST/SOC2 Readiness, Staff Training Tracker)."),
        ("Scope (out)",
         "Third-party SaaS internals (CAQH, FSMB, SAM.gov, OIG, ABMS) — only contract surface tested. "
         "Production data integrity tests (run against staging only). Penetration testing (separate engagement)."),
        ("Test layers",
         "Unit (Vitest, no I/O) → Integration (Vitest + Testcontainers Postgres + mocked HTTP) → "
         "API contract (JSON Schema + FHIR profiles) → E2E (Playwright Chromium + Firefox) → "
         "Accessibility (axe) → Performance (k6) → Security (OWASP ZAP baseline + manual)."),
        ("Test types per module",
         "Functional · Negative · Boundary · Security · UI · UX · Accessibility · Performance · "
         "Integration · Data integrity · Regression · API contract."),
        ("Roles exercised",
         "PROVIDER (portal), SPECIALIST (default staff), MANAGER (sup), COMMITTEE_MEMBER (vote-only), "
         "ADMIN (full). Anonymous (sign-in, public verify token pages, public REST API w/ key, FHIR)."),
        ("Risk-based prioritization",
         "P0 = blocker (auth, PHI, audit, sanctions, primary credentialing flow). "
         "P1 = high (committee, enrollments, monitoring, comms). "
         "P2 = medium (analytics, admin polish, AI governance). "
         "P3 = low (cosmetic, edge browsers)."),
        ("Entry criteria",
         "All deploy migrations applied · Worker containers healthy · Seed data loaded · "
         "Test users for every role created · API key with fhir:read scope generated · "
         "Test SendGrid sandbox ready · Test SFTP server (or stub) reachable."),
        ("Exit criteria",
         "100% P0 pass · ≥95% P1 pass · ≥85% P2 pass · zero open Critical/High defects · "
         "All in-flight Mediums have a documented mitigation/owner · Smoke regression green for 3 consecutive runs."),
        ("Suspension criteria",
         "P0 failure that blocks ≥3 downstream cases (e.g. login broken) · DB schema drift detected · "
         "Worker container in restart loop · ≥3 unrelated 5xx in any 10-minute window."),
        ("Resumption criteria",
         "Blocking defect fix verified by retest · Re-baseline smoke regression · Re-confirm seed data."),
        ("Test data strategy",
         "Synthetic only — no real PHI ever. Generated via @faker-js/faker + curated fixtures under "
         "tests/support/fixtures/. Each E2E run resets to a known seed."),
        ("Defect lifecycle",
         "NEW → TRIAGED (priority + owner) → IN_PROGRESS → RESOLVED → VERIFIED → CLOSED. "
         "Re-open allowed if regression. SLA: P0 fix <24h, P1 <3d, P2 <7d."),
        ("Tooling",
         "Vitest, Playwright, axe-core, k6, OWASP ZAP, Postman/Insomnia for API, Bull Board for queues, "
         "Prisma Studio for DB inspection, Datadog/Grafana for metrics, Sentry for runtime errors."),
        ("Risks & mitigations",
         "External PSV site downtime → use mocks for unit/integration; document live test cadence. "
         "Flaky bot UI scrapes → wrap in retry harness; fail fast vs hang. "
         "Slow CI from Playwright → shard across browsers; reuse storage state. "
         "Data leakage between tenants → mandatory IDOR test per provider-scoped endpoint."),
        ("Compliance crosswalk",
         "NCQA CR 1-7 · HITRUST CSF v11 r2 · SOC 2 Type II Trust Services Criteria · JC NPG 12 · "
         "CMS-0057-F (FHIR Provider Directory) · NY OMIG screening cadence · NPDB Continuous Query."),
    ]
    r = 3
    for label, value in sections:
        ws.cell(row=r, column=1, value=label).font = BODY_BOLD
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=GREY_LIGHT)
        ws.cell(row=r, column=2, value=value).font = BODY
        ws.cell(row=r, column=2).alignment = WRAP
        for col in (1, 2):
            ws.cell(row=r, column=col).border = BORDER_ALL
        ws.row_dimensions[r].height = max(40, len(value) // 6)
        r += 1


# ---------------------------------------------------------------------------
# Sheet: Approach & Pattern
# ---------------------------------------------------------------------------

def build_approach(wb: Workbook) -> None:
    ws = wb.create_sheet("Approach & Pattern")
    autosize(ws, {1: 6, 2: 28, 3: 90})
    banner(ws, 1, "Streamlined Test Execution Pattern", 3)

    steps = [
        ("1", "Prepare environment",
         "Bring up dev stack (`docker compose up -d`). Verify smoke (00 — Environment & Smoke). "
         "Confirm Prisma client matches schema. Seed test data."),
        ("2", "Triage by priority",
         "Run all P0 cases first across every module. Stop and triage on any P0 fail."),
        ("3", "Module-by-module sweep",
         "Within priority, work top-to-bottom by module ID (00, 01, 02 ... 23). "
         "Within a module, follow the row order — preconditions of later cases often depend on earlier ones."),
        ("4", "Two-pane execution",
         "Open this XLSX in one pane; app + DevTools + Bull Board + Prisma Studio in the other. "
         "Mark Status (Pass/Fail/Blocked/N/A) immediately. Capture screenshots/logs into Defect Log if Fail."),
        ("5", "Negative-after-positive",
         "For every Functional pass, run the matching Negative case to verify rejection. "
         "Don't trust 'happy path' alone."),
        ("6", "Cross-cutting verification",
         "After functional sweep, run Security (PHI, IDOR, headers), Accessibility (axe), "
         "Performance (k6 baseline), and UX consistency passes."),
        ("7", "Defect handling",
         "Log every Fail in 'Defect Log' sheet with screenshot/log/repro. "
         "Set priority. Notify owner via the project tracker. Link defect ID back to test row."),
        ("8", "Retest after fix",
         "On fix delivery, re-run the originally-failed case PLUS the immediately-prior + next case "
         "to catch nearby regressions. Update Status, retain Actual + Notes history."),
        ("9", "Sign-off",
         "When exit criteria met: run smoke regression 3 times; fill Sign-off sheet; "
         "archive XLSX in docs/testing/."),
        ("10", "Continuous improvement",
         "Add any newly-found defect class as a permanent test row in next revision. "
         "Bump file timestamp; never overwrite history."),
    ]
    r = 3
    for n, label, desc in steps:
        ws.cell(row=r, column=1, value=n).font = BODY_BOLD
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=2, value=label).font = BODY_BOLD
        ws.cell(row=r, column=3, value=desc).font = BODY
        ws.cell(row=r, column=3).alignment = WRAP
        for col in (1, 2, 3):
            c = ws.cell(row=r, column=col)
            c.border = BORDER_ALL
            if col != 1:
                c.alignment = WRAP
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=ACCENT_LIGHT)
        ws.row_dimensions[r].height = 60
        r += 1


# ---------------------------------------------------------------------------
# Sheet: Coverage Matrix
# ---------------------------------------------------------------------------

def build_coverage(wb: Workbook, cases: list[TestCase]) -> None:
    ws = wb.create_sheet("Coverage Matrix")
    types = ["Functional", "Negative", "Boundary", "Security", "UI", "UX",
             "A11y", "Performance", "Integration", "API", "Data", "Regression"]
    autosize(ws, {1: 36})
    for i, _ in enumerate(types, start=2):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.column_dimensions[get_column_letter(len(types) + 2)].width = 10

    banner(ws, 1, "Module × Test-Type Coverage Matrix", len(types) + 2)

    headers = ["Module"] + types + ["Total"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = HEADER
        c.alignment = CENTER
        fill(c, ACCENT)
        c.border = BORDER_ALL

    modules = sorted({c.module for c in cases})
    counts: dict[str, dict[str, int]] = {m: {t: 0 for t in types} for m in modules}
    for case in cases:
        if case.test_type in counts[case.module]:
            counts[case.module][case.test_type] += 1
        else:
            counts[case.module][case.test_type] = counts[case.module].get(case.test_type, 0) + 1

    r = 3
    for m in modules:
        ws.cell(row=r, column=1, value=m).font = BODY_BOLD
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=GREY_LIGHT)
        total = 0
        for i, t in enumerate(types, start=2):
            v = counts[m].get(t, 0)
            total += v
            c = ws.cell(row=r, column=i, value=v if v else "")
            c.alignment = CENTER
            c.border = BORDER_ALL
            if v:
                fill(c, ACCENT_LIGHT)
        c = ws.cell(row=r, column=len(types) + 2, value=total)
        c.font = BODY_BOLD
        c.alignment = CENTER
        fill(c, NAVY)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.border = BORDER_ALL
        ws.cell(row=r, column=1).border = BORDER_ALL
        r += 1

    # Grand total row
    ws.cell(row=r, column=1, value="GRAND TOTAL").font = HEADER
    fill(ws.cell(row=r, column=1), NAVY)
    ws.cell(row=r, column=1).border = BORDER_ALL
    grand = 0
    for i, t in enumerate(types, start=2):
        col_total = sum(counts[m].get(t, 0) for m in modules)
        grand += col_total
        c = ws.cell(row=r, column=i, value=col_total)
        c.font = HEADER
        c.alignment = CENTER
        fill(c, NAVY)
        c.border = BORDER_ALL
    c = ws.cell(row=r, column=len(types) + 2, value=grand)
    c.font = HEADER
    c.alignment = CENTER
    fill(c, "059669")
    c.border = BORDER_ALL


# ---------------------------------------------------------------------------
# Sheet: Master Test Plan
# ---------------------------------------------------------------------------

MTP_HEADERS = [
    "Test ID", "Module", "Submodule / Feature", "Test Type", "Priority",
    "Role", "Test Case Title", "Pre-conditions", "Test Steps",
    "Input Fields / Data", "Expected Outcome", "Why this matters",
    "Actual Result", "Status", "Defect ID", "Tester", "Test Date", "Notes",
]


def build_master(wb: Workbook, cases: list[TestCase]) -> None:
    ws = wb.create_sheet("Master Test Plan")

    widths = {
        1: 12,  # Test ID
        2: 28,  # Module
        3: 22,  # Submodule
        4: 13,  # Test Type
        5: 8,   # Priority
        6: 12,  # Role
        7: 50,  # Title
        8: 36,  # Preconditions
        9: 50,  # Steps
        10: 28, # Inputs
        11: 50, # Expected
        12: 50, # Why
        13: 30, # Actual
        14: 11, # Status
        15: 12, # Defect ID
        16: 14, # Tester
        17: 12, # Date
        18: 30, # Notes
    }
    autosize(ws, widths)

    banner(ws, 1, "Master Test Plan — Checklist", len(MTP_HEADERS))

    for col, h in enumerate(MTP_HEADERS, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = HEADER
        c.alignment = CENTER
        fill(c, ACCENT)
        c.border = BORDER_ALL
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "B3"

    r = 3
    for case in cases:
        _id_counter["n"] += 1
        tid = f"TC-{_id_counter['n']:04d}"
        row_values = [
            tid, case.module, case.submodule, case.test_type, case.priority,
            case.role, case.title, case.preconditions, case.steps,
            case.inputs, case.expected, case.explanation,
            "", "", "", "", "", "",  # actual, status, defect, tester, date, notes
        ]
        for col, v in enumerate(row_values, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = BODY
            c.alignment = WRAP
            c.border = BORDER_ALL
            # Alternating row shading
            if r % 2 == 1:
                if col not in (14,):  # leave Status without alt fill
                    fill(c, ROW_ALT)

        # Color-code priority
        prio = case.priority
        prio_cell = ws.cell(row=r, column=5)
        prio_cell.alignment = CENTER
        if prio == "P0":
            fill(prio_cell, RED_LIGHT)
            prio_cell.font = Font(name="Calibri", size=10, bold=True, color=RED)
        elif prio == "P1":
            fill(prio_cell, AMBER_LIGHT)
            prio_cell.font = Font(name="Calibri", size=10, bold=True, color=AMBER)
        elif prio == "P2":
            fill(prio_cell, ACCENT_LIGHT)
            prio_cell.font = Font(name="Calibri", size=10, bold=True, color=ACCENT)
        else:
            fill(prio_cell, GREY_LIGHT)

        # Color-code test type
        type_cell = ws.cell(row=r, column=4)
        type_cell.alignment = CENTER
        type_color_map = {
            "Functional": "DBEAFE",
            "Negative": "FEE2E2",
            "Boundary": "FEF3C7",
            "Security": "FECACA",
            "UI": "E0E7FF",
            "UX": "EDE9FE",
            "A11y": "DCFCE7",
            "Performance": "FED7AA",
            "Integration": "BFDBFE",
            "API": "C7D2FE",
            "Data": "FBCFE8",
            "Regression": "BBF7D0",
        }
        if case.test_type in type_color_map:
            fill(type_cell, type_color_map[case.test_type])

        # Reasonable default row height
        approx = max(
            len(case.steps) // 28,
            len(case.expected) // 28,
            len(case.explanation) // 28,
            len(case.preconditions) // 28,
            3,
        )
        ws.row_dimensions[r].height = max(45, approx * 14)
        r += 1

    last_row = r - 1

    # Data validation: Status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Pass,Fail,Blocked,Not Run,N/A"',
        allow_blank=True,
        showErrorMessage=True,
    )
    status_dv.error = "Select Pass, Fail, Blocked, Not Run, or N/A"
    status_dv.errorTitle = "Invalid status"
    ws.add_data_validation(status_dv)
    status_dv.add(f"N3:N{last_row}")

    # Tester / role example dropdown
    tester_note = DataValidation(
        type="list",
        formula1='"Specialist,Manager,Admin,Provider,Committee,Anonymous,DevOps"',
        allow_blank=True,
    )
    ws.add_data_validation(tester_note)
    tester_note.add(f"F3:F{last_row}")

    # Conditional formatting on Status column (col N = 14)
    ws.conditional_formatting.add(
        f"N3:N{last_row}",
        CellIsRule(operator="equal", formula=['"Pass"'],
                   fill=PatternFill("solid", fgColor=GREEN_LIGHT)),
    )
    ws.conditional_formatting.add(
        f"N3:N{last_row}",
        CellIsRule(operator="equal", formula=['"Fail"'],
                   fill=PatternFill("solid", fgColor=RED_LIGHT)),
    )
    ws.conditional_formatting.add(
        f"N3:N{last_row}",
        CellIsRule(operator="equal", formula=['"Blocked"'],
                   fill=PatternFill("solid", fgColor=AMBER_LIGHT)),
    )
    ws.conditional_formatting.add(
        f"N3:N{last_row}",
        CellIsRule(operator="equal", formula=['"N/A"'],
                   fill=PatternFill("solid", fgColor=GREY_LIGHT)),
    )

    # Auto filter
    ws.auto_filter.ref = f"A2:R{last_row}"


# ---------------------------------------------------------------------------
# Sheet: Defect Log
# ---------------------------------------------------------------------------

def build_defect_log(wb: Workbook) -> None:
    ws = wb.create_sheet("Defect Log")
    headers = [
        "Defect ID", "Linked Test ID", "Module", "Title",
        "Severity", "Priority", "Status",
        "Reproduction Steps", "Expected", "Actual",
        "Environment", "Browser / Device", "Reporter",
        "Reported Date", "Owner", "Resolution", "Resolved Date", "Verified By",
    ]
    widths = [12, 14, 26, 36, 11, 10, 12, 50, 40, 40, 18, 18, 14, 12, 14, 50, 12, 14]
    autosize(ws, {i + 1: w for i, w in enumerate(widths)})

    banner(ws, 1, "Defect Log", len(headers))

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = HEADER
        c.alignment = CENTER
        fill(c, RED)
        c.border = BORDER_ALL
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "B3"

    # Pre-create 200 blank rows with borders + dropdowns so testers can fill.
    for r in range(3, 203):
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=r, column=col, value="")
            c.font = BODY
            c.alignment = WRAP
            c.border = BORDER_ALL
            if r % 2 == 1:
                fill(c, ROW_ALT)

    sev = DataValidation(type="list",
                         formula1='"Critical,High,Medium,Low,Cosmetic"',
                         allow_blank=True)
    ws.add_data_validation(sev)
    sev.add(f"E3:E202")

    prio = DataValidation(type="list",
                          formula1='"P0,P1,P2,P3"',
                          allow_blank=True)
    ws.add_data_validation(prio)
    prio.add(f"F3:F202")

    status = DataValidation(type="list",
                            formula1='"NEW,TRIAGED,IN_PROGRESS,RESOLVED,VERIFIED,CLOSED,REOPENED,WONT_FIX"',
                            allow_blank=True)
    ws.add_data_validation(status)
    status.add(f"G3:G202")

    # Conditional formatting on severity
    ws.conditional_formatting.add(
        "E3:E202",
        CellIsRule(operator="equal", formula=['"Critical"'],
                   fill=PatternFill("solid", fgColor=RED_LIGHT)),
    )
    ws.conditional_formatting.add(
        "E3:E202",
        CellIsRule(operator="equal", formula=['"High"'],
                   fill=PatternFill("solid", fgColor=AMBER_LIGHT)),
    )

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}202"


# ---------------------------------------------------------------------------
# Sheet: Sign-off
# ---------------------------------------------------------------------------

def build_signoff(wb: Workbook, total: int) -> None:
    ws = wb.create_sheet("Sign-off")
    autosize(ws, {1: 30, 2: 60})
    banner(ws, 1, "Test Run Sign-off", 2)

    rows = [
        ("Test cycle / release", ""),
        ("Build / commit SHA", ""),
        ("Environment", "Local Docker dev OR Staging URL"),
        ("Total cases planned", str(total)),
        ("Cases executed", "(count)"),
        ("Pass count", "=COUNTIF('Master Test Plan'!N:N,\"Pass\")"),
        ("Fail count", "=COUNTIF('Master Test Plan'!N:N,\"Fail\")"),
        ("Blocked count", "=COUNTIF('Master Test Plan'!N:N,\"Blocked\")"),
        ("Not Run count", "=COUNTIF('Master Test Plan'!N:N,\"Not Run\")"),
        ("N/A count", "=COUNTIF('Master Test Plan'!N:N,\"N/A\")"),
        ("Pass rate", "=IFERROR(B7/(B7+B8+B9),\"N/A\")"),
        ("Open Critical defects", "=COUNTIFS('Defect Log'!E:E,\"Critical\",'Defect Log'!G:G,\"<>CLOSED\")"),
        ("Open High defects", "=COUNTIFS('Defect Log'!E:E,\"High\",'Defect Log'!G:G,\"<>CLOSED\")"),
        ("Recommendation", "GO / NO-GO"),
        ("Notes", ""),
        ("QA lead", "Name / Signature / Date"),
        ("Engineering lead", "Name / Signature / Date"),
        ("Compliance lead", "Name / Signature / Date"),
        ("Product owner", "Name / Signature / Date"),
    ]
    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = BODY_BOLD
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=GREY_LIGHT)
        ws.cell(row=r, column=1).border = BORDER_ALL
        c = ws.cell(row=r, column=2, value=value)
        c.font = BODY
        c.alignment = WRAP
        c.border = BORDER_ALL
        ws.row_dimensions[r].height = 24
        r += 1


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    cases = all_cases()
    print(f"Built {len(cases)} test cases across "
          f"{len({c.module for c in cases})} modules.")

    ts = dt.datetime.now()
    out_dir = os.path.join(os.path.dirname(__file__))
    os.makedirs(out_dir, exist_ok=True)
    fname = f"ESSEN_Credentialing_Master_Test_Plan_{ts:%Y%m%d_%H%M%S}.xlsx"
    out_path = os.path.join(out_dir, fname)

    wb = Workbook()
    build_cover(wb, ts, len(cases))
    build_strategy(wb)
    build_approach(wb)
    build_coverage(wb, cases)
    build_master(wb, cases)
    build_defect_log(wb)
    build_signoff(wb, len(cases))

    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
