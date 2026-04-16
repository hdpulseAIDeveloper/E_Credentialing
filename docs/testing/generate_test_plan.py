"""
ESSEN Credentialing Platform — Comprehensive QA Test Plan Generator
Master Quality Engineering test plan: 27 sections, 500+ test cases.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

wb = openpyxl.Workbook()

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
MODULE_FONT = Font(name="Calibri", bold=True, color="1F4E79", size=11)
MODULE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBMOD_FONT = Font(name="Calibri", bold=True, color="2E75B6", size=10)
SUBMOD_FILL = PatternFill(start_color="E9F0F7", end_color="E9F0F7", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

COLUMNS = [
    ("A", 9,  "TC#"),
    ("B", 20, "Module"),
    ("C", 22, "Sub-Module / Feature"),
    ("D", 13, "Test Type"),
    ("E", 55, "Test Case Description"),
    ("F", 32, "Input / Precondition"),
    ("G", 38, "Expected Outcome"),
    ("H", 38, "Why This Is Expected"),
    ("I", 8,  "Priority"),
    ("J", 10, "Result\n(P/F/B)"),
    ("K", 38, "Actual Outcome / Notes"),
    ("L", 12, "Tester"),
    ("M", 12, "Date"),
]

def setup_sheet(ws, title):
    ws.title = title
    ws.freeze_panes = "A2"
    for col_letter, width, header in COLUMNS:
        ws.column_dimensions[col_letter].width = width
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN_BORDER
    ws.auto_filter.ref = "A1:M1"

tc = [0]
priority_counts = {"P0": 0, "P1": 0, "P2": 0, "P3": 0}
module_counts = {}

def mod_hdr(ws, row, name):
    for c, _, _ in COLUMNS:
        cell = ws[f"{c}{row}"]
        cell.font = MODULE_FONT
        cell.fill = MODULE_FILL
        cell.border = THIN_BORDER
    ws[f"B{row}"] = name
    return row + 1

def sub_hdr(ws, row, name):
    for c, _, _ in COLUMNS:
        cell = ws[f"{c}{row}"]
        cell.font = SUBMOD_FONT
        cell.fill = SUBMOD_FILL
        cell.border = THIN_BORDER
    ws[f"C{row}"] = name
    return row + 1

def tc_add(ws, row, mod, sub, ttype, desc, inp, exp, why, pri="P1"):
    tc[0] += 1
    priority_counts[pri] = priority_counts.get(pri, 0) + 1
    module_counts[mod] = module_counts.get(mod, 0) + 1
    vals = [f"TC-{tc[0]:04d}", mod, sub, ttype, desc, inp, exp, why, pri, "", "", "", ""]
    for i, (col, _, _) in enumerate(COLUMNS):
        cell = ws[f"{col}{row}"]
        cell.value = vals[i]
        cell.alignment = WRAP
        cell.border = THIN_BORDER
    return row + 1

# ═════════════════════════════════════════════════════════════════════
ws = wb.active
setup_sheet(ws, "Master Test Plan")
r = 2

###################################################################
# 0. INFRASTRUCTURE
###################################################################
r = mod_hdr(ws, r, "0. INFRASTRUCTURE & ENVIRONMENT")
r = sub_hdr(ws, r, "Docker Containers")
for desc, inp, exp, why, pri in [
    ("docker compose up starts all services", "docker compose -f docker-compose.dev.yml up --build", "ecred-web :6015, ecred-worker :6025, postgres :5433, redis :6379 all healthy", "All services must run for testing", "P0"),
    ("Web container responds on :6015", "GET http://localhost:6015", "200 OK HTML", "Web app reachable", "P0"),
    ("Worker health check on :6025", "GET http://localhost:6025/health", "JSON { status: 'ok' }", "Worker operational", "P0"),
    ("Bull Board accessible at :6025/bull-board", "GET http://localhost:6025/bull-board", "UI loads; 3 queues: psv-bot, enrollment-bot, scheduled-jobs", "Job monitoring", "P0"),
    ("Redis connectivity from web container", "docker exec ecred-web redis-cli -h redis ping", "PONG", "Redis required for BullMQ and Socket.io", "P0"),
    ("Containers restart on crash gracefully", "docker restart ecred-web; wait 30s", "Container restarts; health returns 200", "Resilience", "P2"),
]:
    r = tc_add(ws, r, "Infra", "Docker", "Smoke", desc, inp, exp, why, pri)

r = sub_hdr(ws, r, "Database")
for desc, inp, exp, why, pri in [
    ("Prisma migrations fully applied", "npx prisma migrate status", "All migrations applied, no pending", "Schema current", "P0"),
    ("Seed data: provider types exist", "SELECT count(*) FROM provider_types", "≥6 types (MD, DO, PA, NP, LCSW, LMHC)", "Seed required", "P0"),
    ("Seed data: admin user exists", "SELECT * FROM users WHERE role='ADMIN'", "At least 1 admin user with isActive=true", "Admin needed for testing", "P0"),
    ("Database indexes present", "Check pg_indexes for core tables", "Indexes on provider.status, expirable.expirationDate, audit_log.timestamp", "Query performance", "P2"),
    ("/api/health confirms DB connectivity", "GET /api/health", "200 JSON with db: 'connected'", "Health check validates deps", "P0"),
]:
    r = tc_add(ws, r, "Infra", "Database", "Smoke", desc, inp, exp, why, pri)

###################################################################
# 1. AUTH & RBAC
###################################################################
r = mod_hdr(ws, r, "1. AUTHENTICATION & AUTHORIZATION")
r = sub_hdr(ws, r, "Sign In")
for desc, inp, exp, why, pri in [
    ("/auth/signin page renders", "Navigate to /auth/signin", "Sign-in form with Microsoft SSO button and email/password fields", "Auth entry point", "P0"),
    ("Login with valid credentials", "Admin seed email + password", "Redirect to /dashboard; session cookie set; sidebar shows user name", "Core auth", "P0"),
    ("Login with wrong password", "Valid email, bad password", "Error displayed; no redirect; no cookie set", "Rejects invalid creds", "P0"),
    ("Login with non-existent email", "nonexistent@test.com", "Generic error (no user enumeration)", "Prevents email enumeration", "P0"),
    ("Login with deactivated account", "isActive=false user", "Error: 'Account is not active'; no session", "Deactivated blocked", "P0"),
    ("Login with empty email", "Empty email, any password", "Validation: email required", "Client validation", "P1"),
    ("Login with empty password", "Valid email, empty password", "Validation: password required", "Client validation", "P1"),
    ("Session persists across page navigation", "Login, navigate /providers, /enrollments, /committee", "Session maintained; no re-auth required", "Persistent session", "P0"),
    ("Session expires after inactivity", "Login, wait beyond maxAge, refresh", "Redirect to /auth/signin", "Session timeout", "P2"),
]:
    r = tc_add(ws, r, "Auth", "Sign In", "Functional" if "Login" in desc else "UI", desc, inp, exp, why, pri)

r = tc_add(ws, r, "Auth", "Sign Out", "Functional", "Sign out clears session", "Click Sign Out in sidebar", "Redirect to /auth/signin; session cookie removed; back button doesn't restore session", "Clean logout", "P0")
r = tc_add(ws, r, "Auth", "Sign Out", "Security", "After sign out, accessing protected page redirects", "Sign out then navigate to /dashboard", "Redirect to /auth/signin with callbackUrl", "Post-logout protection", "P0")

r = sub_hdr(ws, r, "Session / JWT")
for desc, inp, exp, why, pri in [
    ("JWT contains user.id, user.role, user.email", "GET /api/auth/session after login", "JSON: { user: { id, role, email, name } }", "Session integrity for RBAC", "P0"),
    ("JWT role matches database user role", "Login as SPECIALIST; check session", "session.user.role === 'SPECIALIST'", "Role consistency", "P0"),
    ("Tampering with JWT cookie rejected", "Modify JWT cookie value manually", "Next request returns UNAUTHORIZED or redirect to signin", "JWT integrity", "P1"),
]:
    r = tc_add(ws, r, "Auth", "Session/JWT", "Security", desc, inp, exp, why, pri)

r = sub_hdr(ws, r, "Role-Based Access Control")
roles_matrix = [
    ("ADMIN accesses /admin", "Login ADMIN → /admin", "200 OK; admin pages render", "P0"),
    ("ADMIN accesses /committee", "Login ADMIN → /committee", "200 OK", "P0"),
    ("ADMIN accesses /providers", "Login ADMIN → /providers", "200 OK", "P0"),
    ("MANAGER accesses /admin", "Login MANAGER → /admin", "200 OK; limited admin functions", "P0"),
    ("MANAGER can't create users (admin-only)", "MANAGER calls admin.createUser", "TRPCError FORBIDDEN", "P0"),
    ("SPECIALIST blocked from /admin", "Login SPECIALIST → /admin", "Redirect to /dashboard", "P0"),
    ("SPECIALIST can access /providers", "Login SPECIALIST → /providers", "200 OK", "P0"),
    ("COMMITTEE_MEMBER accesses /committee", "Login COMMITTEE_MEMBER → /committee", "200 OK", "P0"),
    ("COMMITTEE_MEMBER blocked from /admin", "COMMITTEE_MEMBER → /admin", "Redirect to /dashboard", "P0"),
    ("Unauthenticated → redirect from /dashboard", "No cookies → /dashboard", "Redirect /auth/signin?callbackUrl=/dashboard", "P0"),
    ("Unauthenticated → redirect from /providers/[id]", "No cookies → /providers/uuid", "Redirect /auth/signin", "P0"),
    ("staffProcedure rejects unauthenticated", "Call provider.list without session", "UNAUTHORIZED error", "P0"),
    ("managerProcedure rejects SPECIALIST", "SPECIALIST calls committee.recordDecision", "FORBIDDEN error", "P0"),
    ("adminProcedure rejects MANAGER", "MANAGER calls admin.deleteUser", "FORBIDDEN error", "P1"),
    ("Public routes accessible: /, /auth/*, /api/health", "No cookies → /, /auth/signin, /api/health", "200 OK on all", "P0"),
    ("Callback URL preserved after auth redirect", "Navigate to /enrollments?status=PENDING unauthenticated", "After login, redirect back to /enrollments?status=PENDING", "P1"),
]
for desc, inp, exp, pri in roles_matrix:
        r = tc_add(ws, r, "Auth", "RBAC", "Security", desc, inp, exp, "Role enforcement at middleware and procedure level", pri)

r = sub_hdr(ws, r, "Registration")
for desc, inp, exp, why, pri in [
    ("Register page renders at /auth/register", "Navigate to /auth/register", "Form: firstName, lastName, email, password, phone(optional)", "Self-registration", "P1"),
    ("Register with valid data", "firstName:Test, lastName:User, email:test@essen.com, password:Test1234!", "User created; redirect to signin; can login", "Account creation", "P1"),
    ("Register with duplicate email", "Same email as existing user", "409 Conflict error", "Duplicate prevention", "P1"),
    ("Register with weak password (<8 chars)", "password: 'abc'", "400: password must be ≥8 chars", "Password policy", "P1"),
    ("Register without uppercase in password", "password: 'test1234!'", "400: must contain uppercase", "Complexity requirement", "P1"),
    ("Register without digit in password", "password: 'TestPassword!'", "400: must contain digit", "Complexity requirement", "P1"),
    ("Register with invalid email format", "email: 'notanemail'", "400: invalid email format", "Email validation", "P1"),
    ("Newly registered user gets SPECIALIST role", "Register → login → check session", "role === 'SPECIALIST'", "Default role assignment", "P1"),
]:
    r = tc_add(ws, r, "Auth", "Registration", "Functional", desc, inp, exp, why, pri)

###################################################################
# 2. PROVIDER ONBOARDING
###################################################################
r = mod_hdr(ws, r, "2. PROVIDER ONBOARDING")
r = sub_hdr(ws, r, "Provider Creation")
for desc, inp, exp, why, pri in [
    ("AddProviderModal opens from /providers", "Click 'Add Provider' button", "Modal with: First Name, Last Name, Provider Type (dropdown), NPI, Specialist, Email, Phone", "Staff creates providers", "P0"),
    ("Create provider with minimum required fields", "First:'John', Last:'Smith', Type:'MD'", "Provider created (INVITED); checklist auto-created; appears in list; audit log", "Core workflow", "P0"),
    ("Create provider with all optional fields", "Add NPI:1234567890, specialist, email, phone", "All fields saved to provider + profile; specialist relation set", "Full data capture", "P0"),
    ("Missing first name → validation error", "Last:'Smith', Type:'MD' (no first name)", "Error: first name required", "Required field", "P0"),
    ("Missing last name → validation error", "First:'John', Type:'MD' (no last name)", "Error: last name required", "Required field", "P0"),
    ("Missing provider type → validation error", "First:'John', Last:'Smith' (no type)", "Error: provider type required", "Required field", "P0"),
    ("Invalid NPI (5 digits)", "NPI: '12345'", "Error: NPI must be 10 digits", "NPI format", "P1"),
    ("Invalid NPI (letters)", "NPI: 'ABCDEFGHIJ'", "Error: NPI must be numeric", "NPI format", "P1"),
    ("Duplicate NPI warning", "NPI already used by another provider", "Warning or error about duplicate NPI", "Data integrity", "P1"),
    ("Provider appears in list immediately", "After creation, view /providers", "New provider at top of list (or filtered)", "Real-time list update", "P0"),
    ("Provider created with initial status INVITED", "Check provider.status after creation", "status = 'INVITED'", "Correct initial state", "P0"),
    ("Checklist auto-generated based on provider type", "Create MD provider; check checklist", "Checklist items match MD document requirements", "Type-specific requirements", "P1"),
]:
    r = tc_add(ws, r, "Onboarding", "Create Provider", "Functional", desc, inp, exp, why, pri)

r = sub_hdr(ws, r, "Provider Invite")
for desc, inp, exp, why, pri in [
    ("Send invite generates magic link", "Click 'Send Invite' on INVITED provider", "inviteSentAt set; inviteToken generated; inviteTokenExpiresAt = now + 72h", "Initiates provider self-service", "P0"),
    ("Second invite overwrites token", "Send invite again to same provider", "New token generated; old token invalidated", "Fresh token", "P1"),
    ("Invite audit log created", "After sending invite", "AuditLog: action='provider.invite.sent'", "Audit trail", "P1"),
    ("Magic link URL structure", "Inspect generated link", "URL: /application?token=<JWT>", "Correct URL format", "P0"),
]:
    r = tc_add(ws, r, "Onboarding", "Provider Invite", "Functional", desc, inp, exp, why, pri)

r = sub_hdr(ws, r, "Application Form (10 Sections)")
sections = [
    ("0: Personal Info", "firstName, lastName, middleName, dob, gender, ssn", "SSN: XXX-XX-XXXX format", "Status → ONBOARDING_IN_PROGRESS"),
    ("1: Contact Info", "mobilePhone, email, addressLine1, city, state, zip", "Valid email; US state dropdown", "Contact data persisted to ProviderProfile"),
    ("2: Professional IDs", "npi, dea, caqhId, medicarePtan, medicaidId, ecfmg", "NPI: exactly 10 digits", "Identifiers saved to Provider model"),
    ("3: Education", "schoolName, country, graduationYear, degree, internship, residency, fellowship", "Graduation year: 4-digit", "Education data snapshot saved"),
    ("4: Board Certifications", "boardName, specialty, certDate, expDate, certNumber, boardEligible", "Repeatable array; Add/Remove rows", "Board data saved as JSON snapshot"),
    ("5: Work History", "employer, position, startDate, endDate, city, state, reason, isCurrent", "Repeatable; min 5 years coverage", "Work history persisted"),
    ("6: Malpractice", "carrier, policyNumber, coverageAmount, effectiveDate, expDate, hasClaims, claimDetails", "Conditional: claimDetails required if hasClaims=true", "Malpractice data saved"),
    ("7: Hospital Affiliations", "hospital, city, state, privilegeType, appointmentDate, status", "Repeatable array", "Affiliation data saved"),
    ("8: Licenses", "state, licenseNumber, type, issueDate, expDate, isPrimary", "Repeatable; creates License records", "License records created in DB"),
    ("9: Attestation", "attestation1-4 checkboxes, signatureName, signatureDate", "All 4 checkboxes = true; signature required; status → DOCUMENTS_PENDING", "Application complete"),
]
for sec_name, fields, validation, outcome in sections:
    r = tc_add(ws, r, "Onboarding", f"App Sec {sec_name}", "Functional",
               f"Complete section {sec_name} with valid data → Save & Continue",
               f"Fields: {fields}", f"{outcome}; POST /api/application/save-section succeeds",
               "Progressive form saves to DB", "P0")
    r = tc_add(ws, r, "Onboarding", f"App Sec {sec_name}", "Validation",
               f"Submit section {sec_name} with missing required fields",
               f"Validation: {validation}", "Zod validation errors shown inline; form blocked",
               "Client-side validation prevents bad data", "P1")
    r = tc_add(ws, r, "Onboarding", f"App Sec {sec_name}", "Negative",
               f"Submit section {sec_name} with boundary values",
               "Max-length strings, special chars, unicode", "Data saved correctly or validation errors shown",
               "Edge case handling", "P2")

r = tc_add(ws, r, "Onboarding", "App Navigation", "Functional", "Navigate Previous/Next between sections", "Click Previous from section 5", "Returns to section 4; all data preserved", "Bidirectional navigation", "P1")
r = tc_add(ws, r, "Onboarding", "App Navigation", "Functional", "Progress bar reflects current section", "Advance to section 3", "Progress bar shows 3/10 filled", "Visual progress indicator", "P1")
r = tc_add(ws, r, "Onboarding", "App Token", "Negative", "Access /application with expired token", "token=<expired JWT>", "Error: 'Token expired or invalid'", "Expired link rejected", "P0")
r = tc_add(ws, r, "Onboarding", "App Token", "Negative", "Access /application with malformed token", "token=garbage_string", "Error state; no form displayed", "Invalid token rejected", "P0")
r = tc_add(ws, r, "Onboarding", "App Token", "Negative", "Access /application with no token", "Navigate to /application (no query param)", "Error: 'token required'", "Token mandatory", "P0")

r = sub_hdr(ws, r, "Document Upload")
for desc, inp, exp, why, pri in [
    ("Upload valid PDF document", "Select PDF < 10MB via /application/documents", "Upload succeeds; document record created; checklist item → RECEIVED", "Document collection", "P0"),
    ("Upload valid image (JPG/PNG)", "Select JPG image < 10MB", "Upload succeeds; document record created", "Image document support", "P1"),
    ("Reject file > 10MB", "Select 15MB PDF", "Error: file too large", "Size limit", "P1"),
    ("Reject invalid MIME type (.exe)", "Upload .exe file", "Error: unsupported file type", "MIME restriction", "P1"),
    ("Reject invalid MIME type (.bat)", "Upload .bat file", "Error: unsupported file type", "MIME restriction", "P1"),
    ("Multiple file upload (if supported)", "Select 3 PDFs simultaneously", "All 3 uploaded; 3 document records created", "Batch upload", "P2"),
    ("Upload while unauthenticated", "Clear session; POST to /api/upload", "401 Unauthorized", "Auth required for upload", "P0"),
]:
    r = tc_add(ws, r, "Onboarding", "Document Upload", "Functional", desc, inp, exp, why, pri)

###################################################################
# 3. STAFF DASHBOARD
###################################################################
r = mod_hdr(ws, r, "3. STAFF DASHBOARD & PROVIDER MANAGEMENT")
r = sub_hdr(ws, r, "Provider List (/providers)")
for desc, inp, exp, why, pri in [
    ("Page renders with data table", "Navigate /providers", "Table: Name, Type, NPI, Status, Specialist, Actions; pagination", "Primary view", "P0"),
    ("Search by provider name", "Search: 'Smith'", "Only matching providers shown", "Name search", "P0"),
    ("Search is case-insensitive", "Search: 'sMiTh'", "Same results as 'Smith'", "Search usability", "P1"),
    ("Search by partial match", "Search: 'Smi'", "Providers with 'Smi' in first or last name", "Partial match", "P1"),
    ("Filter by status", "Select COMMITTEE_READY", "Only committee-ready providers shown", "Status filter", "P0"),
    ("Filter by provider type", "Select 'MD'", "Only MD providers shown", "Type filter", "P1"),
    ("Filter by assigned specialist", "Select specialist user", "Only providers assigned to that specialist", "Assignment filter", "P1"),
    ("Clear filters resets view", "Click Clear; all filters removed", "Full unfiltered list shown", "Filter reset", "P1"),
    ("Pagination: first page", "Default view", "Page 1 of N; correct item count per page", "Pagination", "P1"),
    ("Pagination: navigate to next", "Click Next", "Page 2 data loads; different providers", "Page navigation", "P1"),
    ("Pagination: navigate to previous", "On page 2, click Previous", "Returns to page 1", "Page navigation", "P1"),
    ("Combined filters + pagination", "Status=APPROVED + page 2", "Correct subset of approved providers on page 2", "Filter + pagination", "P2"),
    ("Empty search returns all", "Clear search field, submit", "All providers shown", "Empty search", "P1"),
    ("No results state", "Search: 'ZZZZZZZZZZZ'", "No providers found message", "Empty state", "P1"),
    ("Row click navigates to detail", "Click provider row", "Navigate to /providers/[id]", "Detail navigation", "P0"),
]:
    r = tc_add(ws, r, "Staff Dashboard", "Provider List", "Functional", desc, inp, exp, why, pri)

r = sub_hdr(ws, r, "Provider Detail Page (15 Tabs)")
tabs = [
    ("Overview", "Provider card, COI tracking panel, onsite meeting panel, identifiers (NPI/DEA/CAQH), timeline milestones"),
    ("Documents", "ChecklistPanel with per-document-type status; upload action; verify/reject; OCR trigger"),
    ("Verifications", "PSV verification records: credential type, status, verified date, flagged indicator, source URL"),
    ("Tasks", "Task list with create/edit/complete/delete; assign to staff; priority HIGH/MED/LOW; due date; comments"),
    ("Communications", "Comms log: email, SMS, phone logs, internal notes; addInternalNote, logPhoneCall, sendFollowUpEmail, sendSms"),
    ("Enrollments", "Enrollment table; status, payer, type; AddEnrollmentModal; follow-up history"),
    ("Expirables", "Expirable list: type, expiration, days left, urgency color; create/edit/delete actions"),
    ("Privileges", "Hospital privileges: facility, type (Active/Courtesy), status, approved date, notes"),
    ("Sanctions/NPDB", "OIG sanctions + SAM sanctions + NPDB records; status badges; acknowledge action for flagged"),
    ("Recredentialing", "Recred cycles: cycle#, due date, status, committee session link; status transitions"),
    ("Work History", "WH verification requests: employer, contact, status (PENDING/SENT/RECEIVED); sendRequest action"),
    ("References", "Reference requests: referee name, relationship, status; Likert ratings if received"),
    ("OPPE/FPPE", "Practice evaluations: type, period, evaluator, status, findings text; complete/cancel actions"),
    ("CME", "CME credits: activity name, category (1/2), credits, date; total credits sum; generate CV"),
    ("Audit Trail", "Timeline: actor, action, timestamp, entity, before/after state diffs; pagination"),
]
for tab, content in tabs:
    r = tc_add(ws, r, "Staff Dashboard", f"Tab: {tab}", "UI",
               f"{tab} tab renders correctly with all expected content",
               f"Navigate to /providers/[id]?tab={tab.lower().replace('/', '-').replace(' ', '-')}",
               f"Content visible: {content}", "Provider hub completeness", "P0")

r = sub_hdr(ws, r, "Provider Actions")
for desc, inp, exp, why, pri in [
    ("Edit provider demographics", "Change first name via header Edit → Save", "Name updated; audit log; page refreshes", "Data editing", "P0"),
    ("Edit provider NPI", "Change NPI to valid 10-digit → Save", "NPI updated; audit log", "ID editing", "P0"),
    ("Edit with invalid data rejected", "Set NPI to 5 digits → Save", "Validation error; no save", "Edit validation", "P1"),
    ("Transition: INVITED → ONBOARDING_IN_PROGRESS", "Via ProviderHeaderActions", "Status updates; badge changes; audit log", "Valid transition", "P0"),
    ("Transition: PSV_COMPLETE → COMMITTEE_READY", "Via header actions", "Status updates; provider appears in committee queue", "Committee readiness", "P0"),
    ("Invalid transition: APPROVED → INVITED", "Attempt via API", "Error: 'Cannot transition from APPROVED to INVITED'", "State machine enforcement", "P0"),
    ("Delete provider (manager only)", "MANAGER calls provider.delete(id)", "Provider status → INACTIVE; soft deleted", "Soft delete", "P1"),
    ("Delete provider (specialist denied)", "SPECIALIST calls provider.delete(id)", "FORBIDDEN error", "Role enforcement", "P1"),
    ("iCIMS Import", "Enter valid iCIMS ID → Import", "Provider created from HR data; profile populated; audit log", "HR integration", "P1"),
    ("iCIMS Import with invalid ID", "Enter non-existent iCIMS ID", "Error: employee not found", "Invalid import", "P2"),
    ("CAQH Sync", "Click CAQH Sync (provider has caqhId)", "CAQH data snapshot saved to profile.caqhDataSnapshot; audit log", "External data", "P1"),
    ("COI update", "Set coiStatus=SENT_TO_BROKER, broker name", "COI fields updated; audit log", "COI workflow", "P1"),
    ("Onsite meeting update", "Set meeting date, attendees, notes", "Meeting data saved", "Meeting tracking", "P2"),
    ("Assign specialist", "Change assignedSpecialist via edit", "Specialist changed; appears in specialist filter", "Work assignment", "P1"),
]:
    r = tc_add(ws, r, "Staff Dashboard", "Provider Actions", "Functional", desc, inp, exp, why, pri)

r = sub_hdr(ws, r, "Tasks & Communications")
for desc, inp, exp, why, pri in [
    ("Create task for provider", "task.create: title, priority, dueDate, assignee, provider", "Task created; visible in task list; assigned user sees it in dashboard", "Task management", "P0"),
    ("Complete task", "Mark task as COMPLETED", "completedAt set; task moves to completed section", "Task lifecycle", "P0"),
    ("Add comment to task", "task.addComment: taskId, content", "Comment appended; visible in task thread", "Collaboration", "P1"),
    ("Send follow-up email", "communication.sendFollowUpEmail: providerId, subject, body", "Communication record created (EMAIL); audit log", "Provider outreach", "P1"),
    ("Log phone call", "communication.logPhoneCall: providerId, notes", "Communication record created (PHONE)", "Call logging", "P1"),
    ("Add internal note", "communication.addInternalNote: providerId, note", "Communication record created (INTERNAL_NOTE)", "Staff notes", "P1"),
    ("Send SMS", "communication.sendSms: providerId, message", "Communication record created (SMS)", "SMS outreach", "P1"),
]:
    r = tc_add(ws, r, "Staff Dashboard", "Tasks/Comms", "Functional", desc, inp, exp, why, pri)

###################################################################
# 4. COMMITTEE
###################################################################
r = mod_hdr(ws, r, "4. COMMITTEE DASHBOARD")
for desc, inp, exp, why, pri in [
    ("/committee page loads with stats and queue", "Navigate /committee", "Stats cards, Active Sessions list, Committee Queue (initial + recred)", "Committee overview", "P0"),
    ("Queue shows initial credentialing providers", "Provider with COMMITTEE_READY status exists", "Provider appears in initial credentialing queue section", "Queue categorization", "P0"),
    ("Queue shows recredentialing providers", "RecredentialingCycle with COMMITTEE_READY status", "Provider appears in recredentialing queue section", "Recred queue", "P1"),
    ("Create committee session", "Fill: date, time, location, select 2 committee members → Create", "Session created SCHEDULED; audit log; appears in active sessions", "Session creation", "P0"),
    ("Create session requires date", "Submit without date", "Validation: sessionDate required", "Required field", "P1"),
    ("Add provider to session", "Open session → select provider from queue → Add", "Provider added; status → COMMITTEE_IN_REVIEW; agenda order auto-assigned", "Session population", "P0"),
    ("Remove provider from session", "Click Remove on provider in session", "Provider removed; returns to queue", "Session editing", "P1"),
    ("Update agenda order", "Drag/reorder providers in session", "Agenda orders updated", "Agenda management", "P2"),
    ("Approve provider (manager)", "Click Approve on provider in session", "Decision: APPROVED; provider.status → APPROVED; initialApprovalDate = now; recred cycle created 36mo", "Approval workflow", "P0"),
    ("Deny provider with reason", "Click Deny → enter reason text → Confirm", "Decision: DENIED; provider.status → DENIED; denial reason stored; audit log", "Denial with reason", "P0"),
    ("Defer provider", "Click Defer", "Decision: DEFERRED; provider.status → DEFERRED; remains available for future session", "Deferred handling", "P1"),
    ("Conditional approval with notes", "Click Conditional → add conditions text", "Decision: CONDITIONAL; notes stored", "Conditional approval", "P1"),
    ("Session: SCHEDULED → IN_PROGRESS", "updateSessionStatus(status: IN_PROGRESS)", "Session status updates; timestamp set", "Session lifecycle", "P0"),
    ("Session: IN_PROGRESS → COMPLETED", "updateSessionStatus(status: COMPLETED)", "Session completed; all decisions finalized", "Session completion", "P0"),
    ("Generate committee agenda HTML", "generateAgenda(sessionId)", "HTML agenda with provider names, types, summary; agendaVersion incremented", "Agenda generation", "P1"),
    ("Send agenda to committee members", "sendAgenda(sessionId)", "Communication records created for each member; audit log", "Agenda distribution", "P1"),
    ("Generate provider summary for committee", "generateSummary(providerId, sessionId)", "HTML summary with licenses, verifications, documents status", "Decision support", "P1"),
    ("Non-manager cannot record decisions", "SPECIALIST calls committee.recordDecision", "FORBIDDEN error", "Role enforcement", "P0"),
    ("Recred cycle completed on re-approval", "Re-approve provider with active recred cycle (COMMITTEE_READY)", "Existing recred cycle: committeeSessionId set, status → COMPLETED", "Cycle linkage", "P1"),
]:
    r = tc_add(ws, r, "Committee", "Committee", "Functional", desc, inp, exp, why, pri)

###################################################################
# 5. ENROLLMENTS
###################################################################
r = mod_hdr(ws, r, "5. ENROLLMENTS")
for desc, inp, exp, why, pri in [
    ("/enrollments page renders with stats and table", "Navigate /enrollments", "Stat cards: Draft/Submitted/Pending/Enrolled/Overdue; filter bar; paginated table", "Enrollment management", "P0"),
    ("Create enrollment (delegated)", "Via AddEnrollmentModal: payer=Anthem, type=DELEGATED, method=PORTAL, followUpCadenceDays=14", "Enrollment DRAFT created; follow-up cadence set; audit log", "Enrollment creation", "P0"),
    ("Create enrollment (facility/BTC)", "Type=FACILITY", "Enrollment created with BTC tracking fields", "Facility enrollment", "P1"),
    ("Create enrollment (direct)", "Type=DIRECT", "Enrollment created", "Direct enrollment", "P1"),
    ("Search by payer name", "Search: 'Anthem'", "Only Anthem enrollments shown", "Payer search", "P1"),
    ("Filter by status", "Filter: SUBMITTED", "Only submitted enrollments", "Status filter", "P1"),
    ("Filter by type", "Filter: DELEGATED", "Only delegated enrollments", "Type filter", "P1"),
    ("Update enrollment status", "Change DRAFT → SUBMITTED with submittedDate", "Status updated; submittedDate set; audit log", "Status workflow", "P0"),
    ("Add follow-up to enrollment", "enrollment.addFollowUp: outcome, nextFollowUpDate", "Follow-up record created; nextFollowUpAt updated", "Follow-up tracking", "P0"),
    ("Delete (withdraw) enrollment", "enrollment.delete(id)", "Enrollment status → WITHDRAWN", "Soft delete", "P1"),
    ("Generate roster CSV", "enrollment.generateRoster with selected enrollment IDs", "CSV string generated with provider/payer data", "Roster generation", "P1"),
    ("EFT/ERA status tracking", "Update eftStatus, eraStatus on enrollment", "Fields persisted; visible in detail", "Payment tracking", "P2"),
    ("Enrollment detail page loads", "Navigate /enrollments/[id]", "Payer, type, status, dates, follow-up history visible", "Detail view", "P0"),
    ("Overdue follow-ups appear on dashboard", "Enrollment with follow-up cadence past due", "Appears in dashboard Overdue Follow-ups section", "Cross-module alert", "P1"),
]:
    r = tc_add(ws, r, "Enrollments", "Enrollments", "Functional", desc, inp, exp, why, pri)

###################################################################
# 6. PSV BOTS
###################################################################
r = mod_hdr(ws, r, "6. CREDENTIALING BOTS (PSV)")
bot_types = [
    ("LICENSE_VERIFICATION", "State license verification", "BotRun QUEUED; job enqueued to psv-bot; verification record created on completion"),
    ("DEA_VERIFICATION", "DEA number verification", "BotRun QUEUED; DEA verified or flagged"),
    ("BOARD_NCCPA", "NCCPA board cert check", "BotRun QUEUED; board status verified"),
    ("BOARD_ABIM", "ABIM board cert check", "BotRun QUEUED; ABIM status verified"),
    ("BOARD_ABFM", "ABFM board cert check", "BotRun QUEUED; ABFM status verified"),
    ("OIG_SANCTIONS", "OIG exclusion list check", "SanctionsCheck + BotRun created; CLEAR or FLAGGED result"),
    ("SAM_SANCTIONS", "SAM.gov exclusion check", "SanctionsCheck + BotRun created; CLEAR or FLAGGED result"),
    ("NPDB", "NPDB query", "NpdbRecord + BotRun created"),
    ("EDUCATION_AMA", "AMA education verify (placeholder)", "BotRun QUEUED; worker logs 'not yet implemented' — no crash"),
    ("EDUCATION_ECFMG", "ECFMG education verify (placeholder)", "BotRun QUEUED; worker logs 'not yet implemented' — no crash"),
]
for bt, bdesc, bexp in bot_types:
    r = tc_add(ws, r, "PSV Bots", f"Bot: {bt}", "Functional",
               f"Trigger {bt} bot for provider",
               f"bot.triggerBot(providerId, botType={bt})", bexp,
               f"PSV: {bdesc}", "P1")

for desc, inp, exp, why, pri in [
    ("Bot run history page renders", "/providers/[id]/bots", "Bot runs list: status, type, triggered by, started/completed dates", "Bot monitoring", "P1"),
    ("Acknowledge flagged verification (manager)", "Click Acknowledge on flagged VR", "acknowledgedBy/At set; flag resolved", "Flagged review", "P1"),
    ("Acknowledge denied for specialist", "SPECIALIST calls bot.acknowledgeFlag", "FORBIDDEN error", "Role enforcement", "P1"),
    ("Bot status shows in BotStatusPanel", "Trigger bot; view panel", "Panel shows QUEUED → RUNNING → COMPLETED lifecycle", "Status tracking", "P1"),
    ("Bot enum maps to job name correctly", "Trigger each bot type", "BOT_TYPE_MAP maps enum to correct BullMQ job name", "Enum mapping", "P1"),
]:
    r = tc_add(ws, r, "PSV Bots", "Bot Operations", "Functional", desc, inp, exp, why, pri)

###################################################################
# 7. SANCTIONS
###################################################################
r = mod_hdr(ws, r, "7. SANCTIONS CHECKING")
for desc, inp, exp, why, pri in [
    ("Trigger OIG check", "sanctions.triggerCheck(providerId, source: OIG)", "SanctionsCheck created; BotRun queued; result=CLEAR or FLAGGED", "OIG monitoring", "P0"),
    ("Trigger SAM check", "sanctions.triggerCheck(providerId, source: SAM_GOV)", "SanctionsCheck created; BotRun queued", "SAM monitoring", "P0"),
    ("View provider sanctions history", "sanctions.listByProvider(providerId)", "Chronological sanctions checks with source, result, date", "History view", "P1"),
    ("View all flagged sanctions", "sanctions.getFlagged()", "All FLAGGED + unacknowledged entries", "Urgent attention", "P0"),
    ("Acknowledge flagged sanction (manager)", "sanctions.acknowledge(id)", "isAcknowledged=true; acknowledgedBy/At set; audit log", "Manager review", "P0"),
    ("Acknowledge denied for specialist", "SPECIALIST calls sanctions.acknowledge", "FORBIDDEN", "Role enforcement", "P1"),
    ("Weekly scheduled job runs", "Via Bull Board: trigger sanctions-weekly", "All approved providers checked; results stored", "Automation", "P1"),
    ("Sanctions appear on provider detail", "Check Sanctions/NPDB tab on provider", "OIG + SAM results with status badges", "Cross-module display", "P1"),
    ("Flagged sanction appears on compliance page", "Provider has FLAGGED sanction", "Compliance card shows decreased sanctions compliance", "Cross-module impact", "P1"),
]:
    r = tc_add(ws, r, "Sanctions", "Sanctions", "Functional", desc, inp, exp, why, pri)

###################################################################
# 8. EXPIRABLES
###################################################################
r = mod_hdr(ws, r, "8. EXPIRABLES TRACKING")
for desc, inp, exp, why, pri in [
    ("/expirables renders with urgency cards", "Navigate /expirables", "Cards: Expired, <7d, <30d, <60d, <90d; filter bar; data table", "Expirables hub", "P0"),
    ("Urgency cards match data counts", "Count by urgency bucket", "Card numbers = DB count", "Data accuracy", "P1"),
    ("Filter by type (STATE_LICENSE)", "Select STATE_LICENSE from dropdown", "Only license expirables shown", "Type filter", "P1"),
    ("Filter by urgency (EXPIRED)", "Select EXPIRED", "Only expired items", "Urgency filter", "P1"),
    ("Search expirables", "Search: provider name", "Matching provider's expirables shown", "Search", "P1"),
    ("Row colors by urgency", "View table with mixed urgencies", "Red: expired; Yellow: <30d; Normal: >30d", "Visual urgency", "P1"),
    ("Create expirable", "expirable.create: providerId, type, expirationDate", "Created as CURRENT; appears in list + provider tab", "Create", "P1"),
    ("Update expirable", "expirable.update: change expirationDate", "Date updated; urgency recalculated", "Edit", "P1"),
    ("Delete expirable", "expirable.delete(id)", "Removed from active list", "Delete", "P2"),
    ("Provider detail expirables tab", "View Expirables tab on provider", "Provider-specific expirables with same urgency coloring", "Provider context", "P0"),
    ("Daily scan job updates statuses", "Via Bull Board: trigger expirables-scan", "Statuses updated: CURRENT → EXPIRING_SOON based on dates; tasks created for managers", "Automation", "P1"),
    ("Expirables appear on dashboard sidebar", "Dashboard → Upcoming Expirations", "Shows closest expiring items with clickable link to /expirables", "Dashboard integration", "P1"),
]:
    r = tc_add(ws, r, "Expirables", "Expirables", "Functional", desc, inp, exp, why, pri)

###################################################################
# 9. NY MEDICAID / ETIN
###################################################################
r = mod_hdr(ws, r, "9. NY MEDICAID / ETIN")
for desc, inp, exp, why, pri in [
    ("/medicaid page renders", "Navigate /medicaid", "Summary stats + enrollment list", "Medicaid management", "P1"),
    ("Create new Medicaid enrollment", "/medicaid/new → wizard", "MedicaidEnrollment created; audit log", "Enrollment creation", "P1"),
    ("Enrollment wizard: select provider", "Pick approved provider from dropdown", "Provider selected; NPI/name populated", "Provider picker", "P1"),
    ("Confirm ETIN number", "medicaid.confirmEtin(id, etinNumber)", "etinNumber saved; etinConfirmedDate = now", "ETIN tracking", "P1"),
    ("Update status through workflow", "SUBMITTED → PENDING → ENROLLED", "Status updates; dates set", "Status workflow", "P1"),
    ("Record eMedNY submission", "medicaid.recordSubmission(id, confirmationNumber)", "Submission recorded; audit log", "Submission tracking", "P1"),
    ("Medicaid detail page", "/medicaid/[id]", "Full detail with status, ETIN, PSP, submissions", "Detail view", "P1"),
    ("Delete enrollment (manager)", "medicaid.delete(id) as MANAGER", "Enrollment marked deleted", "Soft delete", "P2"),
    ("Delete denied for specialist", "SPECIALIST calls medicaid.delete", "FORBIDDEN", "Role enforcement", "P2"),
    ("Summary stats accuracy", "medicaid.getSummary()", "Counts match actual DB records per status", "Data accuracy", "P1"),
]:
    r = tc_add(ws, r, "Medicaid", "Medicaid/ETIN", "Functional", desc, inp, exp, why, pri)

###################################################################
# 10. HOSPITAL PRIVILEGES
###################################################################
r = mod_hdr(ws, r, "10. HOSPITAL PRIVILEGES")
for desc, inp, exp, why, pri in [
    ("Privileges panel on provider detail", "Provider detail → Privileges tab", "Table: facility, type, status, approved date, expiry", "Privilege view", "P0"),
    ("Update privilege status: APPLIED → APPROVED", "provider.updateHospitalPrivilege(id, status:APPROVED)", "Status updated; approvedDate set; audit log", "Status workflow", "P1"),
    ("Update privilege with notes", "Add notes to privilege update", "Notes persisted and visible", "Documentation", "P2"),
    ("Multiple facilities per provider", "Provider with 3 hospital privileges", "All 3 visible in tab", "Multi-facility", "P1"),
]:
    r = tc_add(ws, r, "Privileges", "Hospital Privileges", "Functional", desc, inp, exp, why, pri)

###################################################################
# 11. NPDB
###################################################################
r = mod_hdr(ws, r, "11. NPDB")
for desc, inp, exp, why, pri in [
    ("Trigger NPDB initial query", "npdb.triggerQuery(providerId, queryType:INITIAL)", "NpdbRecord created (PENDING); job enqueued to psv-bot (npdb-query)", "NPDB initiation", "P0"),
    ("Trigger NPDB on-demand query", "npdb.triggerQuery(providerId, queryType:ON_DEMAND)", "NpdbRecord + BotRun created", "On-demand query", "P1"),
    ("View provider NPDB records", "npdb.listByProvider(providerId)", "Chronological NPDB records with query type, result, date", "History", "P1"),
    ("View adverse NPDB findings", "npdb.getAdverse()", "All unacknowledged adverse findings", "Urgent attention", "P1"),
    ("Acknowledge adverse finding (manager)", "npdb.acknowledge(id)", "isAcknowledged=true; acknowledgedBy/At set; audit log", "Manager review", "P1"),
    ("NPDB data appears on provider Sanctions/NPDB tab", "Provider detail → Sanctions/NPDB tab", "NPDB records section with result badges", "Cross-module display", "P0"),
    ("NPDB score reflected in scorecards", "Provider with ADVERSE npdb record", "Scorecard NPDB column shows low score", "Score impact", "P1"),
]:
    r = tc_add(ws, r, "NPDB", "NPDB", "Functional", desc, inp, exp, why, pri)

###################################################################
# 12. RECREDENTIALING
###################################################################
r = mod_hdr(ws, r, "12. RECREDENTIALING")
for desc, inp, exp, why, pri in [
    ("/recredentialing page renders", "Navigate /recredentialing", "Summary cards, filters, cycle table", "Recred overview", "P0"),
    ("Auto-cycle on approval", "Approve provider in committee", "RecredentialingCycle: cycleNumber=1, dueDate=approval+36mo, status=PENDING", "Auto-creation", "P0"),
    ("Manual cycle creation (manager)", "recredentialing.create(providerId, dueDate)", "Cycle created; audit log", "Manual creation", "P1"),
    ("Cycle status: PENDING → IN_PROGRESS", "recredentialing.updateStatus(id, IN_PROGRESS)", "Status updated; inProgressAt set", "Workflow", "P1"),
    ("Cycle status: IN_PROGRESS → PSV_RUNNING", "updateStatus(PSV_RUNNING)", "Updated; PSV bots should be triggered", "PSV phase", "P1"),
    ("Cycle status: PSV_RUNNING → COMMITTEE_READY", "updateStatus(COMMITTEE_READY)", "Updated; appears in committee recred queue", "Committee ready", "P1"),
    ("Bulk initiate (manager)", "recredentialing.initiateBulk()", "Cycles created for all eligible providers without active cycles", "Bulk operation", "P1"),
    ("Delete cycle (PENDING only, manager)", "recredentialing.delete(id) where status=PENDING", "Deleted", "Delete pending", "P2"),
    ("Delete non-PENDING cycle denied", "Delete cycle with IN_PROGRESS status", "Error: can only delete PENDING cycles", "Status restriction", "P2"),
    ("Daily check marks overdue", "Cycle with dueDate in past, status≠COMPLETED", "Worker job sets status → OVERDUE", "Overdue detection", "P1"),
    ("Overdue appears on dashboard alerts", "Overdue cycle exists", "Dashboard Cross-Module Alerts: 'N overdue recredentialing cycles'", "Alert integration", "P1"),
    ("Recred on provider detail tab", "Provider detail → Recredentialing tab", "Cycle list with due dates, status, session links", "Provider context", "P0"),
    ("Scorecard recred column", "Approved provider with active cycle", "Scorecards show recred completion status", "Score impact", "P1"),
]:
    r = tc_add(ws, r, "Recredentialing", "Recredentialing", "Functional", desc, inp, exp, why, pri)

###################################################################
# 13. COMPLIANCE & REPORTING
###################################################################
r = mod_hdr(ws, r, "13. COMPLIANCE & REPORTING")
r = sub_hdr(ws, r, "NCQA Compliance (/compliance)")
for desc, inp, exp, why, pri in [
    ("Compliance page renders with 6 cards + score", "Navigate /compliance", "Overall NCQA score circle; 6 category cards", "Compliance dashboard", "P0"),
    ("PSV card: checks LICENSE, DEA, BOARD", "Approved providers with varying PSV status", "Card shows N/Total verified across license, DEA, board cert types", "PSV breadth", "P1"),
    ("Sanctions card says 'Weekly'", "Read card description", "'Weekly OIG and SAM.gov exclusion checks...'", "Corrected label", "P1"),
    ("NPDB card: continuous query status", "Providers with/without NPDB queries", "Shows NPDB query completion rate", "NPDB compliance", "P1"),
    ("180-day credentialing card", "Mix of fast/slow credentialing", "Shows % completed within 180 days", "Timeliness metric", "P1"),
    ("Recredentialing card", "Mix of on-time and overdue cycles", "Shows % on-time recred cycle completion", "Recred compliance", "P1"),
    ("File completeness card", "Providers with varying doc completion", "Shows % with complete document checklist", "Documentation compliance", "P1"),
    ("Overall score is weighted average", "Known provider mix", "Score matches manual weighted calculation", "Score accuracy", "P1"),
]:
    r = tc_add(ws, r, "Compliance", "NCQA Dashboard", "Functional", desc, inp, exp, why, pri)

r = sub_hdr(ws, r, "Reports (/reports)")
for desc, inp, exp, why, pri in [
    ("Reports page renders", "Navigate /reports", "Compliance metrics summary; export cards; saved reports list", "Reports hub", "P0"),
    ("Export Providers CSV", "/reports/export?type=providers → Download", "CSV with columns: name, npi, status, type, specialist, dates", "Data export", "P0"),
    ("Export Enrollments CSV", "type=enrollments", "CSV with payer, type, status, dates", "Export", "P1"),
    ("Export Expirables CSV", "type=expirables", "CSV with type, expiration, status, provider", "Export", "P1"),
    ("Export Recredentialing CSV", "type=recredentialing", "CSV with cycle data", "Export", "P1"),
    ("Export Sanctions (Coming Soon)", "type=sanctions", "'Coming Soon' label or alert shown", "Graceful placeholder", "P2"),
    ("Save report", "report.saveReport(name, type, filters)", "Saved report appears in list", "Report saving", "P1"),
    ("Delete saved report", "report.deleteReport(id)", "Report removed from list", "Report management", "P2"),
    ("Compliance summary data", "report.complianceSummary()", "Returns computed compliance metrics matching dashboard", "API parity", "P1"),
]:
    r = tc_add(ws, r, "Reports", "Reports & Export", "Functional", desc, inp, exp, why, pri)

###################################################################
# 14. VERIFICATIONS
###################################################################
r = mod_hdr(ws, r, "14. VERIFICATIONS (Work History & References)")
r = sub_hdr(ws, r, "Work History")
for desc, inp, exp, why, pri in [
    ("/verifications page with Work History tab", "Navigate /verifications?tab=work-history", "Table: provider, employer, contact, status, sent/received dates", "WH management", "P0"),
    ("Create WH request", "workHistory.create(providerId, employerName, contactEmail)", "Request created PENDING; responseToken generated (UUID)", "Request creation", "P0"),
    ("Send WH request email", "workHistory.sendRequest(id)", "Status → SENT; requestSentAt set; audit log", "Email outreach", "P0"),
    ("Send reminder", "workHistory.sendReminder(id)", "reminderSentAt updated; audit log", "Follow-up", "P1"),
    ("Delete WH request", "workHistory.delete(id)", "Request deleted", "Cleanup", "P2"),
    ("Public form loads with token", "/verify/work-history/[validToken]", "Form displays: provider name, employer fields, verified checkbox, dates, comments", "External form", "P0"),
    ("Public form submission", "Fill form → Submit", "Status → RECEIVED; responseData saved; respondedAt set; audit log", "Data collection", "P0"),
    ("Re-submit rejected (already responded)", "Visit same token URL after submission", "Message: 'already been submitted'", "Duplicate prevention", "P1"),
    ("Invalid token shows error", "/verify/work-history/[invalidToken]", "Error: verification not found", "Token validation", "P1"),
    ("WH status on provider detail", "Provider detail → Work History tab", "Requests with status badges", "Cross-module view", "P0"),
]:
    r = tc_add(ws, r, "Verifications", "Work History", "Functional", desc, inp, exp, why, pri)

r = sub_hdr(ws, r, "Professional References")
for desc, inp, exp, why, pri in [
    ("References tab on /verifications", "Navigate /verifications?tab=references", "Table: provider, referee, relationship, status, dates", "Ref management", "P0"),
    ("Create reference request", "reference.create(providerId, refereeName, email, relationship)", "Request PENDING; responseToken generated", "Request creation", "P0"),
    ("Send reference request", "reference.sendRequest(id)", "Status → SENT; requestSentAt; audit log", "Email outreach", "P0"),
    ("Send reminder", "reference.sendReminder(id)", "reminderSentAt updated; audit log", "Follow-up", "P1"),
    ("Public reference form loads", "/verify/reference/[validToken]", "Form: Likert scales (5 categories), recommendation dropdown, comments", "External form", "P0"),
    ("Submit reference form", "Fill ratings + recommendation → Submit", "Status → RECEIVED; ratings saved; audit log", "Data collection", "P0"),
    ("Re-submit rejected", "Same token after submission", "'Already submitted' message", "Duplicate prevention", "P1"),
    ("Provider detail references tab", "Provider → References tab", "Reference requests with status and ratings if received", "Cross-module", "P0"),
    ("Pending references on dashboard alerts", "Pending references exist", "Dashboard: 'N pending professional references'", "Alert integration", "P1"),
]:
    r = tc_add(ws, r, "Verifications", "References", "Functional", desc, inp, exp, why, pri)

###################################################################
# 15. ROSTER
###################################################################
r = mod_hdr(ws, r, "15. ROSTER MANAGEMENT")
for desc, inp, exp, why, pri in [
    ("/roster page renders", "Navigate /roster", "Roster list with payer, format, latest submission", "Roster management", "P1"),
    ("Create roster (manager)", "roster.createRoster(payerName, format, template)", "PayerRoster created; appears in list", "Roster setup", "P1"),
    ("Generate submission", "roster.generateSubmission(rosterId)", "RosterSubmission created: providerCount, csv blob, GENERATED", "CSV generation", "P1"),
    ("Validate submission", "roster.validateSubmission(submissionId)", "Validation runs; errors/warnings noted", "Quality check", "P1"),
    ("Submit roster", "roster.submitRoster(submissionId)", "Status → SUBMITTED; submittedAt set", "Submission tracking", "P1"),
    ("Acknowledge roster", "roster.acknowledgeRoster(submissionId)", "Status → ACKNOWLEDGED", "Completion", "P2"),
    ("Delete roster (manager)", "roster.deleteRoster(id)", "Roster removed", "Cleanup", "P2"),
]:
    r = tc_add(ws, r, "Roster", "Roster Management", "Functional", desc, inp, exp, why, pri)

###################################################################
# 16. OPPE/FPPE
###################################################################
r = mod_hdr(ws, r, "16. OPPE/FPPE EVALUATIONS")
for desc, inp, exp, why, pri in [
    ("/evaluations page renders", "Navigate /evaluations", "Summary stats; table with provider, type, period, evaluator, status", "Evaluation hub", "P0"),
    ("Create OPPE evaluation (manager)", "evaluation.create(OPPE, providerId, periodStart/End, dueDate)", "Evaluation SCHEDULED; audit log", "OPPE creation", "P1"),
    ("Create FPPE evaluation (manager)", "evaluation.create(FPPE, providerId, privilegeId, ...)", "Evaluation SCHEDULED; linked to hospital privilege", "FPPE creation", "P1"),
    ("Update evaluation status", "evaluation.update(id, status:IN_PROGRESS)", "Status updated; timestamps set", "Workflow", "P1"),
    ("Complete evaluation with findings", "update(id, status:COMPLETED, findings text)", "Completed; findings stored", "Completion", "P1"),
    ("Delete evaluation (SCHEDULED only, manager)", "evaluation.delete(id) where SCHEDULED", "Deleted", "Cleanup", "P2"),
    ("Delete non-SCHEDULED denied", "Delete IN_PROGRESS evaluation", "Error: only SCHEDULED can be deleted", "Restriction", "P2"),
    ("Dashboard evaluation alerts", "Evaluations due soon exist", "Dashboard: 'N evaluations due within 30 days'", "Alert", "P1"),
    ("Provider OPPE/FPPE tab", "Provider → OPPE/FPPE tab", "Evaluations with type, period, evaluator, findings", "Provider context", "P0"),
    ("Scorecards not affected by eval data (yet)", "Check if evaluation data in scorecards", "Evaluations tracked separately from scorecard", "Scope clarity", "P2"),
]:
    r = tc_add(ws, r, "Evaluations", "OPPE/FPPE", "Functional", desc, inp, exp, why, pri)

###################################################################
# 17. PRIVILEGING LIBRARY
###################################################################
r = mod_hdr(ws, r, "17. PRIVILEGING LIBRARY")
for desc, inp, exp, why, pri in [
    ("/admin/privileging renders (admin)", "Navigate as ADMIN", "Privilege category list with item counts", "Library management", "P1"),
    ("Create category", "privileging.createCategory(name, specialty)", "Category created; listed", "Catalog management", "P1"),
    ("Create item with CPT/ICD codes", "privileging.createItem(categoryId, name, cptCodes:['99213'], icd10Codes:['J06.9'], isCore:true)", "Item created under category", "Delineation", "P1"),
    ("Search privileges", "privileging.search(query:'99213')", "Items matching CPT code returned", "Code lookup", "P2"),
    ("Update category", "privileging.updateCategory(id, newName)", "Name updated", "Editing", "P2"),
    ("Delete empty category (admin)", "privileging.deleteCategory(id) with no items", "Deleted", "Cleanup", "P2"),
    ("Non-admin blocked", "SPECIALIST navigates /admin/privileging", "Redirect to /dashboard", "Role restriction", "P1"),
]:
    r = tc_add(ws, r, "Privileging", "Privileging Library", "Functional", desc, inp, exp, why, pri)

###################################################################
# 18. CME & CV
###################################################################
r = mod_hdr(ws, r, "18. CME & CV")
for desc, inp, exp, why, pri in [
    ("/cme page renders", "Navigate /cme", "Provider list with CME credit totals; progress toward 50-credit target", "CME overview", "P0"),
    ("Create CME credit", "cme.create(providerId, activityName, category:1, credits:2.5, completedDate)", "Credit created; audit log; total updates", "Credit tracking", "P1"),
    ("Update CME credit", "cme.update(id, credits:3.0)", "Updated; audit log; total recalculated", "Credit editing", "P1"),
    ("Delete CME credit", "cme.delete(id)", "Credit removed; total decreases", "Credit removal", "P2"),
    ("CME summary with totals", "cme.getSummary(providerId)", "Total credits, Category 1 total, Category 2 total, compliance status", "Summary report", "P1"),
    ("Generate CV", "cme.generateCv(providerId)", "Text CV: education, licenses, privileges, CME, work history", "CV generation", "P1"),
    ("Provider CME tab", "Provider → CME tab", "Credit list with name, category, credits, date; total sum", "Provider context", "P0"),
    ("CME in scorecards", "Provider with 40/50 credits", "Scorecard CME column shows 80%", "Score impact", "P1"),
    ("CME 50-credit target progress bar", "View /cme page", "Progress bar shows X/50 credits toward target", "Visual tracking", "P1"),
]:
    r = tc_add(ws, r, "CME", "CME & CV", "Functional", desc, inp, exp, why, pri)

###################################################################
# 19. REST API & FHIR
###################################################################
r = mod_hdr(ws, r, "19. PUBLIC REST API & FHIR")
r = sub_hdr(ws, r, "API Key Management")
for desc, inp, exp, why, pri in [
    ("Create API key (admin)", "apiKey.create(name:'Test', permissions:{providers:read:true})", "Key created; plaintext 'essen_<hex>' returned ONCE; hash stored", "Key provisioning", "P0"),
    ("Plaintext only shown once", "Retrieve key after creation", "No endpoint to retrieve plaintext; only hash in DB", "Security", "P0"),
    ("List API keys", "apiKey.list()", "Table: name, permissions, createdAt, lastUsedAt, active", "Key overview", "P1"),
    ("Revoke API key", "apiKey.revoke(id)", "isActive=false; key rejected on next API call", "Key revocation", "P0"),
    ("Delete API key", "apiKey.delete(id)", "Key removed from DB", "Key cleanup", "P1"),
    ("/admin/api-keys page", "Navigate as ADMIN", "API key management UI", "Admin page", "P1"),
]:
    r = tc_add(ws, r, "REST API", "API Keys", "Functional", desc, inp, exp, why, pri)

r = sub_hdr(ws, r, "REST v1 Endpoints")
for desc, inp, exp, why, pri in [
    ("GET /api/v1/providers with valid key", "Bearer token; providers:read permission", "200: paginated provider list (id, name, npi, status, type)", "Provider API", "P0"),
    ("GET /api/v1/providers without auth", "No Authorization header", "401 Unauthorized JSON", "Auth enforcement", "P0"),
    ("GET /api/v1/providers with revoked key", "Revoked/expired key", "403 Forbidden", "Revoked rejection", "P0"),
    ("GET /api/v1/providers wrong permission", "Key has enrollments:read only", "403 insufficient permissions", "Permission granularity", "P1"),
    ("GET /api/v1/providers pagination", "?page=1&limit=10", "10 results; total count header/field", "Pagination", "P1"),
    ("GET /api/v1/providers/[id]", "Valid UUID, providers:read key", "200: provider detail with licenses, enrollments", "Single provider", "P1"),
    ("GET /api/v1/providers/[invalid-id]", "Non-existent UUID", "404 Not Found", "Missing resource", "P1"),
    ("GET /api/v1/enrollments", "enrollments:read key", "200: paginated enrollment list", "Enrollment API", "P1"),
    ("GET /api/v1/sanctions", "sanctions:read key", "200: paginated sanctions list", "Sanctions API", "P1"),
    ("lastUsedAt updated on each call", "Call API twice, check DB", "apiKey.lastUsedAt updated", "Usage tracking", "P2"),
]:
    r = tc_add(ws, r, "REST API", "v1 Endpoints", "Functional", desc, inp, exp, why, pri)

r = sub_hdr(ws, r, "FHIR R4")
for desc, inp, exp, why, pri in [
    ("GET /api/fhir/Practitioner", "API key with fhir:read permission", "200: FHIR Bundle (searchset) with Practitioner resources for approved providers", "FHIR compliance", "P0"),
    ("FHIR without fhir:read", "Key missing fhir:read", "403: OperationOutcome 'lacks fhir:read permission'", "Permission check", "P0"),
    ("FHIR without any key", "No Authorization header", "401", "Auth required", "P0"),
    ("FHIR search by NPI", "?identifier=<npi>", "Bundle filtered to matching NPI practitioner(s)", "NPI search", "P1"),
    ("FHIR search by name", "?name=Smith", "Bundle filtered by name", "Name search", "P1"),
    ("FHIR _count parameter", "?_count=5", "Bundle limited to 5 entries", "Pagination", "P1"),
    ("FHIR response structure", "Inspect response JSON", "resourceType:Bundle, type:searchset, entry[].resource.resourceType:Practitioner", "FHIR R4 structure", "P0"),
    ("Practitioner resource has NPI identifier", "Check entry[0].resource.identifier", "system:http://hl7.org/fhir/sid/us-npi, value:<npi>", "NPI in FHIR", "P1"),
]:
    r = tc_add(ws, r, "REST API", "FHIR R4", "Functional", desc, inp, exp, why, pri)

###################################################################
# 20. TELEHEALTH
###################################################################
r = mod_hdr(ws, r, "20. TELEHEALTH CREDENTIALING")
for desc, inp, exp, why, pri in [
    ("/telehealth page renders", "Navigate /telehealth", "Provider list: platform, training date, certified badge, states", "Telehealth view", "P1"),
    ("License coverage check: all covered", "teleHealthStates=['NY'] + active NY license", "'All covered' in green", "Full coverage", "P1"),
    ("License coverage: missing state", "teleHealthStates=['NY','NJ'] but only NY license", "Warning: 'Missing: NJ' in red", "Gap detection", "P1"),
    ("No telehealth states", "Provider without teleHealthStates", "Dash or empty cell", "Empty state", "P2"),
    ("Training certification badge", "Provider with teleHealthTrainingDate set", "Green 'Certified' badge", "Training status", "P1"),
    ("No training certification", "Provider without training date", "No badge or 'Not certified'", "Missing training", "P1"),
]:
    r = tc_add(ws, r, "Telehealth", "Telehealth", "Functional", desc, inp, exp, why, pri)

###################################################################
# 21. SCORECARDS & ANALYTICS
###################################################################
r = mod_hdr(ws, r, "21. SCORECARDS & ANALYTICS")
r = sub_hdr(ws, r, "Scorecards")
for desc, inp, exp, why, pri in [
    ("/scorecards renders with scoring table", "Navigate /scorecards", "Table: Provider, PSV(20%), Sanctions(20%), Expirables(15%), Recred(15%), Docs(10%), NPDB(10%), CME(10%), Overall", "Performance scoring", "P0"),
    ("Overall = weighted average of 7 categories", "Known provider data", "Overall score = manual calculation of weighted formula", "Score accuracy", "P1"),
    ("PSV score: 100% if all verified", "Provider with all verification records verified", "PSV column: 100 in green", "PSV scoring", "P1"),
    ("Sanctions score: 0% if flagged", "Provider with unacknowledged FLAGGED sanction", "Sanctions column: 0 in red", "Sanctions scoring", "P1"),
    ("Expirables score based on renewal status", "Provider with 1 expired, 1 current", "Expirables column shows partial %", "Expirables scoring", "P1"),
    ("NPDB column shows score", "Provider with CLEAR NPDB", "NPDB: 100%", "NPDB scoring", "P1"),
    ("CME column shows progress", "Provider with 40/50 credits", "CME: 80%", "CME scoring", "P1"),
    ("Color coding: green >90, yellow 50-90, red <50", "Various providers", "Colors match thresholds", "Visual indicators", "P1"),
    ("Search scorecards by provider", "Search: 'Smith'", "Filtered to matching providers", "Search", "P1"),
    ("Pagination on scorecards", "50+ providers", "Pagination works correctly", "Large dataset", "P2"),
]:
    r = tc_add(ws, r, "Scorecards", "Scorecards", "Functional", desc, inp, exp, why, pri)

r = sub_hdr(ws, r, "Analytics")
for desc, inp, exp, why, pri in [
    ("/analytics renders with all sections", "Navigate /analytics", "Pipeline chart, enrollment status, bot status, expirables summary, recred summary", "Analytics hub", "P1"),
    ("Provider pipeline distribution", "Check pipeline chart", "Status counts match DB: SELECT status, count(*) FROM providers GROUP BY status", "Data accuracy", "P1"),
    ("Enrollment status distribution", "Check enrollment section", "Status counts match DB enrollment aggregation", "Data accuracy", "P1"),
    ("Bot run status distribution", "Check bot section", "Completed/failed/running counts match", "Data accuracy", "P1"),
    ("Expirables summary section", "Verify presence", "Status breakdown: CURRENT, EXPIRING_SOON, EXPIRED with color badges", "Expirables analytics", "P1"),
    ("Recredentialing summary section", "Verify presence", "Status breakdown: PENDING, IN_PROGRESS, OVERDUE, COMPLETED", "Recred analytics", "P1"),
]:
    r = tc_add(ws, r, "Analytics", "Analytics", "Functional", desc, inp, exp, why, pri)

###################################################################
# 22. MAIN DASHBOARD INTEGRATION
###################################################################
r = mod_hdr(ws, r, "22. MAIN DASHBOARD INTEGRATION")
for desc, inp, exp, why, pri in [
    ("Dashboard renders with stat cards", "/dashboard", "6 cards: Total, Onboarding, Verification, Committee Ready, Approved, Expiring 30d", "Dashboard overview", "P0"),
    ("Stat card counts match DB", "Cross-check with database", "Each count matches respective DB query", "Data integrity", "P0"),
    ("Pipeline table shows non-terminal providers", "View table", "Excludes APPROVED/DENIED/INACTIVE; shows name, type, status, docs, specialist", "Active pipeline", "P0"),
    ("Task list shows current user tasks", "Login as specialist with assigned tasks", "Only own tasks shown; priority, due date visible", "Personal tasks", "P0"),
    ("Cross-Module Alerts: overdue recred", "Create overdue recred cycle", "'N overdue recredentialing cycles' with link to /recredentialing", "Cross-module", "P1"),
    ("Cross-Module Alerts: evaluations due", "Create eval due within 30d", "'N evaluations due within 30 days' with link", "Cross-module", "P1"),
    ("Cross-Module Alerts: pending WH verifications", "Create pending WH request", "'N pending work history verifications' with link", "Cross-module", "P1"),
    ("Cross-Module Alerts: pending references", "Create pending reference", "'N pending professional references' with link", "Cross-module", "P1"),
    ("Upcoming Expirations clickable → /expirables", "Click 'Upcoming Expirations →'", "Navigate to /expirables", "Quick nav", "P1"),
    ("Overdue Follow-ups clickable → /enrollments", "Click 'Overdue Follow-ups →'", "Navigate to /enrollments?status=overdue", "Quick nav", "P1"),
    ("Status distribution bar", "View bar", "Colored segments proportional to status counts; hover shows numbers", "Visual overview", "P1"),
]:
    r = tc_add(ws, r, "Dashboard", "Main Dashboard", "Functional", desc, inp, exp, why, pri)

###################################################################
# 23. ADMIN
###################################################################
r = mod_hdr(ws, r, "23. ADMIN MANAGEMENT")
r = sub_hdr(ws, r, "Users")
for desc, inp, exp, why, pri in [
    ("User list page", "/admin/users", "Table: name, email, role, active status", "User management", "P0"),
    ("Create user (admin only)", "admin.createUser(email, displayName, role:SPECIALIST)", "User created isActive=true role=SPECIALIST", "Provisioning", "P0"),
    ("Update user role", "admin.updateUser(id, role:MANAGER)", "Role updated", "Role change", "P0"),
    ("Deactivate user", "admin.deactivateUser(id)", "isActive=false; user can't login", "Deprovisioning", "P0"),
    ("Delete user (admin)", "admin.deleteUser(id)", "User soft-deleted", "Removal", "P1"),
    ("User detail page", "/admin/users/[id]", "User info, role, status, actions", "Detail view", "P1"),
    ("Search users", "Search by name or email", "Filtered results", "Search", "P1"),
    ("Filter by role", "Select MANAGER", "Only managers shown", "Role filter", "P1"),
]:
    r = tc_add(ws, r, "Admin", "Users", "Functional", desc, inp, exp, why, pri)

r = sub_hdr(ws, r, "Provider Types & Settings")
for desc, inp, exp, why, pri in [
    ("Provider types list", "/admin/provider-types", "Types with abbreviation, description, requirement count", "Type management", "P1"),
    ("Create provider type", "admin.createProviderType(name, abbreviation)", "Type created; available in provider creation", "Configuration", "P1"),
    ("Set document requirement", "admin.setDocumentRequirement(typeId, docType, isRequired)", "Requirement associated with type; affects checklist generation", "Checklist config", "P1"),
    ("Settings list", "/admin/settings", "Key-value settings with category", "App settings", "P1"),
    ("Upsert setting", "admin.upsertSetting(key, value, category)", "Setting created/updated", "Config management", "P1"),
    ("Delete setting", "admin.deleteSetting(key)", "Setting removed", "Cleanup", "P2"),
]:
    r = tc_add(ws, r, "Admin", "Config", "Functional", desc, inp, exp, why, pri)

r = sub_hdr(ws, r, "Workflows & Training")
for desc, inp, exp, why, pri in [
    ("Workflow list", "/admin/workflows", "Workflows with name, description", "Process docs", "P2"),
    ("Workflow editor (Excalidraw)", "/admin/workflows/[id]", "Excalidraw canvas loads; save persists", "Visual editing", "P2"),
    ("Training list (admin)", "training.listAll()", "All training records", "LMS", "P2"),
    ("Create training", "training.create(userId, courseName, ...)", "Record created", "Training tracking", "P2"),
    ("Training compliance summary", "training.getComplianceSummary()", "Summary of staff training completion", "Compliance", "P2"),
]:
    r = tc_add(ws, r, "Admin", "Workflows/Training", "Functional", desc, inp, exp, why, pri)

###################################################################
# 24. BULK IMPORT
###################################################################
r = mod_hdr(ws, r, "24. BULK IMPORT / EXPORT")
for desc, inp, exp, why, pri in [
    ("BulkImportModal opens with type options", "Trigger import action", "Modal: file upload, type selector (providers/enrollments/expirables)", "Bulk import", "P1"),
    ("Parse valid provider CSV", "Upload CSV: firstName,lastName,providerType,npi", "Preview table shows rows; validation passes; Import enabled", "CSV parsing", "P1"),
    ("Import providers from CSV", "Click Import with valid data", "Providers created via tRPC; result: N imported, M skipped", "Bulk create", "P1"),
    ("Import enrollments from CSV", "Valid enrollment CSV", "Enrollments created", "Bulk create", "P1"),
    ("Import expirables from CSV", "Valid expirable CSV", "Expirables created", "Bulk create", "P1"),
    ("Reject CSV with missing headers", "CSV without 'firstName' column", "Validation error listing missing fields", "Data quality", "P1"),
    ("Reject CSV with invalid data", "Row with invalid NPI (3 digits)", "Row-level error shown; import blocked or skips row", "Validation", "P1"),
    ("Empty CSV file", "Upload empty CSV", "Error: no data rows found", "Edge case", "P2"),
]:
    r = tc_add(ws, r, "Bulk Import", "Bulk Operations", "Functional", desc, inp, exp, why, pri)

###################################################################
# 25. UI/UX CONSISTENCY
###################################################################
r = mod_hdr(ws, r, "25. UI/UX CONSISTENCY & DESIGN QUALITY")
r = sub_hdr(ws, r, "Layout & Navigation")
for desc, inp, exp, why, pri in [
    ("Sidebar: all 16+ nav items visible", "Login as admin", "Dashboard, Providers, Committee, Enrollments, Medicaid, Expirables, Recredentialing, Verifications, Evaluations, Rosters, CME, Telehealth, Reports, Compliance, Scorecards, Analytics + Admin", "Complete nav", "P0"),
    ("Sidebar: active route highlighted", "Navigate /enrollments", "Enrollments item visually selected", "Route context", "P1"),
    ("Sidebar: admin section for ADMIN/MANAGER only", "Login as SPECIALIST vs ADMIN", "SPECIALIST: no admin section; ADMIN: admin section visible", "Role-based nav", "P0"),
    ("Page titles match module names", "Visit each page", "H1 heading matches sidebar label (Dashboard, NCQA Compliance Readiness, etc.)", "Identity", "P1"),
    ("Responsive at 1280px", "Resize to 1280px width", "All pages functional; tables scrollable; sidebar usable", "Responsive", "P2"),
    ("Responsive at 1024px", "Resize to 1024px", "Content readable; no overlap; may need horizontal scroll on tables", "Tablet compat", "P2"),
    ("No horizontal scroll at 1440px+", "Standard desktop view", "No unnecessary horizontal scrollbar", "Layout quality", "P1"),
]:
    r = tc_add(ws, r, "UI/UX", "Layout", "UI", desc, inp, exp, why, pri)

r = sub_hdr(ws, r, "Visual Consistency")
for desc, inp, exp, why, pri in [
    ("Badge colors consistent: APPROVED=green", "Compare APPROVED badges across /providers, /enrollments, /expirables, /recredentialing", "Same green shade on all", "Color system", "P1"),
    ("Badge colors consistent: DENIED/EXPIRED=red", "Compare red badges", "Same red shade consistently", "Color system", "P1"),
    ("Badge colors consistent: PENDING=yellow/amber", "Compare pending badges", "Same yellow/amber consistently", "Color system", "P1"),
    ("Badge colors consistent: IN_PROGRESS=blue", "Compare in-progress badges", "Same blue consistently", "Color system", "P1"),
    ("Table header style consistency", "Compare tables across 10 pages", "Same font weight, background, text color on all headers", "Table design", "P1"),
    ("Table row hover consistent", "Hover rows on different pages", "Same hover background color on all", "Interaction consistency", "P1"),
    ("Card design consistency", "Compare stat cards on dashboard, compliance, enrollments, expirables", "Same border, padding, font hierarchy", "Card system", "P1"),
    ("Button styling consistent", "Compare primary actions across pages", "Same shadcn/ui button variant styles", "Button system", "P1"),
    ("Modal design consistency", "Open AddProviderModal, AddEnrollmentModal, BulkImportModal", "Same dialog wrapper, title style, button placement", "Modal system", "P1"),
    ("Empty state messages", "View pages with no data", "Helpful message: 'No {items} found' with guidance", "Empty states", "P1"),
    ("Loading states (client components)", "Trigger data fetch", "Spinner or skeleton during load", "Loading UX", "P2"),
]:
    r = tc_add(ws, r, "UI/UX", "Visual Consistency", "UI", desc, inp, exp, why, pri)

r = sub_hdr(ws, r, "Forms & Interactions")
for desc, inp, exp, why, pri in [
    ("Validation errors inline near fields", "Submit form missing required field", "Red border on field; error text directly below", "Error proximity", "P1"),
    ("Toast on success actions", "Create provider, approve in committee", "Success toast with confirmation message", "Action feedback", "P1"),
    ("Toast on error actions", "Trigger an error (invalid operation)", "Error toast with descriptive message", "Error feedback", "P1"),
    ("Modal close: X button", "Open modal, click X", "Modal closes without submitting", "Close behavior", "P1"),
    ("Modal close: Escape key", "Open modal, press Escape", "Modal closes", "Keyboard close", "P1"),
    ("Modal close: backdrop click", "Click outside modal", "Modal closes", "Backdrop close", "P1"),
    ("Date format consistency: MMM DD, YYYY", "Check dates across pages", "All dates in consistent format (not raw ISO)", "Date formatting", "P1"),
    ("Relative dates (e.g., '3 days ago')", "Check recent activity timestamps", "Relative format for recent; full date for older", "Time display", "P2"),
    ("Tab URL state (provider detail)", "/providers/[id]?tab=cme", "CME tab selected; content shown", "Deep linking", "P1"),
    ("Browser back/forward with tabs", "Navigate tabs, then back", "Previous tab restores correctly", "History integration", "P1"),
    ("Form autofill works", "Browser autofill on login form", "Email/password fields populate correctly", "Autofill compat", "P2"),
    ("Double-submit prevention", "Rapidly click Submit twice", "Only 1 record created; button disabled during submission", "Idempotency", "P1"),
]:
    r = tc_add(ws, r, "UI/UX", "Forms/Interactions", "UI", desc, inp, exp, why, pri)

###################################################################
# 26. SECURITY
###################################################################
r = mod_hdr(ws, r, "26. SECURITY TESTING")
r = sub_hdr(ws, r, "Data Protection")
for desc, inp, exp, why, pri in [
    ("SSN encrypted at rest (AES-256-GCM)", "Update SSN via provider.update; query DB directly", "Column contains base64 encoded ciphertext, not plaintext XXX-XX-XXXX", "HIPAA PHI", "P0"),
    ("DOB encrypted at rest", "Update DOB; query DB", "Encrypted value in DB", "PHI protection", "P0"),
    ("Decryption returns original value", "Update SSN='123-45-6789'; read via tRPC", "Returned SSN = '123-45-6789'", "Encryption roundtrip", "P0"),
    ("API key stored as SHA-256 hash", "Create key; inspect DB apiKey.keyHash", "Hash value (64 hex chars), not 'essen_...'", "Key security", "P0"),
    ("Password stored as bcrypt hash", "Register user; check DB passwordHash", "bcrypt hash ($2b$...)", "Credential security", "P0"),
    ("Encryption key validation", "Set ENCRYPTION_KEY to invalid value", "App startup error: 'ENCRYPTION_KEY must be 32 bytes'", "Key validation", "P1"),
    ("PHI not in API responses", "GET /api/v1/providers", "No SSN, DOB, or other PHI in REST API response", "API data filtering", "P0"),
]:
    r = tc_add(ws, r, "Security", "Data Protection", "Security", desc, inp, exp, why, pri)

r = sub_hdr(ws, r, "Authorization Boundaries")
for desc, inp, exp, why, pri in [
    ("tRPC UNAUTHORIZED for no session", "Call staffProcedure without auth", "TRPCError code: UNAUTHORIZED", "Auth enforcement", "P0"),
    ("tRPC FORBIDDEN for wrong role", "SPECIALIST calls adminProcedure", "TRPCError code: FORBIDDEN with role list", "Role enforcement", "P0"),
    ("REST API 401 without key", "GET /api/v1/providers no header", "401 JSON", "API auth", "P0"),
    ("REST API 403 with revoked key", "Use revoked key", "403 JSON", "Revoked keys", "P0"),
    ("REST API 403 wrong permission", "Key without providers:read → GET /api/v1/providers", "403 insufficient permissions", "Permission check", "P0"),
    ("FHIR 403 without fhir:read", "Key without fhir:read → GET /api/fhir/Practitioner", "403 OperationOutcome", "FHIR permission", "P0"),
    ("Upload requires auth", "POST /api/upload without session", "401", "Upload auth", "P0"),
    ("IDOR: Upload doesn't check providerId ownership", "Authenticated user uploads to another provider", "KNOWN GAP: Upload accepted — document finding", "Authorization gap", "P0"),
    ("Webhook lacks signature verification", "POST /api/webhooks/sendgrid with spoofed data", "KNOWN GAP: No SendGrid signature check — document", "Webhook security", "P1"),
]:
    r = tc_add(ws, r, "Security", "Authorization", "Security", desc, inp, exp, why, pri)

r = sub_hdr(ws, r, "Input Security")
for desc, inp, exp, why, pri in [
    ("SQL injection via search field", "Search: '; DROP TABLE providers; --", "Prisma parameterized query prevents injection; normal 0-result response", "ORM protection", "P0"),
    ("SQL injection via tRPC input", "provider.list(search: \"' OR 1=1 --\")", "No injection; Prisma parameterization", "ORM protection", "P0"),
    ("XSS in provider notes", "Enter: <script>alert(1)</script> in notes", "Stored as text; React auto-escapes; no script execution on render", "Output encoding", "P1"),
    ("XSS in communication body", "Enter: <img onerror=alert(1) src=x> in email body", "React auto-escapes or sanitizes", "XSS prevention", "P1"),
    ("Path traversal in upload filename", "Filename: ../../etc/passwd", "Filename sanitized; no directory traversal", "Path safety", "P1"),
    ("Oversized JSON body", "POST with 50MB JSON body", "413 or server rejects; no crash", "Input limits", "P2"),
]:
    r = tc_add(ws, r, "Security", "Input Security", "Security", desc, inp, exp, why, pri)

r = sub_hdr(ws, r, "Known Security Gaps (Document)")
for desc, inp, exp, why, pri in [
    ("No rate limiting on login", "100 rapid login attempts", "All processed (no 429); DOCUMENT for remediation", "Brute force gap", "P1"),
    ("No rate limiting on API", "100 rapid API calls", "All processed; DOCUMENT", "API abuse gap", "P1"),
    ("/verify/* routes blocked by middleware", "Unauthenticated → /verify/work-history/[token]", "Redirect to signin; DOCUMENT: should be public", "Public form gap", "P0"),
    ("/api/application/save-section middleware", "POST without session cookie", "DOCUMENT behavior: investigate if blocked", "Provider form gap", "P0"),
    ("No CAPTCHA on registration", "Automated registration attempts", "All succeed; DOCUMENT", "Bot registration gap", "P2"),
]:
    r = tc_add(ws, r, "Security", "Known Gaps", "Security", desc, inp, exp, why, pri)

###################################################################
# 27. AUDIT TRAIL
###################################################################
r = mod_hdr(ws, r, "27. AUDIT TRAIL COMPLETENESS")
audit_actions = [
    ("provider.created", "Create provider", "provider"),
    ("provider.status.changed", "Transition provider status", "provider"),
    ("provider.updated", "Edit provider data", "provider"),
    ("provider.invite.sent", "Send provider invite", "provider"),
    ("committee.session.created", "Create committee session", "committeeSession"),
    ("committee.decision.recorded", "Record committee decision", "committeeSessionEntry"),
    ("enrollment.created", "Create enrollment", "enrollment"),
    ("enrollment.status.updated", "Update enrollment status", "enrollment"),
    ("sanctions.triggered", "Trigger sanctions check", "sanctionsCheck"),
    ("npdb.query.triggered", "Trigger NPDB query", "npdbRecord"),
    ("workHistory.request.sent", "Send WH request", "workHistoryVerification"),
    ("workHistory.reminder.sent", "Send WH reminder", "workHistoryVerification"),
    ("workHistory.response.submitted", "Public WH form submitted", "workHistoryVerification"),
    ("reference.request.sent", "Send reference request", "professionalReference"),
    ("reference.reminder.sent", "Send reference reminder", "professionalReference"),
    ("reference.response.submitted", "Public reference form submitted", "professionalReference"),
    ("cme.updated", "Update CME credit", "cmeCredit"),
    ("recredentialing.created", "Create recred cycle", "recredentialingCycle"),
    ("apiKey.created", "Create API key", "apiKey"),
]
for action, trigger, entity in audit_actions:
    r = tc_add(ws, r, "Audit Trail", f"Action: {action}", "Functional",
               f"Verify audit log for: {trigger}",
               f"Perform: {trigger}; query audit_logs WHERE action='{action}'",
               f"Row: actorId (current user), entityType='{entity}', entityId, timestamp, providerId if applicable",
               "NCQA audit trail requirement", "P1")

r = tc_add(ws, r, "Audit Trail", "Audit Immutability", "Security", "No endpoints to delete/modify audit logs", "Search codebase for auditLog.delete or auditLog.update", "No mutation endpoints for audit_logs", "Audit integrity", "P0")
r = tc_add(ws, r, "Audit Trail", "Audit Display", "UI", "Audit trail renders on provider detail", "Provider → Audit Trail tab", "Timeline with actor name, action, timestamp, entity; pagination", "Audit visibility", "P0")

###################################################################
# 28. WORKER JOBS
###################################################################
r = mod_hdr(ws, r, "28. SCHEDULED WORKER JOBS")
jobs = [
    ("expirables-scan", "24h", "Updates statuses (CURRENT→EXPIRING_SOON→EXPIRED); creates tasks; sends emails", "P1"),
    ("sanctions-weekly", "7d", "OIG+SAM checks for all approved providers", "P1"),
    ("sanctions-monthly", "30d", "Additional monthly sanctions batch run", "P1"),
    ("license-poll", "24h", "Re-verifies licenses expiring within 90 days", "P1"),
    ("recredentialing-check", "24h", "Marks overdue cycles; creates new cycles from initialApprovalDate", "P1"),
    ("follow-up-cadence", "1h", "Creates tasks for overdue enrollment follow-ups", "P1"),
]
for job, sched, desc, pri in jobs:
    r = tc_add(ws, r, "Workers", f"Job: {job}", "Functional",
               f"Verify {job} executes correctly",
               f"Schedule: every {sched}; trigger via Bull Board or manual add",
               desc, "Background automation", pri)
    r = tc_add(ws, r, "Workers", f"Job: {job}", "Negative",
               f"Verify {job} handles empty dataset",
               "Run with no qualifying records",
               "Job completes without error; 0 records processed",
               "Graceful empty handling", "P2")

r = tc_add(ws, r, "Workers", "Queue Health", "Smoke", "Bull Board shows all 3 queues", "GET :6025/bull-board", "psv-bot, enrollment-bot, scheduled-jobs queues visible with stats", "Queue monitoring", "P0")

###################################################################
# 29. CROSS-MODULE INTEGRATION
###################################################################
r = mod_hdr(ws, r, "29. CROSS-MODULE INTEGRATION FLOWS")
for desc, inp, exp, why, pri in [
    ("E2E: Create → Invite → Application → Documents → PSV → Committee → Approve", "Full provider lifecycle", "Provider moves INVITED → ONBOARDING → DOCS_PENDING → PSV_IN_PROGRESS → PSV_COMPLETE → COMMITTEE_READY → APPROVED; recred cycle created", "Core happy path end-to-end", "P0"),
    ("Approval creates recredentialing cycle", "Approve provider in committee", "RecredentialingCycle auto-created: cycleNumber=1, dueDate=approval+36mo", "Committee → Recred integration", "P0"),
    ("Approval sets initialApprovalDate", "First-time approval", "provider.initialApprovalDate = committee decision date", "Approval tracking", "P0"),
    ("Committee queue has both initial + recred providers", "Mix of COMMITTEE_READY and recred cycles", "Queue returns { initialCredentialing: [...], recredentialing: [...] }", "Queue categorization", "P1"),
    ("Overdue recred → dashboard alert", "Recred cycle with past dueDate", "Dashboard Cross-Module Alerts shows count with link", "Dashboard integration", "P1"),
    ("Expirable scan → provider task", "Expirable within 30d", "Task created for assigned specialist or manager", "Expirable → Task", "P1"),
    ("Enrollment follow-up cadence → task", "Enrollment with overdue follow-up", "follow-up-cadence job creates task", "Enrollment → Task", "P1"),
    ("Sanctions flagged → compliance decrease", "Flag sanction for provider", "Compliance page shows decreased sanctions score", "Sanctions → Compliance", "P1"),
    ("PSV verified → compliance increase", "Verify license for provider", "Compliance PSV card numerator increases", "PSV → Compliance", "P1"),
    ("CME credits → scorecard", "Add 50 CME credits to provider", "Scorecard CME column = 100%", "CME → Scorecard", "P1"),
    ("NPDB clear → scorecard", "NPDB query returns CLEAR", "Scorecard NPDB = 100%", "NPDB → Scorecard", "P1"),
    ("Telehealth states vs active licenses", "Set teleHealthStates then create/remove licenses", "Telehealth page updates coverage display", "Telehealth → License cross-ref", "P1"),
    ("Application form saves to multiple tables", "Complete all 10 sections", "Provider, ProviderProfile, License records created/updated", "Form → DB mapping", "P0"),
    ("Attestation → status DOCUMENTS_PENDING", "Submit attestation (section 9)", "provider.status changes; applicationSubmittedAt set", "Form → Status", "P0"),
    ("Provider detail includes ALL module data", "provider.getById includes 15+ relations", "All tabs have data from respective module tables", "Data aggregation", "P0"),
    ("Bulk import creates valid providers", "Import CSV → providers appear in list with proper status", "Created as INVITED; checklist generated per type", "Import → Provider lifecycle", "P1"),
    ("Export includes cross-module data", "Export providers CSV", "CSV includes status, type, specialist — all from different tables", "Export completeness", "P1"),
    ("Roster generation from enrollments", "Generate roster from selected enrollments", "CSV contains provider data from enrollment + provider tables", "Roster → Enrollment", "P1"),
    ("Webhook updates communication status", "SendGrid webhook fires for email event", "Communication record updated with delivery status", "Webhook → Comms", "P2"),
]:
    r = tc_add(ws, r, "Integration", "Cross-Module", "Integration", desc, inp, exp, why, pri)

###################################################################
# 30. DATA INTEGRITY
###################################################################
r = mod_hdr(ws, r, "30. DATA INTEGRITY & EDGE CASES")
for desc, inp, exp, why, pri in [
    ("UUID format for all IDs", "Inspect any created record ID", "Valid UUID v4 format", "ID consistency", "P1"),
    ("Timestamps in UTC", "Create record; check DB timestamp", "stored in UTC; displayed in local timezone", "Timezone handling", "P1"),
    ("Soft delete doesn't remove data", "Delete provider (set INACTIVE); query DB", "Record exists with INACTIVE status; not physically deleted", "Soft delete pattern", "P1"),
    ("Foreign key integrity: enrollment → provider", "Delete provider with enrollments", "Provider set INACTIVE; enrollments still reference provider", "FK integrity", "P1"),
    ("Cascade: delete provider with tasks", "Delete provider; check tasks", "Tasks still exist (orphaned) or cleaned up properly", "Cascade behavior", "P2"),
    ("Concurrent provider creation (same NPI)", "Simultaneously create 2 providers with same NPI", "Either 1 succeeds + 1 fails, or unique constraint violation", "Concurrency", "P2"),
    ("Large text in notes field", "Enter 10,000 character note", "Saved and displayed correctly (possibly with scroll)", "Large text handling", "P2"),
    ("Special characters in provider name", "Name: O'Brien-Smith Jr.", "Saved and displayed correctly with apostrophe/hyphen", "Special chars", "P1"),
    ("Unicode in text fields", "Enter: Ñoño Müller-Schröder", "Saved and displayed correctly", "Unicode support", "P1"),
    ("Empty optional fields", "Create provider with all optional fields blank", "Provider created; empty fields show dash or blank (not 'null')", "Null display", "P1"),
    ("Provider type display consistency", "Check type abbreviation across all pages", "Same abbreviation (MD, DO, PA, NP) used everywhere", "Display consistency", "P1"),
    ("Date edge: leap year", "Set expiration to Feb 29, 2028", "Date saved and displayed correctly", "Leap year", "P2"),
    ("Date edge: year boundary", "Expirable expiring Dec 31 → Jan 1 urgency", "Correct urgency calculation across year boundary", "Year boundary", "P2"),
    ("Pagination total count accuracy", "50 providers; page 1 of 5 (limit=10)", "Total shows 50; 5 pages; each page has 10 (last may have fewer)", "Pagination math", "P1"),
]:
    r = tc_add(ws, r, "Data Integrity", "Edge Cases", "Functional", desc, inp, exp, why, pri)


# ═════════════════════════════════════════════════════════════════════
# TEST STRATEGY SHEET
# ═════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Test Strategy")
strat = [
    ["ESSEN CREDENTIALING PLATFORM — COMPREHENSIVE TEST STRATEGY"],
    [""],
    ["Section", "Detail"],
    ["1. Objective", "Verify the ESSEN Credentialing Platform is fully functional, resilient, secure, consistent, and user-friendly across all 20 modules. Cover the entire provider lifecycle from onboarding through recredentialing, including committee workflows, enrollment management, PSV automation, compliance reporting, public APIs, analytics, and cross-module integration."],
    ["2. Scope", f"All 20 functional modules, 42+ page routes, 22 tRPC routers (150+ procedures), 14 REST/API endpoints, 6 scheduled worker jobs, 10+ bot types, UI/UX consistency, security controls, RBAC enforcement, audit trail integrity, data protection, and cross-module integration flows. Total: {tc[0]} test cases."],
    ["3. Test Approach", "Manual functional testing using checklist methodology. Each test case specifies: unique ID, module, sub-module, test type, description, input/precondition, expected outcome, rationale, priority, and placeholder for actuals. Tests are sequenced by dependency: Infrastructure → Auth → CRUD per module → Workflows → Cross-module Integration → Security → Data Integrity."],
    ["4. Test Environment", "Local Docker stack: web container (:6015), worker container (:6025), PostgreSQL (:5433 → :5432 internal), Redis (:6379). Docker Compose dev config. Seed data required (provider types + admin user). Azure AD SSO disabled in dev; use email/password credentials."],
    ["5. Roles Under Test", "ADMIN: full system access including user management, settings, API keys, privileging. MANAGER: admin pages + staff + committee decisions + manager-only mutations. SPECIALIST: staff pages only — providers, enrollments, expirables, tasks. COMMITTEE_MEMBER: committee pages + staff. PROVIDER: application/attestation pages (via magic link). Anonymous: public verification forms, REST API (via key), FHIR."],
    ["6. Priority Framework", "P0 = BLOCKING — Must pass for any release. System-breaking if failed. P1 = HIGH — Core functionality that directly impacts credentialing workflows. P2 = MEDIUM — Enhancement, polish, or edge case. P3 = LOW — Nice-to-have."],
    ["7. Entry Criteria", "1) Docker containers healthy and responding. 2) Database migrated and seeded (≥6 provider types, ≥1 admin user). 3) At least 1 test user per role (ADMIN, MANAGER, SPECIALIST, COMMITTEE_MEMBER). 4) Test providers in varied statuses (INVITED, ONBOARDING, PSV_COMPLETE, COMMITTEE_READY, APPROVED). 5) Sample enrollments, expirables, recred cycles, evaluations, CME credits. 6) Bull Board accessible at :6025."],
    ["8. Exit Criteria", "1) ALL P0 tests PASS. 2) ≥95% P1 tests PASS. 3) All FAILures documented with severity, actual outcome, and root cause. 4) No critical or high security findings unresolved. 5) Cross-module integration happy path verified end-to-end. 6) Test results reviewed by QA lead."],
    ["9. Test Types", "SMOKE: Infrastructure readiness (containers, DB, health). FUNCTIONAL: Feature correctness, CRUD operations, business logic. VALIDATION: Input format enforcement, required fields, data type checks. NEGATIVE: Error handling, boundary conditions, invalid inputs. UI: Visual rendering, layout, consistency, responsiveness. SECURITY: Auth/auth, encryption, injection, IDOR, rate limiting. INTEGRATION: Cross-module data flow, end-to-end workflows."],
    ["10. Known Limitations", "1) Azure AD SSO disabled locally (test with credentials provider). 2) Azure Blob Storage not configured (uploads use local/mock). 3) SendGrid/ACS not configured (emails logged, not sent). 4) Some PSV bots require external site access (NCCPA, DEA, etc.). 5) EDUCATION_AMA and EDUCATION_ECFMG bots are placeholders. 6) /verify/* routes may be blocked by middleware (known gap)."],
    ["11. Test Data Strategy", "Layer 1 (Seed): Provider types, admin user. Layer 2 (Manual): Create 5-10 test providers spanning all statuses. Layer 3 (Module Data): Create enrollments, expirables, recred cycles, evaluations, CME credits, WH verifications, references, sanctions checks, NPDB queries. Layer 4 (Edge Cases): Special characters, unicode, large text, boundary dates."],
    ["12. Defect Classification", "CRITICAL: System crash, data loss, security breach, complete feature failure. MAJOR: Core workflow broken, incorrect data, RBAC bypass. MINOR: UI glitch, formatting issue, non-blocking UX problem. COSMETIC: Visual polish, label text, spacing."],
    ["13. Test Execution Pattern", "Phase 1: Smoke tests (infra health). Phase 2: Auth & RBAC (foundation). Phase 3: Module-by-module functional testing (following sidebar order). Phase 4: Cross-module integration flows. Phase 5: Security testing. Phase 6: UI/UX consistency audit. Phase 7: Data integrity & edge cases. Phase 8: Regression pass on any fixes."],
    ["14. Result Recording", "For each test case: Mark PASS (P), FAIL (F), or BLOCKED (B) in Result column. For FAIL/BLOCKED: describe actual behavior in Actual Outcome column. Include screenshots for UI defects. Note Tester name and Date. Blocked = cannot test due to dependency failure."],
]
for i, row_data in enumerate(strat, 1):
    for j, val in enumerate(row_data):
        cell = ws2.cell(row=i, column=j + 1, value=val)
        if i == 1:
            cell.font = Font(bold=True, size=14, color="1F4E79")
        elif i == 3:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        elif j == 0 and i > 3:
            cell.font = Font(bold=True, color="1F4E79")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 130

# ═════════════════════════════════════════════════════════════════════
# SUMMARY SHEET
# ═════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Summary")
ws3["A1"] = "ESSEN CREDENTIALING — TEST PLAN SUMMARY"
ws3["A1"].font = Font(bold=True, size=14, color="1F4E79")
ws3["A3"] = "Total Test Cases:"
ws3["B3"] = tc[0]
ws3["B3"].font = Font(bold=True, size=16, color="1F4E79")
ws3["A4"] = "Generated:"
ws3["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
ws3["A5"] = "Platform Version:"
ws3["B5"] = "v1.0 — 20 Modules (Post-Integration)"

ws3["A7"] = "PRIORITY BREAKDOWN"
ws3["A7"].font = Font(bold=True, size=12, color="1F4E79")
r3 = 8
for pri in ["P0", "P1", "P2", "P3"]:
    ws3[f"A{r3}"] = pri
    ws3[f"A{r3}"].font = Font(bold=True)
    ws3[f"A{r3}"].border = THIN_BORDER
    ws3[f"B{r3}"] = priority_counts.get(pri, 0)
    ws3[f"B{r3}"].border = THIN_BORDER
    r3 += 1

ws3[f"A{r3}"] = "TOTAL"
ws3[f"A{r3}"].font = Font(bold=True, color="1F4E79")
ws3[f"A{r3}"].border = THIN_BORDER
ws3[f"B{r3}"] = tc[0]
ws3[f"B{r3}"].font = Font(bold=True, color="1F4E79")
ws3[f"B{r3}"].border = THIN_BORDER

r3 += 2
ws3[f"A{r3}"] = "MODULE BREAKDOWN"
ws3[f"A{r3}"].font = Font(bold=True, size=12, color="1F4E79")
r3 += 1
ws3[f"A{r3}"] = "Module"
ws3[f"A{r3}"].font = HEADER_FONT
ws3[f"A{r3}"].fill = HEADER_FILL
ws3[f"B{r3}"] = "Test Cases"
ws3[f"B{r3}"].font = HEADER_FONT
ws3[f"B{r3}"].fill = HEADER_FILL
r3 += 1
for mod, count in sorted(module_counts.items(), key=lambda x: -x[1]):
    ws3[f"A{r3}"] = mod
    ws3[f"A{r3}"].border = THIN_BORDER
    ws3[f"B{r3}"] = count
    ws3[f"B{r3}"].border = THIN_BORDER
    r3 += 1

ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 20

# ═════════════════════════════════════════════════════════════════════
# SAVE
# ═════════════════════════════════════════════════════════════════════
now = datetime.now().strftime("%Y%m%d_%H%M%S")
filename = f"ESSEN_Credentialing_Test_Plan_{now}.xlsx"
filepath = f"c:\\Users\\admin\\development\\HDPulseAI\\EssenApps\\E_Credentialing\\docs\\testing\\{filename}"
wb.save(filepath)
print(f"SAVED: {filepath}")
print(f"Total test cases: {tc[0]}")
print(f"Priority breakdown: {priority_counts}")
print(f"Module breakdown: {module_counts}")
